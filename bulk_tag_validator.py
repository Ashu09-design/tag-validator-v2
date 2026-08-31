import warnings
warnings.filterwarnings("ignore")
import asyncio
import pandas as pd
from playwright.async_api import async_playwright
# playwright-stealth changed its public API between 1.x and 2.x. 1.x exposes a
# module-level `stealth_async(page)`; 2.x exposes a `Stealth` class with
# `apply_stealth_async(page)`. Pin-free support for both, and degrade to a
# no-op if the package is missing entirely, so an environment mismatch can
# never take the whole validator down at import time.
try:
    from playwright_stealth import Stealth as _StealthCls
except ImportError:
    _StealthCls = None
try:
    from playwright_stealth import stealth_async as _stealth_async_fn
except ImportError:
    _stealth_async_fn = None
import os
import time
import re
import sys
import json
import argparse
import datetime
from urllib.parse import unquote, parse_qs, urlparse

# Force UTF-8 on stdout/stderr so log lines with Unicode (->, arrows, em-dashes,
# accented site names) don't crash on Windows' default cp1252 console.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

class _StealthCompat:
    """Uniform `apply_stealth_async(page)` across playwright-stealth versions.

    IMPORTANT: playwright-stealth 1.x is deliberately NOT used. Its injected
    evasion scripts reference helpers that do not exist in the page context
    ("utils is not defined"), which throws on every page and takes the site's
    tag manager down with it — GTM stops initialising, dataLayer freezes and
    no analytics tag ever fires. A click audit run under 1.x therefore reports
    "no tracking" for a page that is in fact fully tracked. Silently wrong
    output is far worse than skipping bot-evasion, so 1.x is skipped unless
    the operator explicitly forces it with TV_FORCE_STEALTH=1.
    """

    def __init__(self):
        self._impl = None
        self._legacy = None
        self._warned = False
        if _StealthCls is not None:
            try:
                self._impl = _StealthCls()          # 2.x — safe
            except Exception:
                self._impl = None
        if self._impl is None and _stealth_async_fn is not None:
            if os.environ.get("TV_FORCE_STEALTH") == "1":
                self._legacy = _stealth_async_fn

    async def apply_stealth_async(self, page):
        try:
            if self._impl is not None:
                await self._impl.apply_stealth_async(page)
            elif self._legacy is not None:
                await self._legacy(page)
            elif _stealth_async_fn is not None and not self._warned:
                self._warned = True
                # Deliberately on stdout, not stderr: nothing has gone wrong.
                # The run is fine — better than fine, since 1.x is what used to
                # break it. Emitting this on stderr made the UI label a routine
                # note as "ERROR" and sent people looking for a fault.
                sys.stdout.write(
                    "[NOTE] Bot-evasion (playwright-stealth 1.x) is not being used: "
                    "that version injects broken JS that stops GTM initialising, "
                    "which would make a fully tagged page look untracked. The audit "
                    "runs normally without it. To enable it, install "
                    "playwright-stealth>=2.0.0." + os.linesep)
                sys.stdout.flush()
        except Exception:
            # Stealth is a nice-to-have for bot detection, never a hard
            # dependency of the audit itself.
            pass


stealth_obj = _StealthCompat()
CONCURRENCY = 3

COOKIE_SELECTORS = [
    '#onetrust-accept-btn-handler',
    '#accept-recommended-btn-handler',
    'button[title="Accept All"]',
    'button[title="Accept"]',
    '#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
    '#CybotCookiebotDialogBodyButtonAccept',
    '#truste-consent-button',
    '#didomi-notice-agree-button',
    '.cc-accept', '.cc-btn.cc-allow',
    '#cookie-accept', '#accept-cookies',
    '[data-action="accept"]',
    'button[aria-label="Accept all cookies"]',
    'button[aria-label="Accept cookies"]',
    'button[aria-label="accept and close"]',
]

COOKIE_TEXT_PATTERNS = [
    "Accept All", "Accept all", "ACCEPT ALL",
    "Accept Cookies", "Accept cookies",
    "Allow All", "Allow all",
    "I Accept", "I agree",
    "Agree", "OK", "Got it",
    "Accept & Close", "Accept and close",
    "Consent", "Continue",
]

REJECT_SELECTORS = [
    '#onetrust-reject-all-handler',
    '#CybotCookiebotDialogBodyButtonDecline',
    '#CybotCookiebotDialogBodyLevelButtonLevelOptinDeclineAll',
    'button[title="Reject All"]',
    'button[title="Reject"]',
    '#truste-consent-required',
    '.cc-deny', '.cc-btn.cc-deny',
    '#cookie-reject', '#reject-cookies',
    '[data-action="reject"]',
    'button[aria-label="Reject all cookies"]',
    'button[aria-label="Reject cookies"]',
    'button[aria-label="Deny"]',
]

REJECT_TEXT_PATTERNS = [
    "Reject All", "Reject all", "REJECT ALL",
    "Reject Cookies", "Reject cookies", "Reject",
    "Decline All", "Decline all", "Decline",
    "Deny All", "Deny all", "Deny",
    "Necessary only", "Only necessary", "Use necessary cookies only",
    "Essential only", "Continue without accepting",
]

# ===== CONSENT SCENARIOS (OneTrust category model) =====
# C0001 Strictly Necessary | C0002 Performance | C0003 Functional
# C0004 Targeting | C0005 Social Media
SCENARIO_GROUPS = {
    "Accept All":  "C0001:1,C0002:1,C0003:1,C0004:1,C0005:1",
    "Reject All":  "C0001:1,C0002:0,C0003:0,C0004:0,C0005:0",
    "Performance": "C0001:1,C0002:1,C0003:0,C0004:0,C0005:0",
    "Functional":  "C0001:1,C0002:0,C0003:1,C0004:0,C0005:0",
    "Targeting":   "C0001:1,C0002:0,C0003:0,C0004:1,C0005:1",
}
# Banner action per scenario:
#   Accept All -> click Accept-All  (full consent)
#   Reject All -> click Reject-All  (no consent)
#   Performance / Functional / Targeting -> DO NOT click any button.
#     Clicking "Accept All" would grant FULL consent and destroy the
#     partial-category meaning. The injected OneTrust OptanonConsent cookie
#     (set per category below) is what governs these scenarios.
def _host_of(url):
    try:
        return urlparse(url).hostname or ""
    except Exception:
        return ""


# Well-known multi-part TLDs where the registrable domain is 3+ parts.
_MULTI_TLDS = {
    "co.uk", "co.jp", "co.kr", "co.in", "co.za", "co.nz", "co.il",
    "com.au", "com.br", "com.cn", "com.mx", "com.sg", "com.hk", "com.tw",
    "com.ar", "com.tr", "com.my", "com.ph", "com.vn", "com.co", "com.pe",
    "org.uk", "org.au", "net.au", "ac.uk", "gov.uk",
    "ne.jp", "or.jp", "ac.jp",
}


def _cookie_domain_for(url):
    """Extract the registrable domain (eTLD+1) for setting cookies.

    OneTrust reads its consent cookie from the root domain. If the page is on
    shop.example.com but we set the cookie on .shop.example.com, the CMP never
    sees it and every tag fires regardless of the chosen scenario.
    """
    h = _host_of(url)
    if not h:
        return None
    if h.startswith("www."):
        h = h[4:]
    if h.replace(".", "").isdigit() or ":" in h:
        return "." + h
    parts = h.split(".")
    if len(parts) <= 2:
        return "." + h
    for n in (3, 2):
        tld_candidate = ".".join(parts[-(n - 1):])
        if tld_candidate in _MULTI_TLDS and len(parts) > n:
            return "." + ".".join(parts[-(n + 1):])
    return "." + ".".join(parts[-2:])


async def accept_cookies(page):
    # German / multilingual fall-back accept phrases — many EU sites surface
    # German first regardless of the URL locale.
    extra_text = COOKIE_TEXT_PATTERNS + [
        "Alle akzeptieren", "Alle Akzeptieren", "Akzeptieren", "Annehmen",
        "Zustimmen", "Auswahl bestätigen", "Einverstanden", "Erlauben",
        "Tout accepter", "Accepter", "Aceptar todas", "Aceptar",
        "Aceitar tudo", "Aceitar", "Accetta tutto",
    ]
    accepted = False
    # Try selectors twice — CMPs can stack a second dialog after the first.
    for attempt in range(2):
        clicked_this_round = False
        for sel in COOKIE_SELECTORS:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=200):
                    await el.click(timeout=2000, force=True)
                    accepted = True
                    clicked_this_round = True
                    await asyncio.sleep(0.4)
                    break
            except: pass
        if not clicked_this_round:
            for text in extra_text:
                try:
                    btn = page.get_by_role("button", name=text, exact=False).first
                    if await btn.is_visible(timeout=200):
                        await btn.click(timeout=2000, force=True)
                        accepted = True
                        clicked_this_round = True
                        await asyncio.sleep(0.4)
                        break
                except: pass
        if not clicked_this_round:
            break
    return accepted


async def reject_cookies(page):
    for sel in REJECT_SELECTORS:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=200):
                await el.click(timeout=2000)
                return True
        except: pass
    for text in REJECT_TEXT_PATTERNS:
        try:
            btn = page.get_by_role("button", name=text, exact=False).first
            if await btn.is_visible(timeout=200):
                await btn.click(timeout=2000)
                return True
        except: pass
    return False

def parse_analytics_payload(url, post_data=""):
    combined = url
    if post_data:
        combined = url + "&" + post_data if "?" in url else url + "?" + post_data
    low = combined.lower()
    result = {
        "is_adobe": False, "adobe_pv": False, "adobe_rsid": "",
        "is_ga4": False, "ga4_pv": False, "ga4_mid": "",
        "is_gtm": False, "gtm_id": "",
        "is_tealium_js": False, "tealium_account": "", "tealium_profile": "", "tealium_env": "",
    }

    if "/b/ss/" in low:
        result["is_adobe"] = True
        m = re.search(r'/b/ss/([^/]+)/', url)
        if m: result["adobe_rsid"] = m.group(1)
        if "pe=" not in low: result["adobe_pv"] = True

    if any(x in low for x in [".omtrdc.net", ".2o7.net", "appmeasurement", "s_code", "satellite-", "launch-"]):
        result["is_adobe"] = True

    if any(x in low for x in ["adobedc.net", "adobedc.demdex", "/ee/v", "/interact"]):
        result["is_adobe"] = True
        if post_data:
            try:
                body = json.loads(post_data)
                events = body.get("events", [])
                for ev in events:
                    xdm = ev.get("xdm", {})
                    if xdm.get("eventType") == "web.webpagedetails.pageViews": result["adobe_pv"] = True
                    web = xdm.get("web", {})
                    if web.get("webPageDetails", {}).get("pageViews", {}).get("value"): result["adobe_pv"] = True
            except: pass

    if "tiqcdn.com" in low and "utag" in low:
        result["is_tealium_js"] = True
        m = re.search(r'tiqcdn\.com/utag/([^/]+)/([^/]+)/([^/]+)/', url, re.I)
        if m:
            result["tealium_account"] = m.group(1)
            result["tealium_profile"] = m.group(2)
            result["tealium_env"] = m.group(3)

    if "googletagmanager.com/gtm.js" in low:
        result["is_gtm"] = True
        m = re.search(r'[?&]id=(GTM-[A-Z0-9]+)', url, re.I)
        if m: result["gtm_id"] = m.group(1).upper()

    if "googletagmanager.com/gtag/js" in low:
        m = re.search(r'[?&]id=(G-[A-Z0-9]+)', url, re.I)
        if m:
            result["ga4_mid"] = m.group(1).upper()
            result["is_ga4"] = True

    if "/g/collect" in low or (("google-analytics.com" in low or "analytics.google.com" in low) and "collect" in low):
        result["is_ga4"] = True
        m = re.search(r'[?&]tid=(G-[A-Z0-9]+)', combined, re.I)
        if m: result["ga4_mid"] = m.group(1).upper()
        if "en=page_view" in low: result["ga4_pv"] = True

    if result["is_ga4"] and post_data and not result["ga4_pv"]:
        if "page_view" in post_data.lower(): result["ga4_pv"] = True

    return result

async def validate_tags(browser, url, index, total):
    results = {
        "URL": url,
        "Tealium_Loaded": "FAIL", "Tealium_Account": "", "Tealium_Profile": "", "Tealium_Env": "",
        "GTM_Loaded": "FAIL", "GTM_ID": "",
        "GA4_Fired": "FAIL", "GA4_Measurement_ID": "", "GA4_PageView": "FAIL",
        "Adobe_Loaded": "FAIL", "Adobe_ReportSuite": "", "Adobe_PageView": "FAIL",
        "Error": ""
    }

    gtm_ids, ga4_ids, adobe_rsids = set(), set(), set()
    tealium_accounts = []
    flags = {"tealium_js": False, "gtm": False, "ga4": False, "ga4_pv": False, "adobe": False, "adobe_pv": False}

    context = None
    try:
        context = await browser.new_context(
            viewport={'width': 1280, 'height': 800},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/125.0.0.0"
        )
        page = await context.new_page()
        await stealth_obj.apply_stealth_async(page)

        cdp = await page.context.new_cdp_session(page)
        await cdp.send("Network.enable")

        def on_cdp_request(params):
            req = params.get("request", {})
            parsed = parse_analytics_payload(req.get("url", ""), req.get("postData", ""))
            if parsed["is_adobe"]: flags["adobe"] = True
            if parsed["adobe_pv"]: flags["adobe_pv"] = True
            if parsed["adobe_rsid"]: adobe_rsids.add(parsed["adobe_rsid"])
            if parsed["is_tealium_js"]:
                flags["tealium_js"] = True
                if parsed["tealium_account"]:
                    tealium_accounts.append({"account": parsed["tealium_account"], "profile": parsed["tealium_profile"], "env": parsed["tealium_env"]})
            if parsed["is_gtm"]:
                flags["gtm"] = True
                if parsed["gtm_id"]: gtm_ids.add(parsed["gtm_id"])
            if parsed["is_ga4"]:
                flags["ga4"] = True
                if parsed["ga4_mid"]: ga4_ids.add(parsed["ga4_mid"])
            if parsed["ga4_pv"]: flags["ga4_pv"] = True

        cdp.on("Network.requestWillBeSent", on_cdp_request)

        def handle_request(request):
            parsed = parse_analytics_payload(request.url, "")
            if parsed["is_adobe"]: flags["adobe"] = True
            if parsed["adobe_pv"]: flags["adobe_pv"] = True
            if parsed["adobe_rsid"]: adobe_rsids.add(parsed["adobe_rsid"])
            if parsed["is_tealium_js"]:
                flags["tealium_js"] = True
                if parsed["tealium_account"] and not tealium_accounts:
                    tealium_accounts.append({"account": parsed["tealium_account"], "profile": parsed["tealium_profile"], "env": parsed["tealium_env"]})
            if parsed["is_gtm"]:
                flags["gtm"] = True
                if parsed["gtm_id"]: gtm_ids.add(parsed["gtm_id"])
            if parsed["is_ga4"]:
                flags["ga4"] = True
                if parsed["ga4_mid"]: ga4_ids.add(parsed["ga4_mid"])
            if parsed["ga4_pv"]: flags["ga4_pv"] = True

        page.on("request", handle_request)

        sys.stdout.write(f"[{index}/{total}] Checking: {url}\n")
        sys.stdout.flush()

        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            results["Error"] = "Timeout" if "Timeout" in str(e) else str(e)[:80]

        await asyncio.sleep(2)
        await accept_cookies(page)
        try: await page.wait_for_load_state("networkidle", timeout=12000)
        except: pass
        await asyncio.sleep(8) # Robust wait for late tags

        # --- FALLBACK: Performance API ---
        try:
            perf_urls = await page.evaluate("performance.getEntriesByType('resource').map(r => r.name)")
            for u in (perf_urls or []):
                p = parse_analytics_payload(u, "")
                if p["is_adobe"]: flags["adobe"] = True
                if p["adobe_pv"]: flags["adobe_pv"] = True
                if p["adobe_rsid"]: adobe_rsids.add(p["adobe_rsid"])
                if p["is_ga4"]: flags["ga4"] = True
                if p["ga4_mid"]: ga4_ids.add(p["ga4_mid"])
                if p["ga4_pv"]: flags["ga4_pv"] = True
                if p["is_tealium_js"]: flags["tealium_js"] = True
        except: pass

        # --- FALLBACK: JS Objects ---
        try:
            js_data = await page.evaluate("""
                (() => {
                    let res = { utag: !!window.utag, gtm: !!window.google_tag_manager, s: !!window.s, alloy: !!window.alloy };
                    if (window.utag && window.utag.cfg) {
                        res.teal_acc = window.utag.cfg.account;
                        res.teal_prof = window.utag.cfg.profile;
                    }
                    if (window.google_tag_manager) {
                        res.gtm_ids = Object.keys(window.google_tag_manager).filter(k => k.startsWith('GTM-') || k.startsWith('G-'));
                    }
                    if (window.dataLayer) {
                        res.dl_pv = window.dataLayer.some(e => e.event === 'page_view' || e.event === 'gtm.js');
                    }
                    return res;
                })()
            """)
            if js_data.get("utag"): flags["tealium_js"] = True
            if js_data.get("gtm"): flags["gtm"] = True
            if js_data.get("s") or js_data.get("alloy"):
                flags["adobe"] = True
                flags["adobe_pv"] = True
            if js_data.get("gtm_ids"):
                for k in js_data["gtm_ids"]:
                    if k.startswith("GTM-"): gtm_ids.add(k)
                    if k.startswith("G-"): ga4_ids.add(k); flags["ga4"] = True
            if js_data.get("dl_pv"): flags["ga4_pv"] = True
        except: pass

        # BUILD RESULTS
        results["Tealium_Loaded"] = "PASS" if flags["tealium_js"] else "FAIL"
        if tealium_accounts:
            results["Tealium_Account"] = tealium_accounts[0]["account"]
            results["Tealium_Profile"] = tealium_accounts[0]["profile"]
            results["Tealium_Env"] = tealium_accounts[0]["env"]
        results["GTM_Loaded"] = "PASS" if flags["gtm"] else "FAIL"
        results["GTM_ID"] = ", ".join(sorted(gtm_ids)) if gtm_ids else ""
        results["GA4_Fired"] = "PASS" if flags["ga4"] else "FAIL"
        results["GA4_Measurement_ID"] = ", ".join(sorted(ga4_ids)) if ga4_ids else ""
        results["GA4_PageView"] = "PASS" if flags["ga4_pv"] else "FAIL"
        results["Adobe_Loaded"] = "PASS" if flags["adobe"] else "FAIL"
        results["Adobe_ReportSuite"] = ", ".join(sorted(adobe_rsids)) if adobe_rsids else ""
        results["Adobe_PageView"] = "PASS" if flags["adobe_pv"] else "FAIL"

        sys.stdout.write(f"[{index}/{total}] Done: {url} | PageView: {'YES' if flags['adobe_pv'] or flags['ga4_pv'] else 'NO'}\n")
        sys.stdout.flush()
        await page.close()
    except Exception as e:
        results["Error"] = f"Fatal: {str(e)[:80]}"
    finally:
        if context: await context.close()
    return results

def parse_ga4_event(url, post_data=""):
    """Extract a list of GA4/UA events and their parameters from a collect hit."""
    url_parts = url.split("?", 1)
    url_qs = url_parts[-1] if len(url_parts) > 1 else ""
    url_params = parse_qs(url_qs)

    lines = []
    if post_data:
        # Split post body into lines (each line represents a separate event in a batch)
        lines = [line.strip() for line in post_data.splitlines() if line.strip()]

    # If no post data lines, fall back to parsing parameters from the URL itself
    if not lines:
        lines = [url_qs]

    events = []
    for line in lines:
        line_params = parse_qs(line)
        # Combine URL common params with line-specific params (line-specific overrides)
        params = {**url_params, **line_params}

        event_name = params.get("en", [""])[0]
        is_ua = False
        if not event_name:
            if params.get("t", [""])[0] == "event":
                event_name = params.get("ea", [""])[0] or "UA Event"
                is_ua = True

        if not event_name:
            continue

        result = {
            "event": unquote(event_name),
            "measurement_id": params.get("tid", [""])[0],
            "params": {}
        }

        if is_ua:
            if params.get("ec"): result["params"]["category"] = unquote(params["ec"][0])
            if params.get("ea"): result["params"]["action"] = unquote(params["ea"][0])
            if params.get("el"): result["params"]["label"] = unquote(params["el"][0])
            if params.get("ev"): result["params"]["value"] = params["ev"][0]
        else:
            # Extract event parameters (ep.*, epn.* for numeric, up.* for user properties)
            for k, v in params.items():
                if k.startswith("ep."):
                    result["params"][k[3:]] = unquote(v[0])
                elif k.startswith("epn."):
                    result["params"][k[4:]] = v[0]  # numeric param
                elif k.startswith("up."):
                    result["params"]["[user] " + k[3:]] = unquote(v[0])

        events.append(result)

    return events


def parse_adobe_link_call(url, post_data=""):
    """Extract Adobe Analytics tracking details from a /b/ss/ hit.

    Handles THREE beacon kinds, all of which a QA engineer sees in Omnibug:
      * classic link tracking  (pe=lnk_o / lnk_d / lnk_e)
      * custom event beacons   (events=event1,event2 with no pe=)
      * page views             (s.t(): no pe=, no events=, has pageName) —
                                these fire on navigation and are the primary
                                signal on page-view-tracked sites."""
    blob = url + ("&" + post_data if post_data else "")
    params = parse_qs(blob.split("?", 1)[-1] if "?" in blob else "")

    pe = params.get("pe", [""])[0]
    events_param = params.get("events", [""])[0]
    page_name = unquote(params.get("pageName", [""])[0])
    g_url = unquote(params.get("g", [""])[0])   # the page URL the beacon was for

    if pe:
        link_type = pe                  # lnk_o / lnk_d / lnk_e
    elif events_param:
        link_type = "event"
    else:
        link_type = "page_view"

    # Human-readable name: link name (pev2) for link hits, else the pageName.
    link_name = unquote(params.get("pev2", [""])[0]) or (page_name if link_type == "page_view" else "")

    result = {
        "link_type": link_type,
        "link_name": link_name,
        "link_url": unquote(params.get("pev1", [""])[0]) or g_url,
        "page_name": page_name,
        "report_suite": "",
        "events": events_param,
        "evars": {},
        "props": {},
    }

    # Extract report suite from URL path: /b/ss/{rsid}/
    m = re.search(r'/b/ss/([^/]+)/', url)
    if m:
        result["report_suite"] = m.group(1)

    # Extract eVars (v1-v255) and props (c1-c75)
    for k, v in params.items():
        if re.match(r'^v\d+$', k):
            result["evars"][f"eVar{k[1:]}"] = unquote(v[0])
        elif re.match(r'^c\d+$', k):
            result["props"][f"prop{k[1:]}"] = unquote(v[0])

    return result

def parse_adobe_websdk_event(post_data):
    """Parse Adobe Web SDK interaction events from POST JSON payload.
    Extracts eventType, webInteraction name, URL, and type."""
    events = []
    if not post_data:
        return events
    try:
        data = json.loads(post_data)
        ev_list = data.get("events", [])
        if not isinstance(ev_list, list):
            ev_list = [ev_list] if isinstance(ev_list, dict) else []
            
        for ev in ev_list:
            if not isinstance(ev, dict):
                continue
            xdm = ev.get("xdm", {})
            if not isinstance(xdm, dict):
                continue
            
            event_type = xdm.get("eventType", "")
            web = xdm.get("web", {})
            if not isinstance(web, dict):
                web = {}
            web_interaction = web.get("webInteraction", {})
            if not isinstance(web_interaction, dict):
                web_interaction = {}
                
            interaction_name = web_interaction.get("name", "")
            interaction_url = web_interaction.get("URL", "")
            interaction_type = web_interaction.get("type", "")
            
            if event_type or interaction_name or interaction_url:
                events.append({
                    "event_type": event_type or "adobe_websdk_event",
                    "interaction_name": interaction_name,
                    "interaction_url": interaction_url,
                    "interaction_type": interaction_type,
                })
    except Exception:
        pass
    return events


# ===== CLICK TRACKING AUDIT =====

DISCOVER_CLICKABLES_JS = r"""
(() => {
    // key -> {el, idx}: the copy of a link kept for this key, and where it
    // sits in results, so a later copy can take its place (see below).
    const seen = new Map();
    const results = [];

    // Tags that are inherently interactive — clicks on these are meaningful.
    const INTERACTIVE_TAGS = new Set([
        'A', 'BUTTON', 'INPUT', 'SELECT', 'TEXTAREA', 'VIDEO', 'AUDIO',
        'SUMMARY', 'AREA', 'DETAILS', 'LABEL', 'OPTION',
    ]);

    // Selectors that target genuinely interactive elements directly.
    const directSelectors = [
        'a', 'area[href]', 'button', 'input[type="submit"]', 'input[type="button"]', 'input[type="image"]',
        'input', 'select', 'textarea',
        'video', 'audio', 'summary',
        '[role="button"]', '[role="link"]', '[role="tab"]', '[role="checkbox"]', '[role="radio"]',
        '[onclick]', '[data-action]', '[data-analytics]', '[data-tracking]', '[data-click]', '[data-event]', '[data-cta]', '[data-ga]',
        '[aria-expanded]', '[data-emu]', '[data-emu-track]', '[data-cmp-clickable]', '[data-track-click]', '[data-click-track]', '[data-analytics-id]',
        '.emu-button', '.emu-link', '.emu-image__link', '.cmp-button', '.cmp-link', '.cmp-image__link',
        '[class*="click-track"]', '[class*="track-click"]', '[class*="emu-button"]', '[class*="emu-link"]'
    ];

    // Deep element collection — walks the light DOM AND every open shadow
    // root so web-component menus (common on enterprise sites) are scanned.
    function collectAll(root, acc) {
        try {
            root.querySelectorAll('*').forEach(n => {
                acc.push(n);
                if (n.shadowRoot) collectAll(n.shadowRoot, acc);
            });
        } catch(e) {}
        return acc;
    }
    const ALL_ELEMENTS = collectAll(document, []);
    function queryDeep(sel) {
        const out = [];
        for (const el of ALL_ELEMENTS) {
            try { if (el.matches && el.matches(sel)) out.push(el); } catch(e) {}
        }
        return out;
    }

    // Class/id-based selectors that often match CONTAINER elements (div, li, span)
    // instead of the actual interactive child. We scan these separately and
    // resolve the real clickable inside them.
    const containerSelectors = [
        '.cta', '.btn', '[class*="btn"]', '[class*="button"]', '[class*="toggle"]', '[class*="accordion"]',
        '[class*="tab"]', '[class*="expand"]', '[class*="collapse"]', '[class*="menu"]', '[class*="video"]',
        '[class*="player"]', '[id*="video"]', '[id*="player"]', '[class*="click"]', '[class*="link"]',
        'form',
    ];
    
    function buildSelector(el) {
        if (el.id) return '#' + CSS.escape(el.id);
        const path = [];
        let cur = el;
        while (cur && cur !== document.body && cur !== document.documentElement) {
            let seg = cur.tagName.toLowerCase();
            if (cur.id) { path.unshift('#' + CSS.escape(cur.id)); break; }
            const parent = cur.parentElement;
            if (parent) {
                const siblings = [...parent.children].filter(c => c.tagName === cur.tagName);
                if (siblings.length > 1) seg += ':nth-of-type(' + (siblings.indexOf(cur) + 1) + ')';
            }
            path.unshift(seg);
            cur = cur.parentElement;
        }
        return path.join(' > ');
    }
    
    function isVisible(el) {
        if (!el) return false;
        const rect = el.getBoundingClientRect();
        if (rect.width < 3 || rect.height < 3) return false;
        const style = window.getComputedStyle(el);
        if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
        return true;
    }

    // Is this element itself interactive (not just a container)?
    function isInteractive(el) {
        if (INTERACTIVE_TAGS.has(el.tagName)) return true;
        // Elements with explicit roles are interactive
        const role = el.getAttribute('role');
        if (role && ['button','link','tab','checkbox','radio','menuitem','switch','option'].includes(role)) return true;
        // Elements with inline event handlers are interactive
        if (el.onclick || el.onmousedown || el.getAttribute('onclick')) return true;
        // data-action or similar click-intent attributes
        if (el.getAttribute('data-action') || el.getAttribute('data-click') || el.getAttribute('data-event')) return true;
        return false;
    }

    // For a container element, try to find the best interactive child.
    // Returns the child element, or null if none found.
    function findInteractiveChild(container) {
        // Prefer direct child links/buttons first
        const direct = container.querySelector('a, button, input[type="submit"], input[type="button"], [role="button"], [role="link"]');
        if (direct && isVisible(direct)) return direct;
        return null;
    }

    // Walk UP to find the nearest interactive ancestor (e.g. a div inside an a).
    // Stops at body. Returns the ancestor or null.
    function findInteractiveAncestor(el) {
        let cur = el.parentElement;
        while (cur && cur !== document.body && cur !== document.documentElement) {
            if (INTERACTIVE_TAGS.has(cur.tagName) || isInteractive(cur)) {
                return cur;
            }
            cur = cur.parentElement;
        }
        return null;
    }

    const elementsToScan = new Set();

    // 1. Gather directly interactive elements (deep: pierces shadow DOM).
    //    CRUCIAL: hidden-but-real links are kept too. Mega-menu / dropdown
    //    links are display:none until their parent is hovered, so a
    //    visible-only scan misses most of the nav tree (observed: 224 of
    //    279 links hidden on an enterprise homepage). The click loop
    //    re-expands menus before clicking, so these are fully testable.
    for (const sel of directSelectors) {
        try {
            queryDeep(sel).forEach(el => {
                if (isVisible(el)) {
                    elementsToScan.add(el);
                } else if (
                    (el.tagName === 'A' && el.getAttribute('href')) ||
                    el.tagName === 'BUTTON'
                ) {
                    elementsToScan.add(el);
                }
            });
        } catch(e) {}
    }

    // 2. Gather from class/id-based selectors — resolve containers to their
    //    interactive children to avoid reporting DIVs/SPANs/LIs as clickables.
    for (const sel of containerSelectors) {
        try {
            queryDeep(sel).forEach(el => {
                if (!isVisible(el)) return;
                if (isInteractive(el)) {
                    // The matched element itself is interactive
                    elementsToScan.add(el);
                } else {
                    // First check: is this container INSIDE an interactive ancestor?
                    // e.g. a div[class*="btn"] inside an <a> — prefer the <a>.
                    const ancestor = findInteractiveAncestor(el);
                    if (ancestor) {
                        elementsToScan.add(ancestor);
                        return;
                    }
                    // Container (div, li, span, etc.) — find the real interactive child
                    const child = findInteractiveChild(el);
                    if (child) {
                        elementsToScan.add(child);
                    }
                    // If no interactive child AND the container has cursor:pointer +
                    // an explicit click handler, keep it (custom JS widget).
                    else if (el.onclick || el.getAttribute('onclick') || el.getAttribute('data-action')) {
                        elementsToScan.add(el);
                    }
                    // Otherwise skip — it's just a container with a class like "tab-panel"
                }
            });
        } catch(e) {}
    }

    // 3. Gather other elements with inline events or cursor: pointer style
    try {
        ALL_ELEMENTS.forEach(el => {
            if (elementsToScan.has(el)) return;
            try {
                if (el.onclick || el.onmousedown || el.getAttribute('onclick')) {
                    if (isVisible(el)) elementsToScan.add(el);
                    return;
                }
                // Only add cursor:pointer elements if they are themselves interactive
                // or are a small leaf element (not a big container).
                const style = window.getComputedStyle(el);
                if (style.cursor === 'pointer') {
                    if (!isVisible(el)) return;
                    // If this element is inside an interactive ancestor (e.g. div inside a),
                    // prefer the ancestor — it's already discovered or will be.
                    if (!INTERACTIVE_TAGS.has(el.tagName) && !isInteractive(el)) {
                        const ancestor = findInteractiveAncestor(el);
                        if (ancestor) {
                            elementsToScan.add(ancestor);
                            return;
                        }
                    }
                    // Skip large container divs that just inherit pointer from children
                    const rect = el.getBoundingClientRect();
                    const isLargeContainer = !INTERACTIVE_TAGS.has(el.tagName)
                        && (rect.width > 600 && rect.height > 200);
                    if (!isLargeContainer) {
                        // Check if it's a container with an interactive child
                        if (INTERACTIVE_TAGS.has(el.tagName) || isInteractive(el)) {
                            elementsToScan.add(el);
                        } else {
                            const child = findInteractiveChild(el);
                            if (child) {
                                elementsToScan.add(child);
                            } else {
                                // Small pointer element with no interactive child — keep it
                                elementsToScan.add(el);
                            }
                        }
                    }
                }
            } catch(e) {}
        });
    } catch(e) {}

    function shouldIgnoreElement(el) {
        let cur = el;
        while (cur && cur !== document.body && cur !== document.documentElement) {
            const id = (cur.id || '').toLowerCase();
            const className = (typeof cur.className === 'string' ? cur.className : '').toLowerCase();
            
            if (id.includes('cookie') || id.includes('consent') || id.includes('onetrust') || id.includes('ot-') || id.includes('privacy') || id.includes('emu-consent') || id.includes('chat-') || id.includes('chatbot') || id.includes('ai-assistant')) {
                return true;
            }
            if (className.includes('cookie') || className.includes('consent') || className.includes('onetrust') || className.includes('ot-') || className.includes('privacy') || className.includes('chat-') || className.includes('chatbot') || className.includes('assistant-chat')) {
                return true;
            }
            cur = cur.parentElement;
        }
        
        const ownText = (el.innerText || el.value || '').toLowerCase().trim();
        if (ownText.includes('cookie settings') || ownText.includes('cookies settings') || ownText.includes('cookie preferences') || ownText.includes('manage cookies') || ownText.includes('accept cookies') || ownText.includes('reject cookies') || ownText.includes('accept all') || ownText.includes('reject all')) {
            return true;
        }
        
        return false;
    }

    // Dedup ONLY by logical identity (tag + text + href + id) — never by
    // pixel coordinates. Tiles and cards often contain multiple distinct
    // links at near-identical positions; coordinate-dedup was skipping them.
    elementsToScan.forEach(el => {
        try {
            if (shouldIgnoreElement(el)) return;

            // Prefer visible innerText; fall back to textContent (innerText
            // is EMPTY for display:none elements — hidden menu links need
            // this), then value, placeholder, aria-label, title, alt — so
            // icon-only buttons / image links still get a meaningful label.
            // Icon-only and image links are common in headers/footers (logo,
            // social icons, back-to-top arrows). A bare "[a]" label makes the
            // report unreadable AND unmatchable during beacon attribution,
            // so dig for any human-readable name the markup offers.
            const imgAlt = (sel) => {
                try {
                    const n = el.querySelector && el.querySelector(sel);
                    if (!n) return '';
                    return (n.getAttribute('alt') || n.getAttribute('aria-label')
                            || n.getAttribute('title') || n.textContent || '').trim();
                } catch (e) { return ''; }
            };
            // Look for any of `names` on the element, then on up to 4
            // ancestors — AEM puts the authored title on the wrapping
            // component div, not on the <a> itself.
            // Each attribute name is searched across the whole ancestor chain
            // before moving to the next name, so a descriptive authored title
            // on a wrapper ("Skinvive by Juvederm") beats a generic component
            // type sitting closer on the element itself ("imageComponent").
            const attrUp = (names) => {
                for (const a of names) {
                    try {
                        let n = el, hops = 0;
                        while (n && n !== document.body && hops < 5) {
                            const v = n.getAttribute && n.getAttribute(a);
                            if (v && v.trim() && v.trim() !== 'true' && v.trim() !== 'false') {
                                return v.trim();
                            }
                            n = n.parentElement; hops++;
                        }
                    } catch (e) {}
                }
                return '';
            };
            // "/content/dam/.../close-modal.svg" -> "close modal"
            const fileLabel = (p) => {
                try {
                    if (!p) return '';
                    let base = String(p).split('?')[0].split('/').pop() || '';
                    base = base.replace(/\.(svg|png|jpe?g|gif|webp|coreimg.*)$/i, '');
                    base = base.replace(/[-_.]+/g, ' ').replace(/\s+/g, ' ').trim();
                    return base.length > 1 ? base : '';
                } catch (e) { return ''; }
            };
            // Last resort before "[button]": a readable form of the id, minus
            // the random AEM hash suffix (button-fd5bb9814f -> "button").
            const idLabel = (n) => {
                try {
                    let s = (n.id || '').replace(/-[0-9a-f]{6,}$/i, '');
                    s = s.replace(/[-_]+/g, ' ').trim();
                    return (s && s.toLowerCase() !== 'button' && s.length > 2) ? s : '';
                } catch (e) { return ''; }
            };
            const text = (el.innerText
                || (el.textContent || '').replace(/\s+/g, ' ').trim()
                || el.value
                || el.placeholder
                || el.getAttribute('aria-label')
                || el.getAttribute('title')
                || el.getAttribute('alt')
                || imgAlt('img')
                || imgAlt('picture img')
                || imgAlt('svg title')
                || imgAlt('svg')
                || imgAlt('use')
                || imgAlt('[aria-label]')
                || imgAlt('[alt]')
                // --- AEM / Adobe EMU component metadata ---
                // Enterprise AEM sites render icon buttons and SVG logos with
                // no text, no alt and no aria-label; the human name lives in
                // authoring attributes instead. Without these the report is a
                // wall of "[button]" rows, and — worse — beacon attribution
                // loses its text key and can bind a hit to the wrong element.
                || attrUp(['data-selector-text', 'data-title', 'data-analytics-id',
                           'data-gtm-attribute', 'aria-label', 'title'])
                || fileLabel(attrUp(['data-icon-reference', 'data-asset', 'data-cmp-src']))
                || idLabel(el)
                || '').replace(/\s+/g, ' ').trim().substring(0, 80);
            const href = (el.href || el.getAttribute('href') || '').trim();

            // Tag each element with its page zone so the audit can click
            // header -> footer -> body in a clean order. Many enterprise
            // sites (AEM/Experience Fragments etc.) don't use semantic
            // <footer>/<header> tags at all — just divs with classes like
            // "cmp-experiencefragment--footer". So walk ancestors checking
            // id/class SUBSTRINGS, not a fixed tag/role selector list.
            // Footer is checked first: footer link lists often sit inside
            // a <nav>, so a header check done first would wrongly claim them.
            let zone = 'body';
            try {
                let cur = el;
                while (cur && cur !== document.body && cur !== document.documentElement) {
                    const id = (cur.id || '').toLowerCase();
                    const cls = (typeof cur.className === 'string' ? cur.className : '').toLowerCase();
                    const role = (cur.getAttribute && cur.getAttribute('role') || '').toLowerCase();
                    if (cur.tagName === 'FOOTER' || role === 'contentinfo' || id.includes('footer') || cls.includes('footer')) {
                        zone = 'footer';
                        break;
                    }
                    if (cur.tagName === 'HEADER' || cur.tagName === 'NAV' || role === 'banner'
                        || id.includes('header') || cls.includes('header')
                        || id.includes('navbar') || cls.includes('navbar')
                        || cls.includes('top-nav') || cls.includes('main-nav')) {
                        zone = 'header';
                        break;
                    }
                    cur = cur.parentElement;
                }
            } catch(e) {}

            // Dedup by normalised text + href + ZONE (case-insensitive,
            // whitespace-collapsed). Zone is part of the key so the same
            // link in header AND footer is clicked in BOTH places — link
            // position often changes the analytics payload, so QA needs
            // each occurrence. Empty-text icons keep DOM identity.
            const textKey = text.toLowerCase().replace(/\s+/g, ' ');
            const hrefKey = href.toLowerCase();
            const key = (textKey
                ? 'T:' + textKey + '|H:' + hrefKey
                : (el.tagName + '|H:' + hrefKey + '|I:' + (el.id || ''))) + '|Z:' + zone;
            // Responsive navs commonly render the SAME link twice: one copy
            // for the desktop bar and one for the mobile drawer, only one of
            // which is laid out. Keeping whichever came first in the DOM can
            // keep the copy with no box at all — it is then unreachable, the
            // audit falls back to the wrapper, and clicking the wrapper's
            // centre hits a SIBLING link. That is how a row for "Children with
            // Spasticity" ended up reporting "Adults with Spasticity". So on a
            // duplicate, prefer the copy that is actually laid out.
            const boxed = (n) => {
                try {
                    const r = n.getBoundingClientRect();
                    if (r.width > 2 && r.height > 2) return true;
                    for (const c of n.getClientRects()) {
                        if (c.width > 2 && c.height > 2) return true;
                    }
                } catch (e) {}
                return false;
            };
            let replaceIdx = -1;
            const prior = seen.get(key);
            if (prior) {
                if (!boxed(prior.el) && boxed(el)) replaceIdx = prior.idx;
                else return;
            }
            seen.set(key, {el: el, idx: replaceIdx >= 0 ? replaceIdx : results.length});

            // STAMP a unique, stable handle straight onto the node. The click
            // loop locates elements by [data-tvuid="..."] instead of a CSS
            // path, which removes the whole class of "clicked the wrong
            // element" bugs: non-unique selectors resolving via .first, and
            // header/footer duplicates that share text+href. Idempotent —
            // discovery runs many times (hover/expand passes) and an already
            // stamped node keeps its original uid.
            let uid = '';
            try {
                uid = el.getAttribute('data-tvuid') || '';
                if (!uid) {
                    window.__tvSeq = (window.__tvSeq || 0) + 1;
                    uid = 'tv' + window.__tvSeq;
                    el.setAttribute('data-tvuid', uid);
                }
            } catch(e) {}

            // Download intent: GA4 file_download / download events only fire
            // for these, and they need the download-interception path so the
            // browser doesn't stall on a real file transfer.
            const dlAttr = el.getAttribute && el.getAttribute('download');
            const isDownload = (dlAttr !== null && dlAttr !== undefined)
                || /\.(pdf|docx?|xlsx?|pptx?|zip|csv|txt|rtf|dmg|exe|pkg|mp4|mp3|wav|avi|mov)(\?|#|$)/i.test(href);

            // Nearest enclosing heading / card title. SDRs routinely have a
            // dozen rows whose Link/Button Name is identical ("Read more")
            // and distinguish them only by Location ("body - krista
            // martins"), so the element needs to carry the label of the card
            // it sits in or those rows cannot be told apart.
            let context = '';
            try {
                let n = el;
                for (let i = 0; i < 8 && n; i++) {
                    n = n.parentElement;
                    if (!n) break;
                    const h = n.querySelector('h1,h2,h3,h4,h5,h6,[class*="title"],[class*="heading"],[class*="name"]');
                    if (h) {
                        const ht = (h.innerText || h.textContent || '').replace(/\s+/g, ' ').trim();
                        if (ht && ht.length <= 90) { context = ht; break; }
                    }
                }
            } catch (e) {}

            // One control answers to several names. A footer logo may render
            // as an aria-label ("Navigate to Skinvive Home Page") while the
            // SDR calls it by its image alt ("SKINVIVE BY JUVEDERM Logo").
            // Picking one and discarding the rest means the row cannot find
            // its own element, so keep every name the markup offers and let
            // matching try them all.
            const labels = [];
            const addLabel = (v) => {
                const t = (v || '').replace(/\s+/g, ' ').trim();
                if (t && t.length <= 90 && !labels.includes(t)) labels.push(t);
            };
            try {
                addLabel(el.innerText);
                addLabel(el.getAttribute('aria-label'));
                addLabel(el.getAttribute('title'));
                addLabel(el.getAttribute('alt'));
                addLabel(el.getAttribute('data-selector-text'));
                addLabel(attrUp(['data-title']));
                const im = el.querySelector && el.querySelector('img');
                if (im) { addLabel(im.getAttribute('alt')); addLabel(im.getAttribute('title')); }
                const sv = el.querySelector && el.querySelector('svg title');
                if (sv) addLabel(sv.textContent);
                const inner = el.querySelector && el.querySelector('[aria-label],[alt]');
                if (inner) {
                    addLabel(inner.getAttribute('aria-label'));
                    addLabel(inner.getAttribute('alt'));
                }
            } catch (e) {}

            // Where the element lives, in the page's own words. `zone` only
            // knows header / footer / body, but SDRs name regions the way the
            // build does — "isi", "hero", "submenu", "footer_promo". The same
            // link often appears in two of them (a Directions for Use in the
            // nav and another inside the ISI block) and the region is the only
            // thing telling those rows apart, so carry the id/class words of
            // every ancestor and let matching compare against them.
            const hints = [];
            // An icon-only control is named by its icon. The back-to-top
            // button renders as "Back to top" but contains <g id="arrow">,
            // and the sheet calls that row "Arrow" — nothing in the rendered
            // label connects the two. Take the id/class words from inside the
            // control as well as from its ancestors.
            try {
                el.querySelectorAll('svg, g, path, use, i, span, img').forEach(k => {
                    const bag = ((k.id || '') + ' ' +
                                 (typeof k.className === 'string' ? k.className : '') + ' ' +
                                 (k.getAttribute('data-icon') || '')).toLowerCase();
                    bag.split(/[^a-z0-9]+/).forEach(tok => {
                        if (tok.length >= 3 && tok.length <= 24 && !hints.includes(tok)) {
                            hints.push(tok);
                        }
                    });
                });
            } catch (e) {}
            try {
                let n = el, hops = 0;
                while (n && n !== document.body && hops < 10) {
                    const bag = ((n.id || '') + ' ' +
                                 (typeof n.className === 'string' ? n.className : '')).toLowerCase();
                    bag.split(/[^a-z0-9]+/).forEach(tok => {
                        if (tok.length >= 2 && tok.length <= 24 && !hints.includes(tok)) {
                            hints.push(tok);
                        }
                    });
                    n = n.parentElement; hops++;
                }
            } catch (e) {}

            const record = {
                selector: buildSelector(el),
                uid: uid,
                context: context,
                labels: labels,
                hints: hints.slice(0, 40),
                tag: el.tagName,
                text: text || `[${el.tagName.toLowerCase()}]`,
                href: href,
                id: el.id || '',
                zone: zone,
                hidden: !isVisible(el),   // needs menu re-expansion before click
                is_download: !!isDownload,
                target: (el.getAttribute && el.getAttribute('target')) || '',
                className: (typeof el.className === 'string') ? el.className.substring(0, 100) : '',
            };
            if (replaceIdx >= 0) results[replaceIdx] = record;
            else results.push(record);
        } catch(e) {}
    });

    // QA-style: scan up to 2000 elements per page (effectively unlimited for
    // any normal site). Hard cap is just a safety net for runaway lists.
    return results.slice(0, 2000);
})()
"""


# Pre-discovery: open every collapsible / dropdown / accordion / hover-menu so
# the link-discovery pass sees the items inside. Without this, sub-menu links
# never get clicked because they aren't in the DOM (or are display:none) at
# discovery time. We do several passes because expanding one menu sometimes
# reveals more collapsibles inside it.
EXPOSE_HIDDEN_JS = r"""
(async () => {
    const sleep = (ms) => new Promise(r => setTimeout(r, ms));

    function fireMouseEnter(el) {
        try {
            ['pointerover','mouseover','pointerenter','mouseenter'].forEach(type => {
                el.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
            });
        } catch(e) {}
    }

    function isVisible(el) {
        if (!el) return false;
        const rect = el.getBoundingClientRect();
        if (rect.width < 1 || rect.height < 1) return false;
        const s = window.getComputedStyle(el);
        return !(s.display === 'none' || s.visibility === 'hidden');
    }

    let totalRevealed = 0;
    // Several rounds — opening one menu can reveal nested collapsibles.
    for (let round = 0; round < 6; round++) {
        let actedThisRound = 0;

        // 1. Open native <details>
        document.querySelectorAll('details:not([open])').forEach(d => {
            try { d.open = true; actedThisRound++; } catch(e) {}
        });

        // 2. Click items that explicitly mark themselves collapsed.
        const collapsedSelectors = [
            '[aria-expanded="false"]',
            '[data-toggle="collapse"]', '[data-bs-toggle="collapse"]',
            '[data-toggle="dropdown"]', '[data-bs-toggle="dropdown"]',
            '[aria-haspopup="true"]:not([aria-expanded="true"])',
            'button[class*="accordion"]', 'button[class*="expand"]',
            'button[class*="toggle"]:not([aria-expanded="true"])',
        ];
        for (const sel of collapsedSelectors) {
            const els = document.querySelectorAll(sel);
            for (const el of els) {
                if (!isVisible(el)) continue;
                try { el.click(); actedThisRound++; } catch(e) {}
                if (actedThisRound > 150) break;   // don't blow up huge pages in one round
            }
            if (actedThisRound > 150) break;
        }

        // 3. Hover-style menus (top nav). Dispatch mouse events on every
        // candidate parent so CSS :hover / JS mouseenter handlers reveal
        // their sub-menus.
        const hoverSelectors = [
            'nav li', 'nav a', '[role="menuitem"]', '[role="menubar"] > *',
            'header li', '.menu li', '.nav li', '[class*="dropdown"] > a',
            '[class*="has-submenu"]', '[class*="has-children"]',
        ];
        for (const sel of hoverSelectors) {
            const els = document.querySelectorAll(sel);
            for (const el of els) {
                if (!isVisible(el)) continue;
                fireMouseEnter(el);
            }
        }

        if (actedThisRound === 0) break;   // nothing left to open
        totalRevealed += actedThisRound;
        await sleep(250);
    }

    return totalRevealed;
})()
"""


# Block ALL navigation during the click audit so a single anchor click doesn't
# take us off the page and end the loop. preventDefault only stops the
# browser-default navigation; the analytics click listeners (Adobe s.tl, GA4
# gtag, GTM dataLayer pushes) still run as normal because they fire in the
# same event tick.
BLOCK_NAVIGATION_JS = r"""
(() => {
    if (window.__navBlockerInstalled) return true;
    window.__navBlockerInstalled = true;

    window.__navBlockerClickHandler = (e) => {
        try {
            const a = e.target && e.target.closest && e.target.closest('a');
            if (a) e.preventDefault();
        } catch (err) {}
    };
    document.addEventListener('click', window.__navBlockerClickHandler, false);

    window.__navBlockerSubmitHandler = (e) => {
        try { e.preventDefault(); } catch (err) {}
    };
    document.addEventListener('submit', window.__navBlockerSubmitHandler, false);

    return true;
})()
"""


# Close any consent UI that is sitting on top of the page.
#
# The menu-discovery pass clicks anything that looks like a menu opener, and
# "Cookies Settings" matches `button[aria-expanded="false"]`. That opens the
# OneTrust Preference Center, whose full-screen dark filter then covers the
# entire header for the rest of the audit: every header link hit-tests as
# obscured, gets clicked by a fallback that the site's own handlers ignore,
# and is reported as untracked. Consent has already been granted explicitly
# before this runs, so dismissing the leftover dialog changes no consent
# state — it just gets the overlay out of the way.
# Make a covered element reachable by a REAL mouse click.
#
# Some clicks open a panel that stays on top of other elements for the rest of
# the run. Falling back to a JS .click() reaches the right node but produces an
# untrusted event, and plenty of analytics handlers (Adobe ActivityMap, and
# this site's own site_link/exit_link handlers) simply ignore those — the
# element then gets reported as untracked even though it is tagged. Instead,
# walk the stack of elements sitting over the target and switch off their
# pointer-events just long enough to deliver a trusted click to the real
# element. Nothing is hidden or removed, and it is put back immediately after.
PIERCE_OVERLAY_JS = r"""
(e) => {
    // An element can render text and still report an empty box: `display:
    // contents` generates no box of its own, and some nav markup relies on
    // it. Measuring only getBoundingClientRect() declares a perfectly visible
    // link unreachable, the audit falls back to its wrapper, and clicking the
    // wrapper hits whichever sibling sits under its centre — which is how a
    // row for "Children with Spasticity" ended up reporting "Adults with
    // Spasticity". Fall back to the element's own client rects, then to its
    // first child that does have a box.
    const rectOf = (n) => {
        let r = n.getBoundingClientRect();
        if (r.width > 2 && r.height > 2) return r;
        for (const c of n.getClientRects()) {
            if (c.width > 2 && c.height > 2) return c;
        }
        for (const k of n.children) {
            const kr = k.getBoundingClientRect();
            if (kr.width > 2 && kr.height > 2) return kr;
        }
        return r;
    };

    // Always undo a previous pierce first. If an earlier one never got its
    // restore (an exception between click and cleanup), overwriting the
    // bookkeeping list would strand those nodes with pointer-events:none for
    // the rest of the audit and quietly break every later element.
    try {
        for (const [n, val, pri] of (window.__tvPierced || [])) {
            if (val) n.style.setProperty('pointer-events', val, pri || '');
            else n.style.removeProperty('pointer-events');
        }
    } catch (err) {}
    window.__tvPierced = [];

    const r = rectOf(e);
    if (!r.width || !r.height) return {ok: false, reason: 'no-box'};
    const pts = [
        [r.x + r.width / 2,    r.y + r.height / 2],
        [r.x + r.width * 0.25, r.y + r.height / 2],
        [r.x + r.width * 0.75, r.y + r.height / 2],
    ].filter(([x, y]) => x >= 0 && y >= 0 && x <= innerWidth && y <= innerHeight);
    if (!pts.length) return {ok: false, reason: 'offscreen'};

    const touched = [];
    const hits = (n) => n && (n === e || e.contains(n) || n.contains(e));
    for (const [x, y] of pts) {
        for (let guard = 0; guard < 10; guard++) {
            const top = document.elementFromPoint(x, y);
            if (hits(top)) {
                window.__tvPierced = touched;
                return {ok: true, x: x, y: y, pierced: touched.length};
            }
            if (!top || top === document.documentElement || top === document.body) break;
            touched.push([top, top.style.getPropertyValue('pointer-events'),
                          top.style.getPropertyPriority('pointer-events')]);
            top.style.setProperty('pointer-events', 'none', 'important');
        }
    }
    window.__tvPierced = touched;
    return {ok: false, reason: 'unreachable', pierced: touched.length};
}
"""

RESTORE_OVERLAY_JS = r"""
(() => {
    const t = window.__tvPierced || [];
    for (const [n, val, pri] of t) {
        try {
            if (val) n.style.setProperty('pointer-events', val, pri || '');
            else n.style.removeProperty('pointer-events');
        } catch (e) {}
    }
    window.__tvPierced = [];
    return t.length;
})()
"""

CLOSE_CONSENT_UI_JS = r"""
(() => {
    let closed = 0;
    try {
        if (window.OneTrust && typeof window.OneTrust.Close === 'function') {
            window.OneTrust.Close();
            closed++;
        }
    } catch (e) {}
    const overlays = [
        '#onetrust-pc-sdk', '.onetrust-pc-dark-filter', '.ot-pc-dark-filter',
        '#onetrust-banner-sdk', '#onetrust-consent-sdk .onetrust-pc-dark-filter',
        '.ot-fade-in.onetrust-pc-dark-filter',
        '#truste-consent-track', '#cookie-consent-overlay',
        '[id*="cookie"][class*="overlay"]', '[class*="cmp-overlay"]',
    ];
    for (const sel of overlays) {
        try {
            document.querySelectorAll(sel).forEach(n => {
                const st = window.getComputedStyle(n);
                if (st && st.display !== 'none' && st.visibility !== 'hidden') {
                    n.style.setProperty('display', 'none', 'important');
                    closed++;
                }
            });
        } catch (e) {}
    }
    return closed;
})()
"""


# Instrument the page's tracking APIs so click-tracking is detected even when
# the network beacon is deferred to the NEXT page view (Adobe ActivityMap),
# batched by a TMS, or suppressed by consent state. Sites like Fresenius track
# every link via ActivityMap — the click stores an s_sq cookie and the beacon
# only rides on the next page load, so pure network capture reports
# "no tracking" for links that ARE tracked. Wrapping the APIs catches the
# call itself, at the moment of click:
#   s.tl(...)                    -> Adobe custom link tracking
#   utag.link(...)               -> Tealium click events
#   gtag('event', ...)           -> GA4 API events
#   dataLayer.push({event:...})  -> GTM/GA4 click triggers
INSTRUMENT_TRACKING_JS = r"""
(() => {
    if (window.__qaTrackInstalled) return true;
    window.__qaTrackInstalled = true;
    window.__qaTrackCaptures = [];
    const push = (type, detail) => {
        try {
            window.__qaTrackCaptures.push({type: type, detail: detail, ts: Date.now()});
            if (window.__qaTrackCaptures.length > 500) window.__qaTrackCaptures.shift();
        } catch(e) {}
    };
    const safeJson = (o) => {
        try { return JSON.parse(JSON.stringify(o)); } catch(e) { return String(o); }
    };

    // Adobe AppMeasurement s.tl — re-check periodically, s is often defined
    // late by the tag manager (Tealium/Launch).
    const wrapS = () => {
        try {
            const s = window.s;
            if (s && typeof s.tl === 'function' && !s.__qaWrapped) {
                const orig = s.tl.bind(s);
                s.__qaWrapped = true;
                s.tl = function(obj, linkType, linkName, vars, doneAction) {
                    push('s.tl', {link_type: String(linkType || ''), link_name: String(linkName || '')});
                    return orig(obj, linkType, linkName, vars, doneAction);
                };
            }
        } catch(e) {}
    };
    // Tealium utag.link
    const wrapU = () => {
        try {
            const u = window.utag;
            if (u && typeof u.link === 'function' && !u.__qaWrapped) {
                const orig = u.link.bind(u);
                u.__qaWrapped = true;
                u.link = function(data, cb, uids) {
                    push('utag.link', safeJson(data || {}));
                    return orig(data, cb, uids);
                };
            }
        } catch(e) {}
    };
    // GA4 gtag
    const wrapG = () => {
        try {
            if (typeof window.gtag === 'function' && !window.gtag.__qaWrapped) {
                const orig = window.gtag;
                const wrapped = function() {
                    try {
                        if (arguments[0] === 'event') {
                            push('gtag', {event: String(arguments[1] || ''), params: safeJson(arguments[2] || {})});
                        }
                    } catch(e) {}
                    return orig.apply(this, arguments);
                };
                wrapped.__qaWrapped = true;
                window.gtag = wrapped;
            }
        } catch(e) {}
    };
    // GTM dataLayer.push — only capture entries with an `event` key
    const wrapDL = () => {
        try {
            const dl = window.dataLayer;
            if (dl && typeof dl.push === 'function' && !dl.__qaWrapped) {
                const orig = dl.push.bind(dl);
                dl.__qaWrapped = true;
                dl.push = function() {
                    try {
                        const a = arguments[0];
                        if (a && typeof a === 'object' && !Array.isArray(a) && a.event) {
                            push('dataLayer', safeJson(a));
                        }
                    } catch(e) {}
                    return orig.apply(null, arguments);
                };
            }
        } catch(e) {}
    };

    // Adobe Client Data Layer (adobeDataLayer) — the event bus AEM Core
    // Components / EMU sites push `cmp:click` and custom events into. It is
    // NOT window.dataLayer, so the GTM wrap above never saw it. On AEM sites
    // this is often the only click signal that fires at click time.
    const wrapACDL = () => {
        try {
            const adl = window.adobeDataLayer;
            if (adl && typeof adl.push === 'function' && !adl.__qaWrapped) {
                const orig = adl.push.bind(adl);
                adl.__qaWrapped = true;
                adl.push = function() {
                    try {
                        for (const a of arguments) {
                            if (a && typeof a === 'object' && !Array.isArray(a)) {
                                if (a.event || a['xdm:eventType']) push('adobeDataLayer', safeJson(a));
                            }
                        }
                    } catch(e) {}
                    return orig.apply(null, arguments);
                };
            }
        } catch(e) {}
    };

    // CLICK WITNESS — records what the browser actually delivered the click
    // to. Coordinate-based CDP clicks can land on an overlay / sticky header
    // instead of the intended node; without this the audit would happily
    // report another element's beacons under this element's row. Capture
    // phase + composedPath so shadow-DOM targets are seen too.
    try {
        if (!window.__qaWitnessInstalled) {
            window.__qaWitnessInstalled = true;
            window.__qaLastClick = null;
            document.addEventListener('click', (e) => {
                try {
                    const path = (e.composedPath && e.composedPath()) || [];
                    const uids = [];
                    for (const n of path) {
                        if (n && n.getAttribute) {
                            const u = n.getAttribute('data-tvuid');
                            if (u) uids.push(u);
                        }
                    }
                    const t = e.target;
                    window.__qaLastClick = {
                        uids: uids,
                        tag: (t && t.tagName) || '',
                        text: ((t && (t.innerText || t.textContent)) || '').replace(/\s+/g, ' ').trim().slice(0, 60),
                        ts: Date.now()
                    };
                } catch(err) {}
            }, true);
        }
    } catch(e) {}

    const wrapAll = () => { wrapS(); wrapU(); wrapG(); wrapDL(); wrapACDL(); };
    wrapAll();
    window.__qaTrackTimer = setInterval(wrapAll, 2000);
    return true;
})()
"""

# Read + clear the click witness. Returns what the last real click actually
# hit, so the audit can prove the intended element received the event.
READ_CLEAR_WITNESS_JS = r"""
(() => {
    const w = window.__qaLastClick || null;
    window.__qaLastClick = null;
    return w;
})()
"""

HARVEST_TRACKING_JS = r"""
(() => {
    const c = window.__qaTrackCaptures || [];
    window.__qaTrackCaptures = [];
    return c;
})()
"""

# Look at what the click pushed WITHOUT consuming it. Used right after a click
# to decide whether a network beacon is still coming (and therefore whether to
# keep the capture window open) before the real harvest runs.
PEEK_TRACKING_JS = r"""
(() => (window.__qaTrackCaptures || []).slice())()
"""

# How long a click's capture window stays open.
#   IDLE  — nothing fired at the API level; close fast, most elements are untracked.
#   MAX   — the API said an event fired, so a network beacon is on its way.
#           Tag managers commonly defer /g/collect by 5s+, and a window that
#           closes earlier pushes the beacon into the NEXT element's results.
CLICK_IDLE_WAIT = 1.8
CLICK_BEACON_MAX_WAIT = 8.0
# Drain at the very end of the page so the last elements' deferred beacons
# still land before reconciliation runs.
FINAL_DRAIN_WAIT = 9.0

# SDR validation judges each row on its own capture window — there is no
# later reconciliation pass to rescue a hit that arrived too late. It also has
# to see EVERY property a dual-tagged site sends to, or it would report an
# event as missing from a property that simply answered a few seconds later.
# So it waits longer, and never closes the window before the minimum.
#
# These are deliberately generous. A row wrongly marked Fail costs a person
# far more time to chase down than the extra seconds cost the run, and the
# observed spread on these sites runs to ~5s for the first hit with the
# second property following behind it.
# How many candidate elements a single SDR row may try before settling for
# the closest fit. Two is usually enough (header vs footer twin); three
# covers a page with a hero, a promo and a nav copy of the same control.
SDR_MAX_CANDIDATES = 3
SDR_MIN_BEACON_WAIT = 10.0
SDR_BEACON_MAX_WAIT = 20.0
# How long to let a page settle before touching it.
SDR_PAGE_SETTLE = 4.0

# Read AND clear Adobe ActivityMap's s_sq cookie. ActivityMap records the
# clicked link into s_sq at click time and only transmits it with the NEXT
# page view — since the audit blocks navigation, the cookie itself is the
# proof that the click was tracked. Cleared after reading so each click
# starts with a clean slate.
READ_CLEAR_SSQ_JS = r"""
(() => {
    const m = document.cookie.match(/(?:^|;\s*)s_sq=([^;]*)/);
    if (!m || !m[1]) return '';
    let val = '';
    try { val = decodeURIComponent(m[1]); } catch(e) { val = m[1]; }
    const host = location.hostname;
    const parts = host.split('.');
    const domains = [host, '.' + host];
    if (parts.length > 2) {
        domains.push('.' + parts.slice(-2).join('.'));
        domains.push('.' + parts.slice(-3).join('.'));
    }
    for (const d of domains) {
        document.cookie = 's_sq=; path=/; domain=' + d + '; expires=Thu, 01 Jan 1970 00:00:00 GMT';
    }
    document.cookie = 's_sq=; path=/; expires=Thu, 01 Jan 1970 00:00:00 GMT';
    return val;
})()
"""


# ===== CLICK ATTRIBUTION SUPPORT =====
# Page-level events that fire on a timer, on scroll, or on tag-manager
# bootstrap. They are NOT caused by the element under test, but they land in
# whatever capture window happens to be open — which is how a scroll event
# ends up printed under some innocent button. Filtered out of click results.
NOISE_EVENTS = {
    'scroll', 'spa_scroll', 'gtm.scrolldepth', 'user_engagement', 'page_view',
    'session_start', 'first_visit', 'gtm.js', 'gtm.dom', 'gtm.load', 'gtm.init',
    'gtm.triggergroup', 'gtm.historychange', 'gtm.timer', 'onetrustloaded',
    'optanonloaded', 'onetrustgroupsupdated', 'optanonconsentupdated',
    'gtm.scrollDepth'.lower(), 'spa_pageview', 'virtual_page_view',
    'cmp:show', 'cmp:loaded', 'consent_update',
}


def _is_noise_event(name):
    return str(name or '').strip().lower() in NOISE_EVENTS


# Parameter keys whose value identifies WHICH element a beacon belongs to.
# Every GA4 click beacon carries the clicked link's own url/text, so a beacon
# can be bound to its true element by content instead of by arrival time —
# which is what makes attribution correct even when a tag manager defers the
# beacon by several seconds into a later element's window.
_LINK_URL_KEYS = ('ep.link_url', 'ep.file_name', 'ep.outbound_url', 'link_url',
                  'gtm.elementurl', 'linkhref', 'ep.page_location')
_LINK_TEXT_KEYS = ('ep.link_text', 'ep.download_file_name', 'ep.file_name',
                   'link_text', 'gtm.elementtext', 'linktext')


def beacon_identity(ev):
    """Extract (link_url, link_text) that identify the element a tracking
    event describes. Works on our parsed GA4 event dicts and on raw
    dataLayer payloads. Returns ('','') when the event carries no element
    identity (e.g. a bare `click` outbound hit)."""
    if not isinstance(ev, dict):
        return '', ''
    params = ev.get('params') if isinstance(ev.get('params'), dict) else {}
    merged = {}
    for src in (params, ev):
        if isinstance(src, dict):
            for k, v in src.items():
                if isinstance(v, (str, int, float)):
                    merged[str(k).lower()] = str(v)
    url = ''
    for k in _LINK_URL_KEYS:
        if merged.get(k):
            url = merged[k]
            break
    txt = ''
    for k in _LINK_TEXT_KEYS:
        if merged.get(k):
            txt = merged[k]
            break
    return url.strip(), txt.strip()


def _attr_key_url(u):
    """Normalise a url for element<->beacon matching. Keeps the fragment,
    because anchor links (#isi) are distinct QA targets, but drops scheme,
    www and trailing slash so relative/absolute forms unify."""
    s = str(u or '').strip().lower()
    if not s:
        return ''
    frag = ''
    if '#' in s:
        s, frag = s.split('#', 1)
        frag = '#' + frag
    s = re.sub(r'^https?://', '', s)
    s = re.sub(r'^www\.', '', s)
    s = s.split('?')[0].rstrip('/')
    return s + frag


def _url_keys(u, base=""):
    """All the normalised forms a url might be written as, for matching.

    A beacon reports the link exactly as the markup wrote it — "#isi",
    "/science", "//cdn/x" — while discovery stored the browser-resolved
    absolute href. Comparing one form against the other silently fails, which
    leaves a genuinely tracked element looking untracked and lets its beacon
    drift onto a neighbour. Emitting both the absolute key and the
    path+fragment key lets either form match.
    """
    s = str(u or '').strip()
    if not s:
        return set()
    keys = set()
    abs_form = s
    if base:
        try:
            from urllib.parse import urljoin
            abs_form = urljoin(base, s)
        except Exception:
            abs_form = s
    for form in (s, abs_form):
        k = _attr_key_url(form)
        if k:
            keys.add(k)
            # host-less form: "host.com/path#frag" -> "/path#frag"
            m = re.match(r'^[^/#?]+\.[^/#?]+(/.*|#.*)?$', k)
            if m:
                tail = m.group(1) or '/'
                keys.add(tail if tail.startswith(('/', '#')) else '/' + tail)
            elif k.startswith(('/', '#')):
                keys.add(k)
    return {k for k in keys if k and k not in ('/', '')}


def _attr_key_text(t):
    return re.sub(r'\s+', ' ', str(t or '')).strip().lower()


_ZONE_KEYS = ('ep.link_location', 'link_location', 'ep.file_location',
              'file_location', 'linklocation', 'downloadlocation')


def beacon_zone(ev):
    """The page region a beacon says the click happened in ('header' /
    'footer' / ''). Sites that tag link_location give us a free tie-breaker
    between the same link duplicated in the header and the footer — without
    it, one twin absorbs the other's events."""
    if not isinstance(ev, dict):
        return ''
    params = ev.get('params') if isinstance(ev.get('params'), dict) else {}
    for src in (params, ev):
        if not isinstance(src, dict):
            continue
        for k, v in src.items():
            if str(k).lower() in _ZONE_KEYS and isinstance(v, str) and v.strip():
                z = v.strip().lower()
                if 'header' in z or 'nav' in z:
                    return 'header'
                if 'footer' in z:
                    return 'footer'
                if 'body' in z or 'main' in z or 'content' in z:
                    return 'body'
    return ''


# Params that change on every hit and say nothing about the event itself.
# Excluded when deciding whether two hits are "the same event".
_VOLATILE_PARAMS = {'hit_timestamp', 'ep.hit_timestamp', '_et', 'et',
                    'client_id_config', 'ep.client_id_config', '_p', 'seg',
                    '_s', 'sid', 'sct', 'ep.page_url', 'page_url'}


def consolidate_ga4_events(events):
    """Collapse hits that are the same event sent to several GA4 properties.

    Sites commonly dual-tag: one click fires `exit_link` to two measurement
    IDs, so a raw list reads "exit_link, exit_link, exit_link" and looks like
    a bug. Group by event name + meaningful params, and report the property
    IDs together, so each row shows what actually happened once.
    """
    grouped = []
    index = {}
    for ev in events or []:
        if not isinstance(ev, dict):
            continue
        name = str(ev.get("event") or "")
        params = ev.get("params") if isinstance(ev.get("params"), dict) else {}
        sig_items = tuple(sorted(
            (str(k), str(v)) for k, v in params.items()
            if str(k).lower() not in _VOLATILE_PARAMS
        ))
        mid = str(ev.get("measurement_id") or "")
        # API-level captures (dataLayer/ACDL/gtag) are a different kind of
        # evidence than a network beacon — never merge the two together.
        is_api = mid.startswith("(")
        key = (name, sig_items, is_api)
        if key in index:
            slot = grouped[index[key]]
            if mid and mid not in slot["_ids"]:
                slot["_ids"].append(mid)
            slot["hit_count"] += 1
            continue
        index[key] = len(grouped)
        grouped.append({
            "event": name,
            "_ids": [mid] if mid else [],
            "params": dict(params),
            "hit_count": 1,
        })

    out = []
    for g in grouped:
        ids = g.pop("_ids")
        g["measurement_ids"] = ids
        # Keep `measurement_id` a plain string so existing report/Excel
        # rendering keeps working unchanged.
        g["measurement_id"] = " + ".join(ids) if ids else ""
        out.append(g)
    return out


def _norm_url(u):
    u = str(u or '').strip().lower()
    u = re.sub(r'^https?://', '', u)
    u = re.sub(r'^www\.', '', u)
    u = u.split('?')[0].split('#')[0]
    return u.strip('/')

def clean_compare_text(text1, text2):
    t1 = _norm_str(text1)
    t2 = _norm_str(text2)
    if not t1 or not t2:
        return False
    if t1 == t2:
        return True
    if len(t1) > 3 and len(t2) > 3:
        if t1 in t2 or t2 in t1:
            return True
    return False

def clean_compare_url(url1, url2):
    u1 = _norm_url(url1)
    u2 = _norm_url(url2)
    if not u1 or not u2:
        return False
    if u1 == u2:
        return True
    if u1 in u2 or u2 in u1:
        return True
    return False

def extract_event_identifiers(ev, ev_type):
    texts = set()
    urls = set()
    
    if ev_type == "ga4":
        if ev.get("event"):
            texts.add(ev["event"])
        for k, v in ev.get("params", {}).items():
            val_str = str(v).strip()
            if not val_str:
                continue
            if val_str.startswith(("http://", "https://", "/", "www.")):
                urls.add(val_str)
            else:
                texts.add(val_str)
    elif ev_type == "adobe":
        if ev.get("link_name"):
            texts.add(ev["link_name"])
        if ev.get("link_url"):
            urls.add(ev["link_url"])
        for k, v in ev.get("evars", {}).items():
            val_str = str(v).strip()
            if val_str.startswith(("http://", "https://", "/", "www.")):
                urls.add(val_str)
            else:
                texts.add(val_str)
        for k, v in ev.get("props", {}).items():
            val_str = str(v).strip()
            if val_str.startswith(("http://", "https://", "/", "www.")):
                urls.add(val_str)
            else:
                texts.add(val_str)
    elif ev_type == "adobe_websdk":
        if ev.get("interaction_name"):
            texts.add(ev["interaction_name"])
        if ev.get("interaction_url"):
            urls.add(ev["interaction_url"])
            
    return texts, urls


async def validate_clicks(browser, url, index, total):
    """Click-level analytics event audit for a single URL.
    Clicks each interactive element and captures GA4 events + Adobe link calls."""
    sys.stdout.write(f"[{index}/{total}] Click audit: {url}\n")
    sys.stdout.flush()

    context = None
    elements_result = []
    try:
        context = await browser.new_context(
            viewport={'width': 1280, 'height': 800},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
        )

        # Force the consent scenario via OneTrust cookies (no-op on non-OT sites)
        dom = _cookie_domain_for(url)
        if dom:
            now = datetime.datetime.utcnow()
            ts = now.strftime("%a+%b+%d+%Y+%H:%M:%S+GMT+0000")
            optanon = (
                f"isGpcEnabled=0&datestamp={ts}&version=202401.1.0&isIABGlobal=false"
                f"&hosts=&consentId=00000000-0000-0000-0000-000000000000"
                f"&interactionCount=1&landingPath=NotLandingPage"
                f"&groups={SCENARIO_GROUPS['Accept All']}&AwaitingReconsent=false"
            )
            try:
                await context.add_cookies([
                    {"name": "OptanonConsent", "value": optanon, "domain": dom, "path": "/"},
                    {"name": "OptanonAlertBoxClosed",
                     "value": now.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                     "domain": dom, "path": "/"},
                ])
            except Exception:
                pass

        page = await context.new_page()
        await stealth_obj.apply_stealth_async(page)

        # Set up CDP for request capture
        cdp = await context.new_cdp_session(page)
        await cdp.send("Network.enable")

        # PERSISTENT request log attached BEFORE navigation so we never miss a
        # beacon. We capture via THREE independent paths (CDP + Playwright
        # context-level + page-level) and dedup by url+ts. Attribution later
        # uses per-click time windows, so requests fired during page load /
        # discovery simply won't match any click — harmless.
        all_requests = []
        seen_req_keys = set()

        # Which element's click window is open right now. Every captured
        # request is stamped with it, so attribution is an explicit label
        # rather than an array slice that a late beacon can slide out of.
        # -1 means "no click in progress" (page load, menu expansion, idle).
        current_cid = {"v": -1}

        def _push_req(req_url, post, ts):
            if not req_url:
                return
            key = (req_url, round(ts, 2))
            if key in seen_req_keys:
                return
            seen_req_keys.add(key)
            all_requests.append({"url": req_url, "post": post or "", "ts": ts,
                                 "cid": current_cid["v"]})

        def on_req_persistent(params):
            try:
                req = params.get("request", {}) or {}
                _push_req(req.get("url", ""), req.get("postData", "") or "", time.time())
            except Exception:
                pass
        cdp.on("Network.requestWillBeSent", on_req_persistent)

        def on_pw_request(request):
            try:
                pd = ""
                try: pd = request.post_data or ""
                except: pass
                _push_req(request.url, pd, time.time())
            except Exception:
                pass
        # Context-level catches requests from ALL pages/popups/frames; page
        # level is a belt-and-suspenders backup.
        try: context.on("request", on_pw_request)
        except Exception: pass
        try: page.on("request", on_pw_request)
        except Exception: pass

        # Load page and settle
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception:
            pass
        await asyncio.sleep(2)
        await accept_cookies(page)
        try:
            await page.wait_for_load_state("networkidle", timeout=12000)
        except:
            pass
        await asyncio.sleep(3)

        # Smooth scroll down and up to trigger lazy-loaded elements and dynamic content
        try:
            sys.stdout.write(f"[{index}/{total}] Scrolling page to load lazy content...\n")
            sys.stdout.flush()
            await page.evaluate("""async () => {
                await new Promise((resolve) => {
                    let totalHeight = 0;
                    const distance = 400;
                    const timer = setInterval(() => {
                        const scrollHeight = document.body.scrollHeight;
                        window.scrollBy(0, distance);
                        totalHeight += distance;
                        if(totalHeight >= scrollHeight || totalHeight > 10000){
                            clearInterval(timer);
                            resolve();
                        }
                    }, 80);
                });
                window.scrollTo(0, 0);
            }""")
            await asyncio.sleep(2)
        except Exception:
            pass

        # Block navigation ONLY during menu expansion phase so menu-opener
        # clicks don't accidentally navigate away during discovery.
        try:
            await page.evaluate(BLOCK_NAVIGATION_JS)
        except Exception:
            pass

        # 1) Expose collapsed menus / accordions / dropdowns (JS internal rounds).
        try:
            revealed = await page.evaluate(EXPOSE_HIDDEN_JS)
            if revealed:
                sys.stdout.write(f"[{index}/{total}] Revealed {revealed} hidden menu/collapsed elements\n")
                sys.stdout.flush()
                await asyncio.sleep(0.8)
        except Exception:
            pass

        # 2) Multi-pass discovery with per-nav-item hover. CSS :hover only
        #    keeps ONE menu open at a time (the one mouse is over), so we
        #    hover each top-nav item one by one and discover its submenu
        #    items while it's open. Accumulate across passes with Python-
        #    side dedup so duplicates (footer/header cookie links etc.)
        #    only count once.
        elements_by_key = {}
        def _add_batch(batch):
            for el in batch:
                t = (el.get("text") or "").strip().lower()
                t = re.sub(r'\s+', ' ', t)
                h = (el.get("href") or "").strip().lower()
                if t:
                    key = "T:" + t + "|H:" + h
                else:
                    key = (el.get("tag", "") + "|H:" + h + "|I:" + (el.get("id") or ""))
                # Zone + frame are part of identity: the same link in header
                # AND footer (or inside an iframe) is a separate QA case.
                key += "|Z:" + (el.get("zone") or "body") + "|F:" + (el.get("frame_url") or "")
                el["fp"] = key          # stable fingerprint, survives re-stamping
                if key not in elements_by_key:
                    elements_by_key[key] = el
                else:
                    # Keep the richest record: a later pass may see the element
                    # while it is visible (real uid + accurate zone) where the
                    # first pass caught it collapsed.
                    prev = elements_by_key[key]
                    if not prev.get("uid") and el.get("uid"):
                        prev["uid"] = el["uid"]

        # Pass A: initial discovery (default page state, no special hovers).
        try: _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
        except Exception: pass

        # Pass B: hover each top-nav item ONE AT A TIME, brief settle, then
        # discover what's now visible — sub-menus revealed under that item.
        # While that menu is open, DRILL DEEPER: hover any submenu items that
        # look like parents themselves so nested flyouts (sub-sub-menus)
        # render and get discovered too. Time-budgeted so a pathological
        # mega-menu can't stall the whole audit.
        try:
            nav_locator = page.locator(
                'header > * a, header > * button, header nav a, header nav button, '
                'nav > a, nav > button, nav > ul > li, nav > ul > li > a, nav > ul > li > button, '
                '[role="menuitem"], [class*="has-submenu"], [class*="has-children"]'
            )
            sub_parent_locator = page.locator(
                '[aria-haspopup="true"], [class*="has-submenu"], [class*="has-children"], '
                'li:has(> ul) > a, li:has(> div ul) > a, [class*="dropdown"] > a, '
                '[class*="flyout"] > a, li[class*="parent"] > a'
            )
            nav_count = min(await nav_locator.count(), 80)
            passb_deadline = time.time() + 150   # hard budget for pass B
            for ni in range(nav_count):
                if time.time() > passb_deadline:
                    break
                try:
                    await nav_locator.nth(ni).hover(timeout=400, force=True, no_wait_after=True)
                    await asyncio.sleep(0.35)        # let sub-menu render
                    before = len(elements_by_key)
                    _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
                    # Only drill deeper when this hover actually revealed
                    # something new — otherwise it's not a menu parent.
                    if len(elements_by_key) > before:
                        try:
                            sub_count = min(await sub_parent_locator.count(), 20)
                            for si in range(sub_count):
                                if time.time() > passb_deadline:
                                    break
                                try:
                                    s_el = sub_parent_locator.nth(si)
                                    if not await s_el.is_visible(timeout=120):
                                        continue
                                    await s_el.hover(timeout=300, force=True, no_wait_after=True)
                                    await asyncio.sleep(0.25)   # let flyout render
                                    _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
                                except Exception:
                                    pass
                        except Exception:
                            pass
                except Exception:
                    pass
            # Move mouse off so footer hover-menus get a chance too.
            try: await page.mouse.move(10, 600)
            except: pass
            await asyncio.sleep(0.3)
        except Exception:
            pass

        # Pass C: one more expose (in case hovers triggered new collapsibles).
        try:
            await page.evaluate(EXPOSE_HIDDEN_JS)
            await asyncio.sleep(0.4)
            _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
        except Exception:
            pass

        # Pass D: AGGRESSIVE menu drill. Many sites (Fresenius and other
        # enterprise mega-menus) open their menus on CLICK, not hover. Find
        # every element that looks like a menu opener (hamburger / burger /
        # aria-haspopup / aria-expanded=false), click it, discover what's
        # newly visible, then drill one level deeper into any "has children"
        # parents inside the opened menu.
        opener_selectors = [
            'button[aria-label*="menu" i]',
            'button[aria-label*="navigation" i]',
            'button[aria-label*="open" i]',
            'button[class*="burger" i]',
            'button[class*="hamburger" i]',
            'button[class*="menu-toggle" i]',
            'button[class*="nav-toggle" i]',
            'button[class*="navigation-toggle" i]',
            'button[class*="navbar-toggle" i]',
            '[aria-haspopup="menu"]',
            '[aria-haspopup="true"]',
            '[data-toggle="menu"]',
            '[data-bs-toggle="offcanvas"]',
            'button[aria-expanded="false"]',
            '[role="button"][aria-controls]',
        ]
        sub_selectors = [
            '[class*="has-children"]:not([class*="has-children--hidden"])',
            '[class*="has-submenu"]',
            '[class*="parent"][role="menuitem"]',
            'li[class*="menu-item"] > a',
            'li[class*="parent"] > a',
            'li[class*="parent"] > button',
            '[aria-haspopup="true"]:not([aria-expanded="true"])',
        ]
        opened_count = 0
        for sel in opener_selectors:
            try:
                count = min(await page.locator(sel).count(), 12)
            except Exception:
                count = 0
            for i in range(count):
                try:
                    el = page.locator(sel).nth(i)
                    if not await el.is_visible(timeout=300):
                        continue
                    # Never open the cookie/consent UI while hunting for
                    # menus: "Cookies Settings" matches these opener
                    # selectors, and the preference center it opens covers
                    # the header for the rest of the audit.
                    try:
                        if await el.evaluate("""e => {
                            let n = e;
                            while (n && n !== document.body) {
                                const id = (n.id || '').toLowerCase();
                                const cls = (typeof n.className === 'string' ? n.className : '').toLowerCase();
                                if (/cookie|consent|onetrust|optanon|^ot-|privacy|truste/.test(id)
                                    || /cookie|consent|onetrust|optanon|ot-pc|ot-sdk|privacy|truste/.test(cls)) {
                                    return true;
                                }
                                n = n.parentElement;
                            }
                            const t = (e.innerText || e.getAttribute('aria-label') || '').toLowerCase();
                            return /cookie|consent|privacy choices|manage choices/.test(t);
                        }"""):
                            continue
                    except Exception:
                        pass
                    await el.click(force=True, timeout=900, no_wait_after=True)
                    await asyncio.sleep(0.55)   # let overlay/menu render
                    opened_count += 1
                    # Discover what just became visible
                    try: _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
                    except: pass
                    # RECURSIVE drill into nested submenus. Each round clicks
                    # every visible "has children" parent (which reveals its
                    # children — possibly more parents), then re-discovers.
                    # Repeat until a round reveals no new clickable elements
                    # (or we hit the round cap). This catches 3rd/4th-level
                    # mega-menus where a submenu opens yet another submenu.
                    # We track which sub-parents we've already clicked (by a
                    # cheap signature) so we don't toggle the same one shut.
                    clicked_subs = set()
                    for _round in range(8):
                        before = len(elements_by_key)
                        round_clicked = 0
                        for ssel in sub_selectors:
                            try:
                                sc = min(await page.locator(ssel).count(), 40)
                            except Exception:
                                sc = 0
                            for j in range(sc):
                                try:
                                    s_el = page.locator(ssel).nth(j)
                                    if not await s_el.is_visible(timeout=150):
                                        continue
                                    # Skip parents already expanded / clicked.
                                    try:
                                        sig = await s_el.evaluate(
                                            "e => (e.tagName||'')+'|'+(e.textContent||'').trim().slice(0,40)+'|'+(e.getAttribute('href')||'')")
                                    except Exception:
                                        sig = f"{ssel}#{j}"
                                    if sig in clicked_subs:
                                        continue
                                    try:
                                        if (await s_el.get_attribute("aria-expanded")) == "true":
                                            clicked_subs.add(sig)
                                            continue
                                    except Exception:
                                        pass
                                    clicked_subs.add(sig)
                                    await s_el.click(force=True, timeout=600, no_wait_after=True)
                                    round_clicked += 1
                                    await asyncio.sleep(0.25)
                                    _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
                                except Exception:
                                    pass
                        after = len(elements_by_key)
                        # Stop when a full round added nothing new and clicked
                        # nothing new — the menu tree is fully expanded.
                        if after == before and round_clicked == 0:
                            break
                    # Close any open overlay/dialog so the next opener can
                    # start from a clean state. Escape works for most CMP /
                    # mega-menu implementations.
                    try: await page.keyboard.press("Escape")
                    except: pass
                    await asyncio.sleep(0.2)
                except Exception:
                    pass
        if opened_count:
            sys.stdout.write(f"[{index}/{total}] Drilled into {opened_count} menu opener(s)\n")
            sys.stdout.flush()

        # Final discovery pass after all drilling.
        try: _add_batch(await page.evaluate(DISCOVER_CLICKABLES_JS))
        except Exception: pass

        # Same-origin iframe discovery — video embeds, form widgets and CMP
        # iframes contain clickables too. Cross-origin frames can't be read
        # (browser security), those throw and are skipped harmlessly.
        try:
            for fr in page.frames:
                if fr == page.main_frame:
                    continue
                try:
                    batch = await fr.evaluate(DISCOVER_CLICKABLES_JS)
                    for b in batch:
                        b["frame_url"] = fr.url
                    if batch:
                        sys.stdout.write(f"[{index}/{total}] Found {len(batch)} clickable(s) inside iframe {fr.url[:60]}\n")
                        sys.stdout.flush()
                    _add_batch(batch)
                except Exception:
                    pass
        except Exception:
            pass

        # KEEP navigation blocked during the click loop. preventDefault only
        # stops the browser-default navigation — GA4 / GTM / Adobe click
        # handlers still run in the same event tick (s.tl fires on the click
        # itself, not on the navigation). Staying on the page keeps expanded
        # menus intact, skips a full page reload per link (3-4x faster), and
        # prevents the failure where one bad navigation kills every element
        # after it (observed: 255/302 skipped after one nav-back failed).
        # JS-driven redirects (window.location=...) can still escape the
        # blocker — the per-element health check below recovers from those.
        try:
            await page.evaluate(BLOCK_NAVIGATION_JS)
        except Exception:
            pass

        # Instrument tracking APIs (s.tl / utag.link / gtag / dataLayer) and
        # clear any ActivityMap s_sq left over from the discovery clicks so
        # the first audited element starts clean.
        try:
            await page.evaluate(INSTRUMENT_TRACKING_JS)
            await page.evaluate(READ_CLEAR_SSQ_JS)
            await page.evaluate(HARVEST_TRACKING_JS)
        except Exception:
            pass

        # Discovery may have left a consent dialog open on top of the page.
        try:
            _cc = await page.evaluate(CLOSE_CONSENT_UI_JS)
            if _cc:
                sys.stdout.write(f"[{index}/{total}] Dismissed {_cc} consent overlay(s) left open by discovery\n")
                sys.stdout.flush()
                await asyncio.sleep(0.4)
        except Exception:
            pass

        elements = list(elements_by_key.values())

        # Order: header first, then footer, then body.
        zone_rank = {"header": 0, "footer": 1, "body": 2}
        elements.sort(key=lambda e: zone_rank.get(e.get("zone", "body"), 2))

        if elements:
            by_zone = {}
            for e in elements:
                by_zone[e.get("zone", "body")] = by_zone.get(e.get("zone", "body"), 0) + 1
            zone_summary = ", ".join(f"{z}={by_zone[z]}" for z in ("header", "footer", "body") if z in by_zone)
            sys.stdout.write(f"[{index}/{total}] Found {len(elements)} clickable elements ({zone_summary})\n")
        else:
            sys.stdout.write(f"[{index}/{total}] Found 0 clickable elements on {url}\n")
        sys.stdout.flush()

        original_url = page.url

        async def _ensure_on_page():
            """Recover the audit page if a click escaped the nav blocker.
            Re-creates the page (with CDP + listeners) if it crashed/closed,
            re-navigates with retries if the URL drifted. Returns True if a
            reload happened (menus will be collapsed again)."""
            nonlocal page, cdp
            try:
                if page.is_closed():
                    page = await context.new_page()
                    try: page.on("request", on_pw_request)
                    except Exception: pass
                    cdp = await context.new_cdp_session(page)
                    await cdp.send("Network.enable")
                    cdp.on("Network.requestWillBeSent", on_req_persistent)
            except Exception:
                pass
            cur = ""
            try:
                cur = page.url
            except Exception:
                pass
            if _norm_url(cur) == _norm_url(original_url):
                return False
            for attempt in range(3):
                try:
                    await page.goto(original_url, wait_until="domcontentloaded", timeout=25000)
                    try:
                        await page.wait_for_load_state("networkidle", timeout=6000)
                    except Exception:
                        pass
                    await asyncio.sleep(1.0)
                    break
                except Exception:
                    await asyncio.sleep(2.0 * (attempt + 1))
            try:
                await page.evaluate(BLOCK_NAVIGATION_JS)
            except Exception:
                pass
            # A reload wipes the API wraps — re-instrument.
            try:
                await page.evaluate(INSTRUMENT_TRACKING_JS)
                await page.evaluate(READ_CLEAR_SSQ_JS)
                await page.evaluate(HARVEST_TRACKING_JS)
            except Exception:
                pass
            # A reload also wipes the data-tvuid stamps, which are how the
            # click loop guarantees it clicks the intended node. Re-run
            # discovery to re-stamp, then re-point each pending element at
            # its NEW uid via the fingerprint (which is reload-stable).
            try:
                fresh = await page.evaluate(DISCOVER_CLICKABLES_JS)
                fp_to_uid = {}
                for f in fresh or []:
                    t = re.sub(r'\s+', ' ', (f.get("text") or "").strip().lower())
                    h = (f.get("href") or "").strip().lower()
                    if t:
                        k = "T:" + t + "|H:" + h
                    else:
                        k = (f.get("tag", "") + "|H:" + h + "|I:" + (f.get("id") or ""))
                    k += "|Z:" + (f.get("zone") or "body") + "|F:" + (f.get("frame_url") or "")
                    if f.get("uid"):
                        fp_to_uid[k] = f["uid"]
                for pending in elements:
                    nu = fp_to_uid.get(pending.get("fp"))
                    if nu:
                        pending["uid"] = nu
            except Exception:
                pass
            return True

        # Request capture listeners were attached before navigation (see top of
        # this function). Clear out everything captured during page load /
        # discovery so the click loop starts from a clean slate — keeps the
        # DEBUG host dump and attribution focused on click-driven beacons.
        all_requests.clear()
        seen_req_keys.clear()

        elements_result = []
        for elem in elements:
            elements_result.append({
                "element": {
                    "selector": elem.get("selector", ""),
                    "tag": elem.get("tag", ""),
                    "text": elem.get("text", ""),
                    "href": elem.get("href", ""),
                    "id": elem.get("id", ""),
                    "zone": elem.get("zone", "body"),
                    "frame_url": elem.get("frame_url", ""),
                    "uid": elem.get("uid", ""),
                    "is_download": bool(elem.get("is_download")),
                },
                "ga4_events": [],
                "adobe_calls": [],
                "adobe_websdk": [],
                "other_analytics": [],
                "datalayer_events": [],   # instant, click-time ground truth
                "network_requests": [],   # RAW capture: every request in this click's window
                "network_count": 0,
                "has_tracking": False,
                "skipped": False,
                "skip_reason": "",
                "click_method": "",
                "click_verified": False,  # witness confirmed the intended node was hit
                "blocked_by": "",         # element covering this one, if any
                "click_method_note": "",
                "attribution": "",        # how the events were bound: witness/content/window
                "total_requests": 0
            })

        GENERIC_VENDORS = [
            ("Segment",   ["api.segment.io/v1/t", "api.segment.io/v1/i", "cdp.segment.io"]),
            ("Tealium",   ["collect.tealiumiq.com/event", "collect.tealiumiq.com/i"]),
            ("Mixpanel",  ["api.mixpanel.com/track", "api-js.mixpanel.com/track"]),
            ("Heap",      ["heapanalytics.com/api/track", "heapanalytics.com/h"]),
            ("Amplitude", ["api.amplitude.com/2/httpapi", "api2.amplitude.com", "api.eu.amplitude.com"]),
            ("Hotjar",    ["vc.hotjar.io", "in.hotjar.com/api/v"]),
            ("FullStory", ["fullstory.com/rec/event", "rs.fullstory.com"]),
            ("MS Clarity",["clarity.ms/collect", "clarity.ms/eus2"]),
            ("Pendo",     ["data.pendo.io/data"]),
            ("Quantum Metric", ["cdn.quantummetric.com/qtm", "qm-record"]),
            ("Snowplow",  ["/com.snowplowanalytics.snowplow/", "/snowplow/track"]),
            # Ad networks: a click that fires one of these is still a tracked
            # click, and the report should say whose tag it was.
            ("Meta / Facebook", ["facebook.com/tr", "connect.facebook.net/signals"]),
            ("Google Ads",      ["googleadservices.com/pagead/conversion",
                                 "google.com/pagead/1p-user-list", "/pagead/viewthroughconversion"]),
            ("DoubleClick",     ["doubleclick.net/activity", "ad.doubleclick.net"]),
            ("LinkedIn",        ["px.ads.linkedin.com", "snap.licdn.com/li.lms-analytics"]),
            ("TikTok",          ["analytics.tiktok.com", "business-api.tiktok.com/track"]),
            ("X / Twitter",     ["static.ads-twitter.com", "analytics.twitter.com"]),
            ("Pinterest",       ["ct.pinterest.com"]),
            ("Snapchat",        ["tr.snapchat.com"]),
            ("Reddit",          ["pixel.reddit.com", "alb.reddit.com"]),
            ("Bing / MS Ads",   ["bat.bing.com/bat", "bat.bing.com/action"]),
            ("Criteo",          ["static.criteo.net", "sslwidget.criteo.com"]),
            ("The Trade Desk",  ["insight.adsrvr.org"]),
        ]

        def _parse_click_requests(reqs, el_res):
            """Parse a list of requests and assign analytics events to el_res."""
            for req in reqs:
                req_low = req["url"].lower()

                # --- GA4 ---
                if "/g/collect" in req_low or (
                    ("google-analytics.com" in req_low or "analytics.google.com" in req_low)
                    and "collect" in req_low
                ):
                    ga4_list = parse_ga4_event(req["url"], req["post"])
                    for ev in ga4_list:
                        # Drop page-level noise (scroll depth, engagement pings,
                        # page_view). These fire on timers and would otherwise be
                        # printed under whichever button happened to be under test.
                        if ev.get("event") and not _is_noise_event(ev["event"]):
                            el_res["ga4_events"].append(ev)
                            el_res["has_tracking"] = True
                            el_res["total_requests"] += 1

                # --- Adobe Classic (/b/ss/) ---
                # Only keep click-level beacons (lnk_o/lnk_d/lnk_e = s.tl calls)
                # Skip page_view (s.t) — those are destination page noise.
                elif "/b/ss/" in req_low:
                    adobe = parse_adobe_link_call(req["url"], req["post"])
                    if adobe and adobe.get("link_type") != "page_view":
                        el_res["adobe_calls"].append(adobe)
                        el_res["has_tracking"] = True
                        el_res["total_requests"] += 1

                # --- Adobe Web SDK (adobedc / omtrdc / demdex) ---
                elif (("adobedc.net" in req_low or "omtrdc.net" in req_low
                       or "demdex.net" in req_low or "2o7.net" in req_low)
                      and req["post"]):
                    ws_events = parse_adobe_websdk_event(req["post"])
                    for w in ws_events:
                        el_res["adobe_websdk"].append(w)
                        el_res["has_tracking"] = True
                        el_res["total_requests"] += 1

                # --- Other vendors ---
                else:
                    matched_vendor = None
                    for vname, sigs in GENERIC_VENDORS:
                        if any(s in req_low for s in sigs):
                            matched_vendor = vname
                            break
                    if matched_vendor:
                        if not any(o["vendor"] == matched_vendor for o in el_res["other_analytics"]):
                            el_res["other_analytics"].append({"vendor": matched_vendor, "url": req["url"][:200]})
                            el_res["has_tracking"] = True
                            el_res["total_requests"] += 1

        # ---- PER-CLICK CAPTURE LOOP ----
        # For each element: bookmark request count → click → wait → diff.
        # Requests that arrive between bookmark and post-wait are directly
        # attributed to this element. No heuristics, no timing guesswork.
        for ei, elem in enumerate(elements):
            el_res = elements_result[ei]
            clicked = False
            click_method = ""
            el = None

            # HEALTH CHECK: if a previous click escaped the nav blocker
            # (JS redirect / window.open), recover before touching locators —
            # otherwise every remaining element fails on a dead page.
            try:
                await _ensure_on_page()
            except Exception:
                pass

            # Resolve the target frame — elements discovered inside a
            # same-origin iframe must be located through that frame.
            target = page
            if elem.get("frame_url"):
                try:
                    for fr in page.frames:
                        if fr.url == elem["frame_url"]:
                            target = fr
                            break
                except Exception:
                    pass

            # PRIMARY: the uid stamped onto the node at discovery time. This
            # is a 1:1 handle — no ambiguity, no .first picking a different
            # node, and header/footer twins that share text+href stay
            # distinct. Everything below is fallback for when a reload wiped
            # the stamps before re-stamping could run.
            uid = elem.get("uid") or ""
            if uid:
                try:
                    cand = target.locator(f'[data-tvuid="{uid}"]')
                    if await cand.count() > 0:
                        el = cand.first
                except Exception:
                    pass

            # FALLBACK 0: the discovery selector.
            if el is None:
                try:
                    cand = target.locator(elem["selector"]).first
                    if await cand.count() > 0:
                        el = cand
                except Exception:
                    pass
            # FALLBACK 1: locate by href — selectors can go stale, but the
            # href is stable. NOTE: el.href (what discovery stored) is the
            # ABSOLUTE url while the DOM attribute is often RELATIVE, and
            # CSS [href=...] matches the raw attribute — so try the absolute
            # form, the path-only form, and an ends-with match.
            if el is None and (elem.get("href") or "").startswith("http"):
                try:
                    from urllib.parse import urlparse as _hp
                    full = elem["href"]
                    path = _hp(full).path or ""
                    cand_sels = [f'a[href="{full.replace(chr(34), "")}"]']
                    if path and path != "/":
                        cand_sels.append(f'a[href="{path}"]')
                        cand_sels.append(f'a[href$="{path}"]')
                    for cs in cand_sels:
                        try:
                            cand = target.locator(cs).first
                            if await cand.count() > 0:
                                el = cand
                                break
                        except Exception:
                            pass
                except Exception:
                    pass
            # FALLBACK 2: locate by exact visible text.
            if el is None:
                txt = (elem.get("text") or "").strip()
                if txt and not txt.startswith("["):
                    try:
                        cand = target.get_by_text(txt, exact=True).first
                        if await cand.count() > 0:
                            el = cand
                    except Exception:
                        pass

            if el is not None:
                # Check if element is visible — after navigating back,
                # submenus will be collapsed. Expand parent menus first.
                try:
                    is_visible = await el.is_visible(timeout=500)
                except Exception:
                    is_visible = False

                if not is_visible:
                    # Re-open collapsed menus/accordions first (they re-collapse
                    # after every navigate-back), then walk up the DOM tree.
                    try:
                        await page.evaluate(EXPOSE_HIDDEN_JS)
                        await asyncio.sleep(0.4)
                    except Exception:
                        pass
                    # JS: walk up DOM tree, expand any collapsed parent menus
                    try:
                        await el.evaluate("""e => {
                            let node = e;
                            const toExpand = [];
                            while (node && node !== document.body) {
                                node = node.parentElement;
                                if (!node) break;
                                const tag = node.tagName || '';
                                const expanded = node.getAttribute('aria-expanded');
                                const hasPopup = node.getAttribute('aria-haspopup');
                                if (tag === 'LI' || tag === 'DIV' || hasPopup) {
                                    if (expanded === 'false' || hasPopup) {
                                        toExpand.unshift(node);
                                    }
                                }
                            }
                            for (const p of toExpand) {
                                const trigger = p.querySelector(':scope > a, :scope > button, :scope > [role="button"]');
                                const t = trigger || p;
                                t.dispatchEvent(new MouseEvent('mouseenter', { bubbles: true }));
                                t.dispatchEvent(new MouseEvent('mouseover', { bubbles: true }));
                                t.click();
                            }
                        }""")
                        await asyncio.sleep(0.5)
                    except Exception:
                        pass

                    # Playwright hover on parent nav items
                    try:
                        parent_sel = elem.get("selector", "")
                        # Extract top-level nav parent (e.g. first li in the selector)
                        parts = parent_sel.split(" > ")
                        for depth in range(2, min(len(parts), 5)):
                            partial = " > ".join(parts[:depth])
                            try:
                                parent_loc = page.locator(partial).first
                                if await parent_loc.is_visible(timeout=300):
                                    await parent_loc.hover(force=True, timeout=500)
                                    await asyncio.sleep(0.3)
                            except Exception:
                                pass
                    except Exception:
                        pass

                    # Re-check visibility
                    try:
                        is_visible = await el.is_visible(timeout=500)
                    except Exception:
                        is_visible = False

                try:
                    await el.scroll_into_view_if_needed(timeout=1500)
                except Exception:
                    pass
                await asyncio.sleep(0.35)

                # ---- DRAIN before the click ----
                # Everything above (menu re-expansion, parent hovers, JS
                # .click() on collapsed parents, scrolling) fires the site's
                # OWN tracking. Previously the capture window opened before
                # all that, so a parent menu item's beacons were reported
                # under its child. Open the window only now, and throw away
                # anything the expansion produced.
                try:
                    await page.evaluate(HARVEST_TRACKING_JS)     # discard
                    await page.evaluate(READ_CLEAR_WITNESS_JS)   # discard
                    await page.evaluate(READ_CLEAR_SSQ_JS)       # discard
                except Exception:
                    pass
                await asyncio.sleep(0.15)
                req_bookmark = len(all_requests)
                current_cid["v"] = ei

                # Resolve a click point that actually HITS this element.
                # A centre-point CDP click is the highest-fidelity input
                # (isTrusted:true, which Adobe ActivityMap and some GTM
                # triggers require) but it is coordinate-based: a sticky
                # header, cookie bar or overlay sitting on top means the
                # click lands on a DIFFERENT element while the audit records
                # it as a success. Hit-test first, and only take the trusted
                # path when the topmost node at that point really is ours.
                hit = None
                try:
                    hit = await el.evaluate("""(e) => {
                        const r = e.getBoundingClientRect();
                        if (!r.width || !r.height) return {ok: false, reason: 'no-box'};
                        const cands = [
                            [r.x + r.width / 2,   r.y + r.height / 2],
                            [r.x + r.width * 0.25, r.y + r.height / 2],
                            [r.x + r.width * 0.75, r.y + r.height / 2],
                            [r.x + Math.min(6, r.width / 2), r.y + Math.min(6, r.height / 2)],
                        ];
                        for (const [x, y] of cands) {
                            if (x < 0 || y < 0 || x > innerWidth || y > innerHeight) continue;
                            const top = document.elementFromPoint(x, y);
                            if (top && (top === e || e.contains(top) || top.contains(e))) {
                                return {ok: true, x: x, y: y};
                            }
                        }
                        return {ok: false, reason: 'obscured'};
                    }""")
                except Exception:
                    hit = None

                # If something is covering the element, try to clear it before
                # falling back. A previous element in this very loop is the
                # usual culprit: clicking a "More from ..." style toggle
                # leaves its panel open on top of the links that come next.
                if not (hit and hit.get("ok")):
                    try:
                        await page.evaluate(CLOSE_CONSENT_UI_JS)
                        await page.keyboard.press("Escape")
                        await asyncio.sleep(0.25)
                        await el.scroll_into_view_if_needed(timeout=1200)
                        await asyncio.sleep(0.2)
                        hit = await el.evaluate("""(e) => {
                            const r = e.getBoundingClientRect();
                            if (!r.width || !r.height) return {ok: false, reason: 'no-box'};
                            const cands = [
                                [r.x + r.width / 2,   r.y + r.height / 2],
                                [r.x + r.width * 0.25, r.y + r.height / 2],
                                [r.x + r.width * 0.75, r.y + r.height / 2],
                            ];
                            let blocker = '';
                            for (const [x, y] of cands) {
                                if (x < 0 || y < 0 || x > innerWidth || y > innerHeight) continue;
                                const top = document.elementFromPoint(x, y);
                                if (top && (top === e || e.contains(top) || top.contains(e))) {
                                    return {ok: true, x: x, y: y};
                                }
                                if (top && !blocker) {
                                    blocker = top.tagName + (top.id ? '#' + top.id : '')
                                            + '.' + String(top.className || '').slice(0, 50);
                                }
                            }
                            return {ok: false, reason: 'obscured', blocker: blocker};
                        }""")
                    except Exception:
                        pass
                    el_res_blocked = (hit or {}).get("blocker") or (hit or {}).get("reason") or ""
                    if hit and not hit.get("ok") and hit.get("blocker"):
                        el_res["blocked_by"] = hit["blocker"]

                # Still covered: neutralise the covering elements so the
                # trusted click can be delivered to the real target.
                pierced = False
                if not (hit and hit.get("ok")):
                    try:
                        hit = await el.evaluate(PIERCE_OVERLAY_JS)
                        pierced = True
                        if hit and hit.get("ok") and hit.get("pierced"):
                            el_res["click_method_note"] = f"pierced {hit['pierced']} overlay(s)"
                    except Exception:
                        pass

                if hit and hit.get("ok"):
                    try:
                        await cdp.send("Input.dispatchMouseEvent", {
                            "type": "mouseMoved", "x": hit["x"], "y": hit["y"]})
                        await cdp.send("Input.dispatchMouseEvent", {
                            "type": "mousePressed", "x": hit["x"], "y": hit["y"],
                            "button": "left", "clickCount": 1})
                        await asyncio.sleep(0.05)
                        await cdp.send("Input.dispatchMouseEvent", {
                            "type": "mouseReleased", "x": hit["x"], "y": hit["y"],
                            "button": "left", "clickCount": 1})
                        clicked = True
                        click_method = "cdp-click"
                    except Exception:
                        pass

                # STILL obscured. Playwright's force-click is NOT a safe
                # fallback here: force only skips actionability checks, it
                # still dispatches at the element's coordinates, so it lands
                # on whatever is covering the element — producing a confident
                # "clicked" with another element's beacons, or none at all.
                # Dispatch on the node itself instead: untrusted, but it is
                # guaranteed to reach the element under test.
                if not clicked:
                    try:
                        await el.evaluate("e => e.click()")
                        clicked = True
                        click_method = "js-click"
                    except Exception:
                        pass

                # Only now try a coordinate click, as a last resort for nodes
                # where .click() is a no-op (e.g. label/summary handling).
                if not clicked:
                    try:
                        await el.click(force=True, timeout=2500, no_wait_after=True)
                        clicked = True
                        click_method = "force-click"
                    except Exception:
                        pass

                # Put any neutralised overlays back before anything else runs.
                if pierced:
                    try:
                        await page.evaluate(RESTORE_OVERLAY_JS)
                    except Exception:
                        pass

                # ---- VERIFY the intended node actually received the click ----
                if clicked:
                    await asyncio.sleep(0.25)
                    witness = None
                    try:
                        witness = await page.evaluate(READ_CLEAR_WITNESS_JS)
                    except Exception:
                        witness = None
                    w_uids = (witness or {}).get("uids") or []
                    if uid and w_uids:
                        if uid in w_uids:
                            el_res["click_verified"] = True
                        else:
                            # A different element swallowed the click. Retry by
                            # dispatching directly on our node rather than
                            # recording the other element's beacons here.
                            # JS dispatch, not a coordinate click — the whole
                            # reason we got here is that coordinates resolve
                            # to somebody else.
                            try:
                                await el.evaluate("e => e.click()")
                                click_method += "+retarget-js"
                            except Exception:
                                pass
                            await asyncio.sleep(0.25)
                            try:
                                w2 = await page.evaluate(READ_CLEAR_WITNESS_JS)
                                if uid in ((w2 or {}).get("uids") or []):
                                    el_res["click_verified"] = True
                            except Exception:
                                pass
                    elif uid and not w_uids:
                        # No witness at all (handler stopped propagation before
                        # our capture listener, or click never landed).
                        el_res["click_verified"] = False

            if clicked:
                # ---- ADAPTIVE WAIT ----
                # Fixed 3s was the single biggest source of wrong data on this
                # stack: GTM defers /g/collect by ~5s, so every tracked click's
                # beacon arrived AFTER the window closed and got recorded under
                # the NEXT element. Read the instant API-level signal first —
                # dataLayer/ACDL push synchronously at click time — and only
                # pay the long wait when we know a beacon is actually coming.
                await asyncio.sleep(0.6)
                try:
                    early_caps = await page.evaluate(PEEK_TRACKING_JS)
                except Exception:
                    early_caps = []
                expect_beacon = any(
                    not _is_noise_event((c.get("detail") or {}).get("event")
                                        if isinstance(c.get("detail"), dict) else "")
                    for c in (early_caps or [])
                ) or bool(elem.get("is_download"))

                if os.environ.get("TV_DEBUG"):
                    try:
                        _dbg = await page.evaluate("""()=>({dl:!!(window.dataLayer&&window.dataLayer.__qaWrapped),
                            inst:!!window.__qaTrackInstalled, caps:(window.__qaTrackCaptures||[]).length,
                            dllen:Array.isArray(window.dataLayer)?window.dataLayer.length:-1,
                            url:location.href.slice(0,60)})""")
                    except Exception as _e:
                        _dbg = {"err": str(_e)[:60]}
                    sys.stdout.write("      DBG ei=%s reqs=%s bm=%s caps=%s %s\n" % (
                        ei, len(all_requests), req_bookmark, len(early_caps or []), _dbg))
                    sys.stdout.flush()
                deadline = time.time() + (CLICK_BEACON_MAX_WAIT if expect_beacon
                                          else CLICK_IDLE_WAIT)
                quiet_needed = 1.2 if expect_beacon else 0.6
                last_seen = len(all_requests)
                last_change = time.time()
                while time.time() < deadline:
                    await asyncio.sleep(0.25)
                    if len(all_requests) != last_seen:
                        last_seen = len(all_requests)
                        last_change = time.time()
                    elif time.time() - last_change >= quiet_needed:
                        break

            label = elem.get("text", "")[:30] or elem.get("selector", "")[:30]
            if not clicked:
                current_cid["v"] = -1     # never leave a window open on a skip
                el_res["skipped"] = True
                el_res["skip_reason"] = "not-found" if el is None else "unclickable"
                sys.stdout.write(
                    f"[{index}/{total}]   [{ei+1}/{len(elements)}] [{elem.get('zone', 'body').upper()[:6]}] "
                    f"\"{label}\" -> [SKIPPED: {el_res['skip_reason']}]\n")
                sys.stdout.flush()
                continue

            # Close this element's window before parsing so nothing that
            # arrives during the (async) parse gets stamped with this cid.
            current_cid["v"] = -1

            # Capture: grab all NEW requests since the bookmark
            click_requests = all_requests[req_bookmark:]
            # RAW network log for this click — full QA visibility, nothing
            # silently dropped. Capped per element as a safety net only.
            el_res["network_requests"] = [
                {"url": r["url"][:300]} for r in click_requests[:200]
            ]
            el_res["network_count"] = len(click_requests)
            el_res["click_method"] = click_method
            el_res["attribution"] = "window"
            _parse_click_requests(click_requests, el_res)

            # Harvest API-LEVEL tracking fired by this click — catches Adobe
            # s.tl, Tealium utag.link, gtag events and GTM dataLayer pushes
            # even when their network beacon is deferred, batched or blocked.
            try:
                api_caps = await page.evaluate(HARVEST_TRACKING_JS)
            except Exception:
                api_caps = []
            for cap in api_caps or []:
                ctype = cap.get("type")
                det = cap.get("detail") or {}
                if ctype == "s.tl":
                    el_res["adobe_calls"].append({
                        "link_type": "s.tl:" + ((det.get("link_type") or "o") if isinstance(det, dict) else "o"),
                        "link_name": (det.get("link_name") or "") if isinstance(det, dict) else "",
                        "evars": {}, "props": {},
                    })
                    el_res["has_tracking"] = True
                elif ctype == "utag.link":
                    detail_str = json.dumps(det)[:180] if isinstance(det, (dict, list)) else str(det)[:180]
                    el_res["other_analytics"].append({"vendor": "Tealium utag.link", "url": detail_str})
                    el_res["has_tracking"] = True
                elif ctype == "gtag":
                    el_res["ga4_events"].append({
                        "event": (det.get("event") or "") if isinstance(det, dict) else str(det),
                        "measurement_id": "(gtag api)",
                        "params": (det.get("params") or {}) if isinstance(det, dict) else {},
                    })
                    el_res["has_tracking"] = True
                elif ctype in ("dataLayer", "adobeDataLayer") and isinstance(det, dict):
                    ev_name = str(det.get("event") or det.get("xdm:eventType") or "")
                    # Scroll-depth / GTM bootstrap / consent events fire on
                    # timers, not on this click. Recording them here is what
                    # made unrelated buttons look "tracked".
                    if _is_noise_event(ev_name):
                        continue
                    params = {k: v for k, v in det.items()
                              if k != "event" and isinstance(v, (str, int, float, bool))}
                    src = "(dataLayer)" if ctype == "dataLayer" else "(adobeDataLayer)"
                    el_res["datalayer_events"].append({
                        "event": ev_name, "source": src, "params": params,
                    })
                    el_res["ga4_events"].append({
                        "event": ev_name,
                        "measurement_id": src,
                        "params": params,
                    })
                    el_res["has_tracking"] = True

            # ActivityMap: the click stores its link record in the s_sq
            # cookie (sent with the NEXT page view, which we block) — the
            # cookie itself is proof the click is Adobe-tracked.
            try:
                s_sq = await page.evaluate(READ_CLEAR_SSQ_JS)
            except Exception:
                s_sq = ""
            if s_sq:
                oid = ""
                m_oid = re.search(r'oid%3D(.*?)%26', s_sq, re.I) or re.search(r'oid=([^&;]*)', s_sq, re.I)
                if m_oid:
                    oid = unquote(unquote(m_oid.group(1))).strip()[:120]
                el_res["adobe_calls"].append({
                    "link_type": "ActivityMap",
                    "link_name": oid or s_sq[:120],
                    "evars": {}, "props": {},
                })
                el_res["has_tracking"] = True

            # Log result for this element
            ga_str = ", ".join(e["event"] for e in el_res["ga4_events"]) or "--"
            aa_parts = []
            for c in el_res["adobe_calls"]:
                name = c.get("link_name") or c.get("events") or c.get("link_type") or "adobe"
                aa_parts.append(name)
            aa_str = ", ".join(aa_parts) or "--"
            if el_res["adobe_websdk"] and aa_str == "--":
                aa_str = ", ".join(w["event_type"] for w in el_res["adobe_websdk"])
            oa_str = ", ".join(sorted({o["vendor"] for o in el_res["other_analytics"]})) if el_res["other_analytics"] else ""
            zone_tag = elem.get("zone", "body").upper()[:6]
            extra = f" Other:[{oa_str}]" if oa_str else ""
            vtag = "" if el_res.get("click_verified") else " [UNVERIFIED]"
            sys.stdout.write(
                f"[{index}/{total}]   [{ei+1}/{len(elements)}] [{zone_tag}] \"{label}\" -> [{click_method}]{vtag} "
                f"GA4:[{ga_str}] Adobe:[{aa_str}]{extra}\n")
            sys.stdout.flush()

            # Leave the page in a neutral state for the next element. Toggles,
            # accordions and "expand" panels stay open after being clicked and
            # will physically cover the elements that come after them, which
            # shows up as a run of unverified clicks with no tracking. Escape
            # closes most overlays; anything this reopens is discarded by the
            # drain at the start of the next element.
            # A consent link opens a panel that would cover the page for every
            # row after it, so dismiss it here rather than refusing to test it.
            try:
                await page.keyboard.press("Escape")
                await asyncio.sleep(0.15)
                await page.evaluate(CLOSE_CONSENT_UI_JS)
            except Exception:
                pass

            # Close any popup tabs the click opened (target=_blank links).
            # Their requests were already captured by the context-level
            # listener; leaving them open would pollute later clicks.
            try:
                for extra_pg in context.pages:
                    if extra_pg != page:
                        try: await extra_pg.close()
                        except Exception: pass
            except Exception:
                pass

            # The nav blocker keeps most clicks on-page; if this one escaped
            # (JS redirect), recover now with retries + page re-create.
            try:
                await _ensure_on_page()
            except Exception:
                pass

        # ---- FINAL DRAIN ----
        # The last elements clicked may still have beacons in flight. Without
        # this they'd simply never be captured.
        if elements_result:
            current_cid["v"] = -1
            await asyncio.sleep(FINAL_DRAIN_WAIT)

        # ================= CONTENT-BASED RECONCILIATION =================
        # Timing alone cannot attribute a beacon that a tag manager defers by
        # several seconds — by the time it lands, a later element's window is
        # open, which is exactly how "button 1's parameters" end up displayed
        # under button 2. But every click beacon carries the clicked link's
        # OWN identity (ep.link_url / ep.link_text / gtm.elementUrl), so the
        # beacon can be bound to its true element by content instead.
        #
        # Pass 1: index every audited element by url and by text.
        # Pass 2: re-read every analytics request, work out which element it
        #         describes, and move it there — overriding the time-window
        #         guess whenever the payload names an element unambiguously.
        try:
            by_url = {}
            by_text = {}
            for i, r in enumerate(elements_result):
                if r.get("skipped"):
                    continue
                for uk in _url_keys(r["element"].get("href"), original_url):
                    by_url.setdefault(uk, []).append(i)
                tk = _attr_key_text(r["element"].get("text"))
                if tk:
                    by_text.setdefault(tk, []).append(i)

            def _pick(cands, fallback_cid, zone):
                """Narrow a candidate list down to one element."""
                if not cands:
                    return None
                if len(cands) == 1:
                    return cands[0]
                # The beacon often names the region it fired in; prefer the
                # candidate sitting in that region. This is what keeps a
                # header logo and a footer logo — identical href, identical
                # text — from absorbing each other's events.
                if zone:
                    zoned = [i for i in cands
                             if elements_result[i]["element"].get("zone") == zone]
                    if len(zoned) == 1:
                        return zoned[0]
                    if zoned:
                        return fallback_cid if fallback_cid in zoned else zoned[0]
                # Otherwise the click window that was open is the best tie-break.
                return fallback_cid if fallback_cid in cands else cands[0]

            def _resolve_owner(link_url, link_text, fallback_cid, zone=''):
                """Which element does this beacon describe? Returns an index
                into elements_result, or None to leave it where it is."""
                tk = _attr_key_text(link_text)
                u_hits = []
                for uk in _url_keys(link_url, original_url):
                    for i in by_url.get(uk, []):
                        if i not in u_hits:
                            u_hits.append(i)
                t_hits = by_text.get(tk, []) if tk else []

                # Strongest: url AND text agree.
                both = [i for i in u_hits if i in t_hits]
                if both:
                    return _pick(both, fallback_cid, zone)
                if u_hits:
                    return _pick(u_hits, fallback_cid, zone)
                if t_hits:
                    return _pick(t_hits, fallback_cid, zone)
                return None

            # Wipe the network-derived events and rebuild them from scratch
            # with correct ownership. API-level captures (dataLayer/ACDL/s.tl)
            # are NOT touched: those are synchronous at click time and were
            # already attributed correctly.
            api_only = []
            for r in elements_result:
                keep_ga4 = [e for e in r["ga4_events"]
                            if str(e.get("measurement_id", "")).startswith("(")]
                api_only.append(keep_ga4)
                r["ga4_events"] = []
                r["adobe_calls"] = [c for c in r["adobe_calls"]
                                    if str(c.get("link_type", "")).startswith("s.tl")
                                    or c.get("link_type") == "ActivityMap"]
                r["adobe_websdk"] = []
                r["other_analytics"] = []
                r["total_requests"] = 0

            moved = 0
            page_level_events = []
            for req in all_requests:
                req_low = req["url"].lower()
                cid = req.get("cid", -1)
                is_ga4 = ("/g/collect" in req_low or (
                    ("google-analytics.com" in req_low or "analytics.google.com" in req_low)
                    and "collect" in req_low))

                if is_ga4:
                    for ev in parse_ga4_event(req["url"], req["post"]):
                        name = ev.get("event") or ""
                        if not name or _is_noise_event(name):
                            continue
                        lu, lt = beacon_identity(ev)
                        owner = _resolve_owner(lu, lt, cid, beacon_zone(ev))
                        if owner is None:
                            # No element identity in the payload. Such hits are
                            # usually page-level (consent state, engagement,
                            # visibility) and merely happened to land inside a
                            # click's window — attributing them would print the
                            # same phantom event under a dozen unrelated
                            # buttons. Keep one only when the element's own
                            # click-time dataLayer capture named that same
                            # event, which proves the click caused it.
                            if not (lu or lt) and 0 <= cid < len(elements_result):
                                api_names = {
                                    str(x.get("event") or "").lower()
                                    for x in elements_result[cid].get("datalayer_events", [])
                                }
                                if name.lower() not in api_names:
                                    page_level_events.append(name)
                                    continue
                            owner = cid if 0 <= cid < len(elements_result) else None
                        if owner is None:
                            continue
                        if owner != cid:
                            moved += 1
                        tgt = elements_result[owner]
                        tgt["ga4_events"].append(ev)
                        tgt["has_tracking"] = True
                        tgt["total_requests"] += 1
                        if tgt["attribution"] != "content":
                            tgt["attribution"] = "content"
                    continue

                # Non-GA4 vendors carry no element identity, so they stay with
                # the click window they arrived in.
                if 0 <= cid < len(elements_result):
                    _parse_click_requests([req], elements_result[cid])

            # Put the API-level GA4 events back.
            for r, keep in zip(elements_result, api_only):
                r["ga4_events"] = keep + r["ga4_events"]
                if r["ga4_events"] or r["adobe_calls"]:
                    r["has_tracking"] = True

            # Collapse the same event sent to multiple GA4 properties.
            for r in elements_result:
                r["ga4_events"] = consolidate_ga4_events(r["ga4_events"])

            # Recompute has_tracking honestly after the rebuild.
            for r in elements_result:
                r["has_tracking"] = bool(r["ga4_events"] or r["adobe_calls"]
                                         or r["adobe_websdk"] or r["other_analytics"])

            sys.stdout.write(f"[{index}/{total}] Reconciled beacons by payload identity "
                             f"({moved} re-attributed to their true element)\n")
            if page_level_events:
                from collections import Counter as _C
                _pl = ", ".join(f"{k}({v})" for k, v in _C(page_level_events).most_common(6))
                sys.stdout.write(f"[{index}/{total}] Excluded page-level (non-click) events: {_pl}\n")
            # The per-element lines above were printed BEFORE reconciliation,
            # so reprint the final, corrected picture — that is what the
            # report and the UI actually contain.
            sys.stdout.write(f"[{index}/{total}] ---- FINAL (post-reconciliation) ----\n")
            for _i, _r in enumerate(elements_result):
                if _r.get("skipped"):
                    continue
                _names = []
                for _e in _r["ga4_events"]:
                    _n = _e.get("event") or ""
                    _ids = _e.get("measurement_ids") or []
                    _real = [x for x in _ids if not str(x).startswith("(")]
                    _names.append(f"{_n}x{len(_real)}" if len(_real) > 1 else _n)
                _lbl = (_r["element"].get("text") or "")[:30]
                _zt = _r["element"].get("zone", "body").upper()[:6]
                sys.stdout.write(
                    f"[{index}/{total}]   [{_i+1}] [{_zt}] \"{_lbl}\" -> "
                    f"{', '.join(_names) if _names else '(no tracking)'}\n")
            sys.stdout.flush()
        except Exception as _rec_err:
            sys.stdout.write(f"[{index}/{total}] [WARN] reconciliation skipped: {str(_rec_err)[:120]}\n")
            sys.stdout.flush()

        # DEBUG: dump unique hosts seen across the whole click loop
        try:
            from collections import Counter
            from urllib.parse import urlparse as _up
            hosts = Counter()
            for _r in all_requests:
                try:
                    h = (_up(_r["url"]).hostname or "").lower()
                    if h: hosts[h] += 1
                except: pass
            total_reqs = len(all_requests)
            tracking_keywords = ("adobedtm", "adobedc", "demdex", "omtrdc", "sc.omtrdc",
                                 "tealium", "tiqcdn", "google-analytics", "googletagmanager",
                                 "/b/ss/", "segment.io", "/g/collect", "2o7.net")
            tracking_hosts = {h: c for h, c in hosts.items()
                              if any(k in h for k in tracking_keywords) or any(k in h for k in ["analytics", "tag", "tracking", "collect"])}
            top10 = ", ".join(f"{h}({c})" for h, c in hosts.most_common(10))
            sys.stdout.write(f"[{index}/{total}] DEBUG total_requests={total_reqs} "
                             f"top_hosts: {top10}\n")
            if tracking_hosts:
                sys.stdout.write(f"[{index}/{total}] DEBUG analytics_hosts_seen: "
                                 f"{', '.join(f'{h}({c})' for h, c in tracking_hosts.items())}\n")
            else:
                sys.stdout.write(f"[{index}/{total}] DEBUG NO analytics hosts in capture — "
                                 f"likely cookie consent not accepted OR site uses 1st-party proxy\n")
            sys.stdout.flush()
        except Exception:
            pass

        try: cdp.remove_listener("Network.requestWillBeSent", on_req_persistent)
        except: pass
        try: page.remove_listener("request", on_pw_request)
        except: pass
        try: context.remove_listener("request", on_pw_request)
        except: pass

        await page.close()
    except Exception as e:
        sys.stdout.write(f"[{index}/{total}] [ERROR] {url}: {str(e)[:80]}\n")
        sys.stdout.flush()
    finally:
        if context:
            try:
                await context.close()
            except:
                pass

    tracked = sum(1 for r in elements_result if r["has_tracking"])
    skipped = sum(1 for r in elements_result if r.get("skipped"))
    sys.stdout.write(f"[{index}/{total}] Done: {url} | {tracked}/{len(elements_result)} elements have tracking"
                     f"{f' | {skipped} skipped' if skipped else ''}\n")
    sys.stdout.flush()

    return {
        "URL": url,
        "Total_Elements": len(elements_result),
        "With_Tracking": tracked,
        "Without_Tracking": len(elements_result) - tracked,
        "Skipped": skipped,
        "Error": "",
        "_click_rich": {
            "URL": url,
            "elements": elements_result
        }
    }


# ===== SDR (Solution Design Reference) VALIDATION =====
# Reads the user's SDR Excel — a per-button QA spec — opens each page, finds
# the matching link/button, clicks it, captures the GA4 events that fire,
# and compares against the SDR's expected event name + parameters. Output
# is PASS/FAIL per row with a parameter-level diff.

SDR_SKIP_VALUES = {'', '-', '--', 'n/a', 'na', 'nan', 'none', 'tbd', 'x'}
SDR_GLOBAL_URLS = {'all pages', 'global', 'all page', 'sitewide', 'site wide'}


SDR_SHOT_DIR = "sdr_shots"

# Ring the element the audit actually clicked, so the screenshot shows which
# control the row was judged on rather than a page that could be anything.
SDR_RING_JS = r"""
(uid) => {
    const el = document.querySelector('[data-tvuid="' + uid + '"]');
    if (!el) return null;

    // An element rendered by `display: contents` reports no box of its own;
    // fall back to its first laid-out child so the crop still finds it.
    const rectOf = (n) => {
        let r = n.getBoundingClientRect();
        if (r.width > 2 && r.height > 2) return r;
        for (const c of n.getClientRects()) {
            if (c.width > 2 && c.height > 2) return c;
        }
        for (const k of n.children) {
            const kr = k.getBoundingClientRect();
            if (kr.width > 2 && kr.height > 2) return kr;
        }
        return r;
    };

    el.setAttribute('data-tv-ring', '1');
    el.style.setProperty('outline', '3px solid #e11d48', 'important');
    el.style.setProperty('outline-offset', '2px', 'important');

    // A sub-menu item, a tray link, an accordion row: most of what a page
    // offers is folded away until something opens it, and a control with no
    // box cannot be photographed. Unfold whatever is hiding it, just long
    // enough for the picture, then put it all back.
    if (!(rectOf(el).width > 2 && rectOf(el).height > 2)) {
        let n = el, guard = 0;
        while (n && n !== document.body && guard++ < 14) {
            const cs = getComputedStyle(n);
            const collapsed = (cs.overflow !== 'visible')
                && (parseFloat(cs.maxHeight) === 0 || parseFloat(cs.height) === 0);
            if (cs.display === 'none' || cs.visibility === 'hidden'
                    || parseFloat(cs.opacity) === 0 || collapsed) {
                n.setAttribute('data-tv-shown', n.getAttribute('style') || '__none__');
                n.style.setProperty('display', 'block', 'important');
                n.style.setProperty('visibility', 'visible', 'important');
                n.style.setProperty('opacity', '1', 'important');
                n.style.setProperty('max-height', 'none', 'important');
                n.style.setProperty('height', 'auto', 'important');
                n.style.setProperty('overflow', 'visible', 'important');
            }
            n = n.parentElement;
        }
    }
    try { el.scrollIntoView({block: 'center', inline: 'center'}); } catch (e) {}

    // A tray or sticky bar the audit opened earlier can sit on top of the
    // control. Pointer-events tricks do not help a photograph — the thing has
    // to be out of the way — so anything painted over the target's centre is
    // hidden for the length of the shot and put back afterwards.
    const r = rectOf(el);
    const cx = r.left + r.width / 2, cy = r.top + r.height / 2;
    let hidden = 0;
    const hide = (n) => {
        n.setAttribute('data-tv-hid', n.style.visibility || '__none__');
        n.style.setProperty('visibility', 'hidden', 'important');
        hidden++;
    };

    // A tray the audit opened earlier — an ISI panel, a modal — can be pinned
    // over the whole viewport. It is not necessarily what elementsFromPoint
    // reports at the target's centre, so full-screen pinned layers are taken
    // out by their own measurement: pinned, large, and not an ancestor of the
    // control being photographed.
    try {
        const vw = window.innerWidth, vh = window.innerHeight;
        for (const n of document.querySelectorAll(
                'div,section,aside,dialog,[role="dialog"]')) {
            if (hidden >= 8) break;
            if (n.contains(el) || el.contains(n)) continue;
            const cs = getComputedStyle(n);
            if (cs.position !== 'fixed' && cs.position !== 'sticky') continue;
            if (cs.visibility === 'hidden' || cs.display === 'none') continue;
            const b = n.getBoundingClientRect();
            if (b.width * b.height < vw * vh * 0.45) continue;
            if (b.top > vh || b.bottom < 0) continue;
            hide(n);
        }
    } catch (e) {}

    try {
        for (const n of document.elementsFromPoint(cx, cy)) {
            if (n === el) break;
            // An ancestor is not an overlay — hiding it would hide the target
            // too. Skip past it rather than stopping, because a tray whose
            // wrapper happens to contain the whole page would otherwise end
            // the search on its first step and nothing would be uncovered.
            if (n.contains(el) || el.contains(n)) continue;
            if (n === document.body || n === document.documentElement) continue;
            if (hidden >= 8) break;
            if (n.hasAttribute('data-tv-hid')) continue;
            hide(n);
        }
    } catch (e) {}

    const r2 = rectOf(el);
    if (!r2.width || !r2.height) return null;
    // Screen coordinates — the same frame the viewport screenshot is in.
    return {x: r2.left, y: r2.top, w: r2.width, h: r2.height, hidden: hidden,
            vw: window.innerWidth, vh: window.innerHeight,
            onscreen: r2.bottom > 0 && r2.top < window.innerHeight
                      && r2.right > 0 && r2.left < window.innerWidth};
}
"""

SDR_UNRING_JS = r"""
() => {
    document.querySelectorAll('[data-tv-ring]').forEach(n => {
        n.style.removeProperty('outline');
        n.style.removeProperty('outline-offset');
        n.removeAttribute('data-tv-ring');
    });
    document.querySelectorAll('[data-tv-shown]').forEach(n => {
        const prev = n.getAttribute('data-tv-shown');
        if (prev && prev !== '__none__') n.setAttribute('style', prev);
        else n.removeAttribute('style');
        n.removeAttribute('data-tv-shown');
    });
    document.querySelectorAll('[data-tv-hid]').forEach(n => {
        const prev = n.getAttribute('data-tv-hid');
        n.style.removeProperty('visibility');
        if (prev && prev !== '__none__') n.style.visibility = prev;
        n.removeAttribute('data-tv-hid');
    });
}
"""


async def _sdr_capture_click_shot(page, out_dir, excel_row, attempt, uid):
    """The control about to be tested, ringed, in the state the audit found it.

    This is the evidence a manual pass would paste beside the row: not a
    description of what was clicked, but a picture of it. If the element
    cannot be ringed there is nothing to show, so no file is written — a
    screenshot of a page with nothing marked on it is not evidence.
    """
    if not uid:
        return ""
    try:
        os.makedirs(out_dir, exist_ok=True)
        box = await page.evaluate(SDR_RING_JS, uid)
        if not box or not box.get("onscreen"):
            try:
                await page.evaluate(SDR_UNRING_JS)
            except Exception:
                pass
            return ""
        await asyncio.sleep(0.35)

        path = os.path.join(out_dir, "row%s_try%d_click.png" % (excel_row, attempt))
        await page.screenshot(path=path, animations="disabled")
        try:
            await page.evaluate(SDR_UNRING_JS)
        except Exception:
            pass
        _crop_to_element(path, box)
        _shrink_png(path)
        return path
    except Exception:
        return ""


def _crop_to_element(path, box, pad_x=170, pad_y=110):
    """Cut the region around the ringed control out of a viewport screenshot.

    The screenshot and the box are both in screen coordinates, so this holds
    for a sticky header and a scrolled page alike.
    """
    try:
        from PIL import Image
        im = Image.open(path)
        # A screenshot can be taken at a higher device pixel ratio than CSS.
        scale = im.width / float(box.get("vw") or im.width)
        left = max(0, int((box["x"] - pad_x) * scale))
        top = max(0, int((box["y"] - pad_y) * scale))
        right = min(im.width, int((box["x"] + box["w"] + pad_x) * scale))
        bottom = min(im.height, int((box["y"] + box["h"] + pad_y) * scale))
        if right - left < 20 or bottom - top < 20:
            return
        im.crop((left, top, right, bottom)).save(path, "PNG")
    except Exception:
        pass


def _shrink_png(path, width=640):
    """Keep a screenshot to what the workbook actually draws.

    Every picture is carried inside the .xlsx, so a hundred-row SDR would
    otherwise hand back a file of tens of megabytes.
    """
    try:
        from PIL import Image
        im = Image.open(path)
        if im.width > width:
            im = im.resize((width, max(1, int(im.height * width / im.width))),
                           Image.LANCZOS)
        im.convert("RGB").save(path, "PNG", optimize=True)
    except Exception:
        pass


def _sdr_panel_html(r):
    """An Omnibug-style read-out of the hit this row was judged on."""
    import html as _h

    status = r.get("status", "")
    tone = {"PASS": "#15803d", "FAIL": "#b42318"}.get(status, "#6b7280")
    ev = r.get("actual_event") or "(no event fired)"
    mid = r.get("measurement_id") or "-"

    # Show every expected parameter with its verdict, then anything else the
    # hit carried — which is exactly the order a person reads Omnibug in.
    rows, seen = [], set()
    for d in r.get("param_diff", []):
        seen.add(d["param"])
        got = d.get("actual") or ""
        ok = d.get("match")
        mark = "✓" if ok and not d.get("cosmetic") else ("≈" if ok else "✕")
        colour = "#15803d" if ok and not d.get("cosmetic") else ("#b45309" if ok else "#b42318")
        rows.append(
            '<tr><td class="m" style="color:%s">%s</td><td class="k">%s</td>'
            '<td class="v">%s</td><td class="e">%s</td></tr>' % (
                colour, mark, _h.escape(str(d["param"])),
                _h.escape(str(got) if got else "(not sent)"),
                "" if ok else "expected " + _h.escape(str(d.get("expected", "")))))
    for k, v in (r.get("actual_params") or {}).items():
        if k in seen:
            continue
        rows.append('<tr><td class="m" style="color:#9aa1ad">.</td><td class="k">%s</td>'
                    '<td class="v">%s</td><td class="e"></td></tr>'
                    % (_h.escape(str(k)), _h.escape(str(v))))
    if not rows:
        rows.append('<tr><td class="m" style="color:#b42318">X</td>'
                    '<td class="k" colspan="3">no GA4 hit captured for this click</td></tr>')

    return """<!doctype html><meta charset="utf-8"><style>
    * { box-sizing: border-box; margin: 0; }
    body { background: #fff; font: 12px/1.45 'Segoe UI', Arial, sans-serif; color: #14171f; }
    .panel { width: 620px; border: 1px solid #d0d5dd; border-radius: 6px; overflow: hidden; }
    .hd { padding: 7px 10px; background: #f7f8fa; border-bottom: 1px solid #e4e7ec;
          display: flex; gap: 8px; align-items: baseline; }
    .hd .ev { font-weight: 700; font-size: 13px; }
    .hd .st { margin-left: auto; font-weight: 700; color: %s; }
    .hd .id { color: #6b7280; font-size: 11px; }
    table { width: 100%%; border-collapse: collapse; }
    td { padding: 4px 8px; border-bottom: 1px solid #f0f2f5; vertical-align: top;
         word-break: break-all; }
    td.m { width: 26px; font-weight: 700; text-align: center; white-space: nowrap;
           font-size: 13px; }
    td.k { width: 168px; color: #3a4152; font-family: Consolas, monospace; }
    td.v { font-family: Consolas, monospace; }
    td.e { width: 190px; color: #b42318; font-size: 11px; }
    tr:last-child td { border-bottom: none; }
    </style>
    <div class="panel">
      <div class="hd"><span class="ev">%s</span><span class="id">%s</span>
        <span class="st">%s</span></div>
      <table>%s</table>
    </div>""" % (tone, _h.escape(str(ev)), _h.escape(str(mid)), status, "".join(rows))


async def _sdr_render_param_panels(browser, results, out_dir):
    """Draw one panel image per row, in a single throwaway page.

    Rendering after the run rather than during it keeps the audit itself at
    the same speed — the data was already captured when the click happened.
    """
    made = 0
    try:
        os.makedirs(out_dir, exist_ok=True)
        ctx = await browser.new_context(viewport={"width": 660, "height": 900},
                                        device_scale_factor=2)
        page = await ctx.new_page()
        for r in results:
            try:
                await page.set_content(_sdr_panel_html(r), wait_until="load")
                el = await page.query_selector(".panel")
                if not el:
                    continue
                path = os.path.join(out_dir, "row%s_params.png" % r.get("excel_row"))
                await el.screenshot(path=path)
                r["shot_params"] = path
                made += 1
            except Exception:
                continue
        await ctx.close()
    except Exception:
        pass
    return made


# A staging site guards itself one of two ways, and which one it is cannot be
# told from the URL. Basic auth is answered by the browser before any page
# exists; a login form only appears once a page has loaded. Both are handled,
# and the credentials are read from the environment rather than the command
# line so they never appear in a process listing.
def _auth_from_env():
    u = os.environ.get("TV_AUTH_USER", "").strip()
    p = os.environ.get("TV_AUTH_PASS", "")
    return (u, p) if (u or p) else ("", "")


LOGIN_FORM_JS = r"""
() => {
    // A visible password box is the one reliable sign of a login form.
    const pw = [...document.querySelectorAll('input[type="password"]')]
        .find(n => {
            const r = n.getBoundingClientRect();
            const cs = getComputedStyle(n);
            return r.width > 20 && r.height > 8
                && cs.visibility !== 'hidden' && cs.display !== 'none';
        });
    if (!pw) return null;

    const form = pw.closest('form');
    const scope = form || document;
    // The user box is the text-ish input just before the password one.
    const candidates = [...scope.querySelectorAll(
        'input[type="text"], input[type="email"], input:not([type])')]
        .filter(n => {
            const r = n.getBoundingClientRect();
            return r.width > 20 && r.height > 8;
        });
    let user = null;
    for (const n of candidates) {
        if (pw.compareDocumentPosition(n) & Node.DOCUMENT_POSITION_PRECEDING) user = n;
    }
    if (!user && candidates.length) user = candidates[0];

    const mark = (n, name) => { if (n) n.setAttribute('data-tv-login', name); };
    mark(pw, 'pass');
    mark(user, 'user');
    const submit = scope.querySelector(
        'button[type="submit"], input[type="submit"], button:not([type])');
    mark(submit, 'submit');
    return {user: !!user, submit: !!submit, inForm: !!form};
}
"""


async def _sdr_goto(page, url, tries=3, timeout=45000):
    """Load a page, allowing for the connection blinking.

    A CDN in front of a client site will drop a connection now and then —
    ERR_SSL_PROTOCOL_ERROR, ERR_TIMED_OUT, a reset — especially when an audit
    is asking it for the same page repeatedly. One attempt turns that blink
    into "Page did not load" against every row on the page, which reads like
    a site-wide outage and throws away a run that was minutes from finishing.
    Returns the last error, or "" when the page came up.
    """
    last = ""
    for attempt in range(tries):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            if attempt:
                sys.stdout.write("[SDR] Page loaded on attempt %d\n" % (attempt + 1))
                sys.stdout.flush()
            return ""
        except Exception as e:
            last = str(e)
            if attempt + 1 < tries:
                await asyncio.sleep(2.5 * (attempt + 1))
    return last


async def _sdr_try_login(page, user, pw):
    """Fill and submit a login form if the page is showing one.

    Returns True when a form was submitted, so the caller can say so once
    rather than guessing from the page that follows.
    """
    if not (user or pw):
        return False
    try:
        found = await page.evaluate(LOGIN_FORM_JS)
    except Exception:
        return False
    if not found:
        return False
    try:
        if found.get("user"):
            await page.fill('[data-tv-login="user"]', user)
        await page.fill('[data-tv-login="pass"]', pw)
        if found.get("submit"):
            await page.click('[data-tv-login="submit"]')
        else:
            await page.press('[data-tv-login="pass"]', "Enter")
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=20000)
        except Exception:
            pass
        await asyncio.sleep(2.0)
        return True
    except Exception:
        return False


NOT_IN_SDR_SHEET = "Not in SDR"


def _sweep_pick_event(evs, text, href):
    """Which of these hits belongs to the control that was just clicked.

    Taking the first hit in the window is wrong: pages run timers, and a
    periodic ping lands inside every window. A click beacon, on the other
    hand, carries the thing that was clicked — its label or its destination —
    so the hit is chosen by what is in it.
    """
    t = re.sub(r"\s+", " ", str(text or "")).strip().lower()
    h = str(href or "").strip().lower().rstrip("/")
    best, best_score = None, 0
    for ev in evs:
        score = 0
        for v in (ev.get("params") or {}).values():
            lv = str(v).strip().lower()
            if not lv:
                continue
            if t and len(t) > 2 and (t in lv or lv in t):
                score += 3
            if h and len(h) > 8 and (h in lv or lv.rstrip("/").endswith(h)):
                score += 3
        if re.search(r"click|download|nav|link|cta|button|outbound|exit",
                     str(ev.get("event") or ""), re.I):
            score += 1
        if score > best_score:
            best, best_score = ev, score
    return best, best_score


def _sweep_resolve(extras_for_page):
    """Settle the verdicts for one page's sweep, once all of it is known.

    An event that shows up in nearly every window is ambient — a timer, an
    engagement ping — and cannot be what any particular button sent. That can
    only be judged after the whole page has been swept, which is why the
    verdict is decided here rather than at each click.
    """
    if not extras_for_page:
        return
    windows = len(extras_for_page)
    seen = {}
    for e in extras_for_page:
        for name in {ev.get("event", "") for ev in e.get("_raw_events", [])}:
            seen[name] = seen.get(name, 0) + 1
    ambient = {n for n, c in seen.items() if windows >= 4 and c >= windows * 0.7}

    for e in extras_for_page:
        evs = e.pop("_raw_events", []) or []
        own = [ev for ev in evs if ev.get("event") not in ambient]
        pick, score = _sweep_pick_event(own or evs, e.get("button_name"),
                                        e.get("link_url"))
        note = ""
        if pick is None or score == 0:
            pick = None
            if evs:
                note = ("only page-level hits fired in this window (%s) — nothing "
                        "that names this control" % ", ".join(sorted(
                            {ev.get("event", "") for ev in evs})[:4]))
            else:
                note = "no GA4 event fired on this click"
        elif score < 3:
            note = ("matched on the event name only — no parameter names this "
                    "control, so read it with care")
        e["status"] = "TAGGED" if pick else "NOT TAGGED"
        e["actual_event"] = (pick or {}).get("event", "")
        e["actual_params"] = (pick or {}).get("params", {}) or {}
        e["measurement_id"] = (pick or {}).get("measurement_id", "")
        e["fired_events"] = sorted({ev.get("event", "") for ev in evs})
        if not e.get("shot_click"):
            note = ((note + "; ") if note else "") + (
                "no screenshot — this control is not rendered on the page as "
                "loaded (it appears only on hover or inside a menu)")
        e["reason"] = note
        if ambient:
            e["ambient_events"] = sorted(ambient)


def _sweep_key(el):
    """Logical identity of a clickable, stable across re-scans.

    Elements are re-stamped on every scan, so a uid taken during row 1 does
    not name the same node by row 20. Text, destination and zone do.
    """
    return (
        re.sub(r"\s+", " ", str(el.get("text") or "")).strip().lower(),
        str(el.get("href") or "").strip().lower(),
        str(el.get("zone") or ""),
    )


def _sdr_clean(v):
    """Normalise one SDR cell. The template uses '-' for 'not applicable'."""
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except Exception:
        pass
    s = str(v).replace('\u00a0', ' ').strip()
    return '' if s.lower() in SDR_SKIP_VALUES else s


def _sdr_norm_header(h):
    """A header reduced to comparable words.

    Different teams write the same column a dozen ways: "GA4 Event Name",
    "event_name", "ga4_event". Lower-casing, dropping any "(link_url)" hint and
    flattening punctuation to spaces makes all of them comparable, so column
    detection is not tied to one house style.
    """
    t = str(h or '').strip().lower()
    t = re.sub(r'\(.*?\)', ' ', t)
    t = re.sub(r'[^a-z0-9]+', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()


# For each structural column: the exact names to prefer, then looser words to
# fall back on. Exact wins first so that "current_page_url" is not mistaken
# for the page URL column and "event_category" is not mistaken for the event.
_SDR_COLUMN_RULES = {
    'url':   (['page url', 'url', 'page', 'page link', 'landing page'],
              ['page url']),
    'loc':   (['location', 'section', 'placement', 'area', 'position'],
              ['location', 'placement']),
    'ctype': (['click type', 'event type', 'interaction type', 'type',
               'interaction', 'action type'],
              ['click type', 'interaction']),
    'name':  (['link button name', 'button name', 'element name', 'link name',
               'name', 'element', 'link', 'cta name', 'component name'],
              ['button name', 'element name', 'link name']),
    'event': (['ga4 event name', 'ga4 event', 'event name', 'event',
               'ga4 event link', 'ga4 events', 'event id'],
              ['ga4 event', 'event name']),
}

# Columns that describe the QA process rather than the tag being tested.
_SDR_NON_PARAM = ('qa', 'comment', 'note', 'status', 'jira', 'ticket', 'owner',
                  'dev', 'remark', 'issue', 'bug', 'priority')


def _sdr_pick_columns(headers):
    """Work out which column is which, for any SDR layout.

    Structural columns are matched on exact normalised names first and only
    then on looser words. Everything left over that has a header and is not
    about the QA process is treated as an expected parameter, named by its own
    header — which is how a sheet whose columns are simply `event_category`
    and `event_label`, with no "(param)" annotation anywhere, still works.
    """
    norm = [_sdr_norm_header(h) for h in headers]
    picked, taken = {}, set()

    for key, (exact, loose) in _SDR_COLUMN_RULES.items():
        found = 0
        for want in exact:
            for i, h in enumerate(norm):
                if h == want and (i + 1) not in taken:
                    found = i + 1
                    break
            if found:
                break
        if not found:
            for want in loose:
                for i, h in enumerate(norm):
                    if want in h and (i + 1) not in taken:
                        found = i + 1
                        break
                if found:
                    break
        if found:
            picked[key] = found
            taken.add(found)

    param_cols, qa_cols = {}, {}
    for i, raw in enumerate(headers):
        col = i + 1
        if col in taken or not str(raw).strip():
            continue
        low = str(raw).strip().lower()
        if any(w in low for w in _SDR_NON_PARAM):
            if 'qa' in low:
                qa_cols[str(raw).strip()] = col
            continue
        # An explicit "(link_url)" annotation names the parameter; otherwise
        # the header itself is the parameter name.
        m = re.search(r'\(([a-z_][a-z0-9_ ]*)\)\s*$', low)
        pname = (m.group(1) if m else _sdr_norm_header(raw)).strip().replace(' ', '_')
        if pname:
            param_cols[col] = pname
    return picked, param_cols, qa_cols


def _sdr_find_header_row(ws, max_scan=8):
    """Locate the real header row.

    These workbooks put a grouping banner on row 1 ("Event Name",
    "Parameter 1"...) and the actual column names on row 2. Rather than
    hard-coding row 2, find the row that carries the anchor columns — some
    sheets do use row 1.
    """
    return _sdr_header_row_from_grid(_sdr_grid(ws, max_scan), max_scan)


def _sdr_grid(ws, limit=None):
    """Read a worksheet into plain row tuples.

    openpyxl's `ws.cell(r, c)` is a linear scan on a read-only worksheet, so
    addressing every cell of a 1000x30 sheet that way takes minutes. Streaming
    the rows once is effectively instant.
    """
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if limit and i + 1 >= limit:
            break
    return rows


def _sdr_cellv(grid, r, c):
    """1-based cell lookup into a grid produced by _sdr_grid."""
    if r < 1 or r > len(grid):
        return None
    row = grid[r - 1]
    return row[c - 1] if 0 < c <= len(row) else None


def _sdr_header_row_from_grid(grid, max_scan=8):
    best, best_score = 1, -1
    for r in range(1, min(max_scan, len(grid)) + 1):
        vals = [str(v or '').strip().lower() for v in grid[r - 1]]
        joined = ' | '.join(vals)
        # Score a row on how many structural columns it looks like it names.
        # Written against normalised names so a sheet headed "url /
        # Element Name / event_name" scores as highly as the house template.
        cells = [_sdr_norm_header(v) for v in grid[r - 1]]
        score = 0
        for _key, (exact, loose) in _SDR_COLUMN_RULES.items():
            if any(c == w for c in cells for w in exact) or \
                    any(w in c for c in cells for w in loose):
                score += 1
        if score > best_score:
            best, best_score = r, score
    return best if best_score >= 2 else 1


def list_sdr_sheets(sdr_path):
    """Describe every sheet so the UI can offer a picker.

    Returns each sheet's name, how many rows look testable and the distinct
    page URLs — enough for someone to recognise which tab is the live SDR.
    """
    import openpyxl
    out = []
    try:
        wb = openpyxl.load_workbook(sdr_path, data_only=True, read_only=True)
    except Exception as e:
        return [{"name": "", "error": str(e)[:120], "testable_rows": 0, "page_urls": []}]
    for ws in wb.worksheets:
        try:
            grid = _sdr_grid(ws)
            hr = _sdr_header_row_from_grid(grid)
            hdr = [str(v or '').strip() for v in (grid[hr - 1] if hr <= len(grid) else ())]
            low = [h.lower() for h in hdr]
            cols, p_cols, _qa = _sdr_pick_columns(hdr)
            c_url = cols.get('url') or 1
            c_name = cols.get('name') or 0
            c_ev = cols.get('event') or 0
            if not c_ev:
                out.append({"name": ws.title, "header_row": hr, "testable_rows": 0,
                            "page_urls": [], "qa_columns": [h for h in hdr if 'qa' in h.lower()]})
                continue
            rows, urls, last = 0, [], ''
            for r in range(hr + 1, len(grid) + 1):
                u = _sdr_clean(_sdr_cellv(grid, r, c_url))
                if u:
                    last = u
                nm = _sdr_clean(_sdr_cellv(grid, r, c_name))
                ev = _sdr_clean(_sdr_cellv(grid, r, c_ev))
                # Count on any identifier — the name column is blank
                # throughout some sheets.
                ident = nm or any(_sdr_clean(_sdr_cellv(grid, r, cc))
                                  for cc in p_cols)
                if ident and ev:
                    rows += 1
                    if last and last not in urls:
                        urls.append(last)
            out.append({
                "name": ws.title,
                "header_row": hr,
                "testable_rows": rows,
                "page_urls": urls[:40],
                "qa_columns": [h for h in hdr if 'qa' in h.lower()],
            })
        except Exception as e:
            out.append({"name": ws.title, "error": str(e)[:120],
                        "testable_rows": 0, "page_urls": []})
    try:
        wb.close()
    except Exception:
        pass
    return out


def parse_sdr_file(sdr_path, sheet_name=None, base_url=""):
    """Parse one SDR sheet into test cases.

    Layout these workbooks share: a banner row, then the real header row, then
    data. `Page URL` is filled only on the first row of each block and blank
    afterwards meaning "same page as above", so it is forward-filled. A column
    named like "Link Text (link_text)" declares that the value underneath is
    the expected GA4 parameter `link_text` — that is what makes automated
    comparison possible at all.

    Each case keeps `excel_row` so results can be written straight back into
    the operator's own sheet, in the format they already use.
    """
    import openpyxl
    wb = openpyxl.load_workbook(sdr_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        sheets = list_sdr_sheets(sdr_path)
        pick = max(sheets, key=lambda s: s.get("testable_rows", 0)) if sheets else None
        ws = wb[pick["name"]] if pick and pick.get("name") in wb.sheetnames else wb.worksheets[0]

    grid = _sdr_grid(ws)
    hr = _sdr_header_row_from_grid(grid)
    headers = [str(v or '').strip() for v in (grid[hr - 1] if hr <= len(grid) else ())]
    low = [h.lower() for h in headers]

    cols, param_cols, qa_cols = _sdr_pick_columns(headers)
    c_url = cols.get('url') or 1
    c_loc = cols.get('loc') or 0
    c_type = cols.get('ctype') or 0
    c_name = cols.get('name') or 0
    c_event = cols.get('event') or 0

    cases = []
    last_url = ''
    for r in range(hr + 1, len(grid) + 1):
        raw_url = _sdr_clean(_sdr_cellv(grid, r, c_url))
        if raw_url:
            last_url = raw_url
        page_url = last_url

        name = _sdr_clean(_sdr_cellv(grid, r, c_name)) if c_name else ''
        expected_event = _sdr_clean(_sdr_cellv(grid, r, c_event)) if c_event else ''
        if not expected_event:
            continue   # spacer / section heading / not a test case

        expected_params = {}
        for col, pname in param_cols.items():
            v = _sdr_clean(_sdr_cellv(grid, r, col))
            # "<link url>", "<provider_name>" and friends are authoring
            # placeholders meaning "whatever is right for this row", not a
            # literal expected value. Comparing against them would fail every
            # such row for no reason, so they are treated as unspecified.
            if v and not re.fullmatch(r'<[^>]*>', v.strip()):
                expected_params[pname] = v

        # Which column names the control is not fixed. Some sheets fill
        # "Link/Button Name" on every row; others leave it blank throughout and
        # identify the control purely by its link text or URL. Requiring the
        # name threw away 60 of 67 rows on one sheet, so take whichever
        # identifier the sheet actually uses.
        if not name:
            name = (expected_params.get('link_text')
                    or expected_params.get('download_file_name') or '')
        if not name:
            lu = expected_params.get('link_url') or ''
            tail = re.sub(r'[?#].*$', '', lu).rstrip('/').rsplit('/', 1)[-1]
            name = tail or lu

        # "All Pages" / "Global" rows still need a concrete page to test on.
        eff = page_url
        if (not eff) or eff.lower() in SDR_GLOBAL_URLS:
            eff = base_url or ''
        cases.append({
            "excel_row": r,
            "sheet": ws.title,
            "page_url_raw": page_url,
            "page_url": eff,
            "applies_all_pages": bool(page_url and page_url.lower() in SDR_GLOBAL_URLS),
            "location": _sdr_clean(_sdr_cellv(grid, r, c_loc)).lower() if c_loc else '',
            "click_type": _sdr_clean(_sdr_cellv(grid, r, c_type)) if c_type else '',
            "button_name": name,
            # A row can name an event and still name nothing findable on the
            # page. It is reported as untested rather than dropped.
            "identifiable": bool(name or expected_params.get('link_text')
                                 or expected_params.get('link_url')),
            "link_text": expected_params.get('link_text', '') or name,
            "link_url": expected_params.get('link_url', ''),
            "expected_event": expected_event,
            "expected_params": expected_params,
        })

    try:
        wb.close()
    except Exception:
        pass
    return {
        "sheet": ws.title,
        "header_row": hr,
        "headers": headers,
        "qa_columns": qa_cols,
        "param_columns": param_cols,
        "cases": cases,
    }


def _norm_str(s):
    """Loose comparison normalisation — case/whitespace insensitive."""
    return re.sub(r'\s+', ' ', str(s or '').strip().lower())


async def _find_sdr_target(page, link_text, link_url):
    """Locate the SDR element by link URL (if given) or visible text.
    Returns a Playwright locator or None."""
    # 1. Exact href match
    if link_url and link_url.startswith('http'):
        try:
            esc = link_url.replace('"', '\\"')
            loc = page.locator(f'a[href="{esc}"]').first
            if await loc.count() > 0:
                return loc
        except Exception:
            pass
        # 2. Match by host+path (drop query/hash)
        try:
            from urllib.parse import urlparse
            p = urlparse(link_url)
            relative = p.path or '/'
            esc = relative.replace('"', '\\"')
            loc = page.locator(f'a[href$="{esc}"]').first
            if await loc.count() > 0:
                return loc
        except Exception:
            pass
    elif link_url and link_url.startswith('/'):
        try:
            esc = link_url.replace('"', '\\"')
            loc = page.locator(f'a[href="{esc}"]').first
            if await loc.count() > 0:
                return loc
        except Exception:
            pass
    # 3. Match by visible text (case-insensitive, partial)
    if link_text:
        try:
            loc = page.get_by_text(link_text, exact=False).first
            if await loc.count() > 0:
                return loc
        except Exception:
            pass
        # 4. Match by aria-label / title
        for attr in ('aria-label', 'title', 'alt'):
            try:
                esc = link_text.replace('"', '\\"')
                loc = page.locator(f'[{attr}*="{esc}" i]').first
                if await loc.count() > 0:
                    return loc
            except Exception:
                pass
    return None


def _sdr_show(v):
    """Collapse a captured value for display.

    Sites sometimes send a parameter with the source markup's newlines and
    indentation still in it. Printing that raw turns one line of a report into
    ten. The value is only tidied for reading — comparison always uses the
    real one.
    """
    s = re.sub(r'\s+', ' ', str(v if v is not None else '')).strip()
    return s if len(s) <= 160 else s[:157] + '...'


def _sdr_param_matches(pname, expected, actual, base_url=""):
    """Compare one expected SDR parameter against what actually fired.

    URL-ish parameters are compared on their normalised forms so that a
    relative value in the SDR ("/search") still matches the absolute URL the
    beacon reports. Everything else is a case- and whitespace-insensitive
    comparison, because SDR authors type values by hand.
    """
    if actual is None:
        return False
    exp, act = str(expected), str(actual)
    if pname in ('link_url', 'download_file_name', 'file_location', 'page_location'):
        ek, ak = _url_keys(exp, base_url), _url_keys(act, base_url)
        if ek and ak and (ek & ak):
            return True
        # download_file_name is often a bare name in one place and a path in
        # the other ("botox-cosmetic_pi" vs "/pdf/botox-cosmetic_pi.pdf").
        e_tail = _norm_str(exp).rstrip('/').split('/')[-1]
        a_tail = _norm_str(act).rstrip('/').split('/')[-1]
        if e_tail and a_tail:
            if e_tail == a_tail:
                return True
            e_stem = re.sub(r'\.[a-z0-9]{1,5}$', '', e_tail)
            a_stem = re.sub(r'\.[a-z0-9]{1,5}$', '', a_tail)
            if e_stem and e_stem == a_stem:
                return True
    return _norm_str(exp) == _norm_str(act)


def _sdr_label_norm(v):
    """Normalise a clickable's label for matching against an SDR row.

    Accordions and disclosure widgets rename themselves as they are used: the
    control an SDR calls "Expand All" reads "Collapse all" once something has
    opened it, and each item is exposed to assistive tech as
    "Expand: How is it made?". Matching raw text means the same button matches
    or misses depending on what an earlier row happened to click, so the
    toggle affordance is stripped and its two states are treated as one.
    """
    s = _norm_str(v)
    if not s:
        return ''
    # "Expand: How is it made?" -> "How is it made?"
    s = re.sub(r'^(expand|collapse|open|close|show|hide|toggle)\s*[:\-–]\s*', '', s)
    # "Collapse all" and "Expand all" are the same control in two states.
    s = re.sub(r'\b(expand|collapse|show|hide|open|close)\b', 'toggle', s)
    return re.sub(r'\s+', ' ', s).strip()


def _sdr_element_fp(e):
    """Identity for a discovered element that survives a page reload.

    data-tvuid is assigned fresh on every scan, so it cannot be used to find
    the same control again after the page has been reloaded between candidate
    attempts. Text, href, region and enclosing heading together do identify
    it — including telling apart three "Find a Provider" buttons that differ
    only by which block they sit in.
    """
    return (
        _norm_str(e.get('text'))[:60],
        (e.get('href') or '').strip().lower(),
        (e.get('zone') or 'body'),
        _norm_str(e.get('context'))[:40],
    )


def _sdr_region(case):
    """The page region the SDR pins this row to, if it names one.

    Two things can say it: the Location column, and an expected parameter such
    as link_location or file_location. When a sheet has a header row and a
    footer row for the very same control — same name, same link, same text —
    this is the only thing that tells them apart, so it is treated as binding
    rather than as a mild preference.
    """
    parts = [str(case.get('location') or '')]
    for k in ('link_location', 'file_location', 'event_type', 'section'):
        v = (case.get('expected_params') or {}).get(k)
        if isinstance(v, str):
            parts.append(v)
    blob = ' '.join(parts).lower()
    # Check footer first: "footer nav" contains both words.
    if 'footer' in blob:
        return 'footer'
    if 'header' in blob or 'menu' in blob or re.search(r'nav', blob):
        return 'header'
    return ''


def _sdr_el_in_region(el, region):
    """Is this element in the region the sheet named?

    `zone` only distinguishes header, footer and body, but a page names its own
    areas — an ISI block, a hero, a promo strip. The sheet uses those names, so
    the element's ancestor id/class words are consulted too. Without this, the
    Directions for Use link inside the ISI (which reports file_location
    "footer") cannot be told apart from the one in the nav.
    """
    if not region:
        return True
    if (el.get('zone') or 'body') == region:
        return True
    return region in (el.get('hints') or [])


def _sdr_location_detail(case):
    """The specific part of an SDR Location, minus the page region.

    "body - krista martins" -> "krista martins". That remainder names the card
    or section a control sits in, and is often the only thing distinguishing
    a dozen rows that all say "Read more".
    """
    raw = str(case.get('location') or '')
    return re.sub(r'^\s*(body|header|footer|menu|nav)\s*[-–:]\s*', '', raw, flags=re.I).strip()


# Differences that are cosmetic rather than functional. An SDR is written by
# hand, so it says "navigation_link" where the tag sends "navigation link",
# or omits a trailing question mark. The value being reported IS the intended
# one; flagging these as failures buries the handful of real defects under a
# hundred spelling notes. They pass, with the difference recorded so the sheet
# can still be tidied up later.
def _sdr_cosmetic_key(v):
    s = _norm_str(v)
    s = s.replace('&', ' and ')
    s = re.sub(r'[_\-–—/\\.:;,!?\'"“”‘’()\[\]]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def _sdr_url_host_only_diff(expected, actual):
    """True when two urls are the same path on different hosts.

    SDRs are often written while a site is on a staging host and the URLs are
    not rewritten for launch, so a sheet says hcpskinvive-dev.pdcoe.dev where
    the live tag reports hcp.skinvivebyjuvederm.com. This still FAILS — a URL
    that does not match is a real defect — but the reason says the difference
    is only the host, which points straight at the fix.
    """
    def split(u):
        u = str(u or '').strip()
        if not u:
            return None
        m = re.match(r'^https?://([^/?#]+)(.*)$', u, re.I)
        if not m:
            return None
        host = m.group(1).lower().lstrip('www.')
        rest = (m.group(2) or '/').rstrip('/') or '/'
        return host, rest
    a, b = split(expected), split(actual)
    if not a or not b:
        return False
    return a[0] != b[0] and a[1] == b[1]


def _sdr_cosmetic_match(expected, actual):
    """True when the two differ only in separators, case, punctuation or
    spacing — i.e. the same value written two ways."""
    if actual is None:
        return False
    e, a = _sdr_cosmetic_key(expected), _sdr_cosmetic_key(actual)
    return bool(e) and e == a


def _sdr_actual_value(params, pname):
    """Read a parameter out of a captured hit.

    GA4 carries user-scoped values as `up.*`, which the parser stores under a
    "[user] " prefix. An SDR simply names the parameter, so look in both
    places — otherwise a value that IS being sent reads as missing.
    """
    if not isinstance(params, dict):
        return None
    if pname in params:
        return params[pname]
    alt = "[user] " + pname
    if alt in params:
        return params[alt]
    low = {str(k).lower(): v for k, v in params.items()}
    return low.get(pname.lower(), low.get(alt.lower()))


def _sdr_score_element(case, el, base_url=""):
    """How well does a discovered element answer this SDR row?

    Returns (total_score, text_score). The SDR identifies a target the way a
    human would — by the button's name, its link URL and roughly where it sits
    on the page — so all three are scored rather than taking the first text
    match, which is how the wrong element used to get clicked when a link
    appears in both the header and the footer.

    text_score is reported separately because link URL alone is a weak
    identifier: whole blocks of SDR rows share one destination (every FAQ
    accordion row points at the FAQ page), and a URL-only match will happily
    collapse all of them onto whichever element owns that href. The caller
    uses text_score to prefer candidates whose label actually agrees.
    """
    score = 0
    text_score = 0
    el_text = _norm_str(el.get('text'))
    _ = el_text
    el_href = el.get('href') or ''
    zone = (el.get('zone') or 'body').lower()

    want_url = case.get('link_url') or ''
    if want_url:
        # Resolve against the page being tested: SDRs write relative values
        # ("/", "/search") while discovery stores the browser-resolved
        # absolute href. A bare "/" is the site root — a real identifier for
        # the logo, and the most common SDR row of all.
        if _url_keys(want_url, base_url) & _url_keys(el_href, base_url):
            score += 10

    def _tokens(v):
        return {t for t in re.split(r'[^a-z0-9]+', _sdr_label_norm(v)) if len(t) > 2}

    # Every name this control answers to, not just the one rendered.
    el_label_list = [l for l in (el.get('labels') or []) if l]
    if el.get('text') and el['text'] not in el_label_list:
        el_label_list.insert(0, el['text'])
    if not el_label_list:
        el_label_list = ['']

    def _agree(want):
        """Best agreement between `want` and any name this element carries."""
        w = _sdr_label_norm(want)
        if not w:
            return 0
        return max(_agree_one(w, lbl) for lbl in el_label_list)

    def _agree_one(w, raw_label):
        el_label = _sdr_label_norm(raw_label)
        el_tokens = _tokens(raw_label)
        if not el_label:
            return 0
        # Ignore trailing punctuation: SDRs write "What is it" for a heading
        # rendered as "What is it?".
        w_trim = w.rstrip('?!.:;')
        el_trim = el_label.rstrip('?!.:;')
        el_text = el_label
        if w == el_text or w_trim == el_trim:
            return 12
        if len(w) > 3 and (w in el_text or el_text in w):
            return 8
        if True:
            # Word overlap catches SDR typos, truncated names and labels the
            # page prefixes ("Expand: How are lines formed?") — the row naming
            # "How Much Resarch Has Gone Into Botox" should still find
            # "How much research has gone into BOTOX Cosmetic?".
            wt = _tokens(w)
            if wt and el_tokens:
                overlap = len(wt & el_tokens) / len(wt)
                if overlap >= 0.8:
                    return 9
                if overlap >= 0.6:
                    return 6
                if overlap >= 0.45:
                    return 3
        return 0

    # The Link/Button Name column is the row's own identity — it is what a
    # person scans the page for. link_text is a parameter *value*, and in real
    # sheets it is often copied from the row above and left unedited, so it is
    # kept as a weaker, secondary signal rather than allowed to outvote the
    # name and drag the match onto a different element.
    name_score = _agree(case.get('button_name') or '')
    ltext_score = _agree(case.get('link_text') or '')

    # Icon-only controls carry no usable label, so let a distinctive word from
    # the sheet match the icon's own name in the markup. Generic words are
    # excluded — every button contains something called "button" or "icon".
    if max(name_score, ltext_score) < 8:
        generic = {'button', 'link', 'icon', 'text', 'item', 'wrapper', 'inner',
                   'outer', 'content', 'container', 'element', 'component',
                   'cmp', 'emu', 'aaaem', 'svg', 'img', 'image', 'span', 'div'}
        el_hints = set(el.get('hints') or [])
        for want in (case.get('button_name') or '', case.get('link_text') or ''):
            words = {w for w in re.split(r'[^a-z0-9]+', _norm_str(want))
                     if len(w) >= 4 and w not in generic}
            if words and words <= el_hints:
                name_score = max(name_score, 9)
                break
    text_score = max(name_score, ltext_score)
    score += text_score

    # The SDR's Location often names the card or section the control sits in
    # ("body - krista martins"). When several controls share a name, that is
    # the only thing that tells them apart, so compare it against the
    # element's enclosing heading.
    loc_detail = _sdr_location_detail(case)
    # The Location column often names the page area in the build's own words
    # ("isi", "footer_promo", "hero"). Those words are in the markup, so match
    # them against it directly.
    if loc_raw_words := {w for w in re.split(r'[^a-z0-9]+', (case.get('location') or '').lower())
                         if len(w) >= 3}:
        if loc_raw_words & set(el.get('hints') or []):
            score += 6

    if loc_detail and len(loc_detail) > 3:
        ctx = _sdr_label_norm(el.get('context'))
        if ctx:
            d_norm = _sdr_label_norm(loc_detail)
            if d_norm and (d_norm == ctx or d_norm in ctx or ctx in d_norm):
                score += 7
            else:
                dt = {t for t in re.split(r'[^a-z0-9]+', d_norm) if len(t) > 2}
                ct = {t for t in re.split(r'[^a-z0-9]+', ctx) if len(t) > 2}
                if dt and ct and len(dt & ct) / len(dt) >= 0.6:
                    score += 4

    loc = (case.get('location') or '').lower()
    if loc:
        if 'header' in loc or 'menu' in loc or 'nav' in loc:
            score += 3 if zone == 'header' else -1
        elif 'footer' in loc:
            score += 3 if zone == 'footer' else -1
        elif 'body' in loc or 'hero' in loc or 'content' in loc:
            score += 2 if zone == 'body' else 0

    ctype = (case.get('click_type') or '').lower()
    if 'download' in ctype and el.get('is_download'):
        score += 3
    return score, text_score, name_score


async def validate_sdr(browser, sdr_path, start_url, sheet_name=None,
                       ga4_id="", ga4_mode="specific", resume=False,
                       progress_path="sdr_results.json", qa_column="",
                       filled_path="sdr_filled.xlsx"):
    """Automate a manual SDR QA pass.

    For every SDR row: open the page it names, find the button it names, click
    it for real, and compare the GA4 event and every expected parameter
    against what actually fired. A row passes only when the event name AND all
    of its parameters match; otherwise it fails with an explicit reason naming
    each parameter that differed, so the sheet says *why*, not just "Fail".

    ga4_id / ga4_mode select which GA4 property counts as the source of truth
    on sites that tag to more than one.
    """
    parsed = parse_sdr_file(sdr_path, sheet_name, base_url=start_url)
    cases = parsed["cases"]
    sys.stdout.write(f"[SDR] Sheet '{parsed['sheet']}' -> {len(cases)} test cases\n")
    if ga4_id:
        sys.stdout.write(f"[SDR] Validating against GA4 property {ga4_id} (mode={ga4_mode})\n")
    else:
        sys.stdout.write("[SDR] No GA4 property selected — accepting a hit on any property\n")
    sys.stdout.flush()
    if not cases:
        return {"meta": parsed, "results": []}

    # ---- resume support ----
    # These runs take tens of minutes. If one is interrupted — the server is
    # restarted, the machine sleeps, someone hits Cancel — everything done so
    # far used to be thrown away, because results were only written at the very
    # end. Progress is now flushed after every page, and a new run can pick up
    # where the last one stopped instead of re-clicking hundreds of elements.
    done_by_row = {}
    if resume and os.path.exists(progress_path):
        try:
            with open(progress_path, "r", encoding="utf-8") as f:
                prev = json.load(f)
            same = (prev.get("sheet") == parsed.get("sheet")
                    and str(prev.get("ga4_id") or "") == str(ga4_id or ""))
            if same:
                for r in prev.get("results", []):
                    if r.get("excel_row") and r.get("status") in ("PASS", "FAIL", "SKIPPED"):
                        done_by_row[r["excel_row"]] = r
        except Exception:
            done_by_row = {}
        if done_by_row:
            sys.stdout.write(f"[SDR] Resuming — {len(done_by_row)} row(s) already validated, "
                             f"{len(cases) - len(done_by_row)} to go\n")
            sys.stdout.flush()

    # Group by page so each URL is loaded once.
    by_page, page_order = {}, []
    for c in cases:
        key = (c.get("page_url") or "").strip()
        if key not in by_page:
            page_order.append(key)
            by_page[key] = []
        by_page[key].append(c)

    results = [done_by_row[r] for r in sorted(done_by_row)] if done_by_row else []
    detected_ids = set()

    def _flush(partial=True):
        """Persist what has been validated so far.

        Written after every page so an interrupted run still leaves a usable
        report behind, and so the UI can show progress mid-run.
        """
        try:
            ordered = sorted(results, key=lambda r: r.get("excel_row") or 0)
            payload = {
                "generated": datetime.datetime.now().isoformat(),
                "start_url": start_url,
                "sheet": parsed.get("sheet", ""),
                "ga4_id": ga4_id,
                "detected_ga4_ids": sorted(detected_ids),
                "partial": bool(partial),
                "completed": len(ordered),
                "total_cases": len(cases),
                "results": ordered,
                "not_in_sdr": list(extras),
            }
            tmp = progress_path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, default=str)
            os.replace(tmp, progress_path)
        except Exception:
            pass
        # Keep a downloadable workbook current as well, so an interrupted run
        # still hands back a filled SDR and not just raw JSON.
        try:
            if ordered:
                write_sdr_verdicts(sdr_path, parsed, ordered, filled_path,
                                   qa_column=qa_column, extras=list(extras))
        except Exception:
            pass

    def _record(case, status, reason, extra=None):
        row = {
            "excel_row": case["excel_row"],
            "page_url": case.get("page_url", ""),
            "location": case.get("location", ""),
            "button_name": case.get("button_name", ""),
            "link_text": case.get("link_text", ""),
            "link_url": case.get("link_url", ""),
            "expected_event": case.get("expected_event", ""),
            "expected_params": case.get("expected_params", {}),
            "status": status,
            "reason": reason,
            "actual_event": "",
            "actual_params": {},
            "fired_events": [],
            "param_diff": [],
            "matched_element": "",
            "click_method": "",
            "click_verified": False,
        }
        if extra:
            row.update(extra)
        results.append(row)
        return row

    auth_user, auth_pass = _auth_from_env()
    ctx_args = {
        "viewport": {'width': 1280, 'height': 800},
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    }
    if auth_user or auth_pass:
        # Answers the server's own challenge, before any page is rendered.
        ctx_args["http_credentials"] = {"username": auth_user, "password": auth_pass}
        sys.stdout.write("[SDR] Credentials supplied for %r — will answer Basic auth "
                         "and fill a login form if one appears" % auth_user + os.linesep)
        sys.stdout.flush()
    context = await browser.new_context(**ctx_args)

    all_requests = []
    seen_req_keys = set()
    current_cid = {"v": -1}
    # Controls found on the page that no SDR row accounts for. The sheet says
    # what was specified; this says what is actually there.
    extras = []

    def _push_req(u, post, ts):
        if not u:
            return
        k = (u, round(ts, 2))
        if k in seen_req_keys:
            return
        seen_req_keys.add(k)
        all_requests.append({"url": u, "post": post or "", "ts": ts, "cid": current_cid["v"]})

    def on_pw_request(request):
        try:
            pd_ = ""
            try:
                pd_ = request.post_data or ""
            except Exception:
                pass
            _push_req(request.url, pd_, time.time())
        except Exception:
            pass

    try:
        context.on("request", on_pw_request)
    except Exception:
        pass

    page = await context.new_page()
    await stealth_obj.apply_stealth_async(page)
    cdp = await context.new_cdp_session(page)
    await cdp.send("Network.enable")

    def on_cdp_req(params):
        try:
            req = params.get("request", {}) or {}
            _push_req(req.get("url", ""), req.get("postData", "") or "", time.time())
        except Exception:
            pass
    cdp.on("Network.requestWillBeSent", on_cdp_req)

    for page_url in page_order:
        # Rows already validated by an earlier interrupted run stay as they are.
        page_cases = [c for c in by_page[page_url]
                      if c["excel_row"] not in done_by_row]
        if not page_cases:
            continue

        # SDR rows sometimes name a placeholder instead of a real page
        # ("<providers name page>"). Those cannot be opened, and calling them
        # failures would be misleading.
        if not page_url or not page_url.lower().startswith("http"):
            for c in page_cases:
                _record(c, "SKIPPED",
                        f"No testable page URL in SDR (value: '{c.get('page_url_raw') or 'blank'}')")
            sys.stdout.write(f"[SDR] === SKIP '{page_url or '(blank)'}' "
                             f"({len(page_cases)} cases, not a real URL) ===\n")
            sys.stdout.flush()
            continue

        sys.stdout.write(f"[SDR] === {page_url} ({len(page_cases)} cases) ===\n")
        sys.stdout.flush()

        err = await _sdr_goto(page, page_url)
        if err:
            # Not a tagging failure. "Fail" in a QA column is a claim about the
            # tag; nothing was tested here, and saying otherwise puts false
            # failures in front of whoever reads the sheet. The row is marked
            # Not Tested with the reason, which is what actually happened.
            for c in page_cases:
                _record(c, "SKIPPED",
                        f"Page did not load after 3 attempts — nothing was tested "
                        f"on this row: {err[:70]}")
            sys.stdout.write("[SDR] %s did not answer after 3 attempts — %d row(s) "
                             "left Not Tested" % (page_url, len(page_cases)) + os.linesep)
            sys.stdout.flush()
            continue

        if await _sdr_try_login(page, auth_user, auth_pass):
            sys.stdout.write("[SDR] Signed in through the login form on the page"
                             + os.linesep)
            sys.stdout.flush()

        await asyncio.sleep(3)
        await accept_cookies(page)
        try:
            await page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass
        await asyncio.sleep(SDR_PAGE_SETTLE)

        # Walk the page top to bottom before discovering anything. Most of
        # these pages build their lower half lazily, so a scan done at the
        # fold simply cannot see the promo blocks, tab strips and footer CTAs
        # that SDR rows point at — those rows then resolve to a same-named
        # control higher up the page and get judged against the wrong element.
        try:
            await page.evaluate("""async () => {
                await new Promise((resolve) => {
                    let total = 0;
                    const step = 600;
                    const timer = setInterval(() => {
                        window.scrollBy(0, step);
                        total += step;
                        if (total >= document.body.scrollHeight || total > 25000) {
                            clearInterval(timer);
                            resolve();
                        }
                    }, 60);
                });
                window.scrollTo(0, 0);
            }""")
            await asyncio.sleep(2.0)
        except Exception:
            pass

        # Same preparation the click audit uses: keep clicks on-page, open
        # every menu so hidden targets exist, wrap the tracking APIs, and get
        # any consent dialog out of the way before measuring.
        try:
            await page.evaluate(BLOCK_NAVIGATION_JS)
        except Exception:
            pass
        try:
            await page.evaluate(EXPOSE_HIDDEN_JS)
            await asyncio.sleep(0.6)
        except Exception:
            pass

        discovered = {}

        def _absorb(batch):
            for b in batch or []:
                if b.get("uid"):
                    discovered[b["uid"]] = b

        try:
            _absorb(await page.evaluate(DISCOVER_CLICKABLES_JS))
        except Exception:
            pass

        # Hover each top-level nav item so CSS-only dropdowns render, and
        # click obvious menu openers — SDR rows routinely target sub-menu
        # links that do not exist in the DOM until their parent opens.
        try:
            nav = page.locator('header a, header button, nav > a, nav > button, '
                               'nav > ul > li > a, nav > ul > li > button, [aria-haspopup]')
            for i in range(min(await nav.count(), 40)):
                try:
                    await nav.nth(i).hover(timeout=350, force=True, no_wait_after=True)
                    await asyncio.sleep(0.2)
                    _absorb(await page.evaluate(DISCOVER_CLICKABLES_JS))
                except Exception:
                    pass
        except Exception:
            pass
        # Only NAVIGATION openers get clicked. Content accordions match the
        # same selectors, and opening them before testing changes the very
        # thing under test: a disclosure button reports its current state, so
        # a pre-expanded "Expand All" fires link_text "collapse all" and the
        # row fails against an SDR that describes the page as a person finds
        # it. Nav menus must be opened to reach sub-menu links; body
        # accordions must be left exactly as the page loaded.
        for sel in ('button[aria-label*="menu" i]', 'button[class*="hamburger" i]',
                    'button[class*="menu-toggle" i]', '[aria-haspopup="true"]',
                    'button[aria-expanded="false"]'):
            try:
                loc = page.locator(sel)
                for i in range(min(await loc.count(), 6)):
                    el_o = loc.nth(i)
                    if not await el_o.is_visible(timeout=250):
                        continue
                    if await el_o.evaluate("""e => {
                        // Skip anything that is not inside a navigation region.
                        let n = e, inNav = false;
                        while (n && n !== document.body) {
                            const tag = n.tagName || '';
                            const role = (n.getAttribute && n.getAttribute('role') || '').toLowerCase();
                            const id = (n.id || '').toLowerCase();
                            const cls = (typeof n.className === 'string' ? n.className : '').toLowerCase();
                            if (tag === 'HEADER' || tag === 'NAV' || role === 'banner'
                                || role === 'navigation' || /header|navbar|nav-|menu/.test(id + ' ' + cls)) {
                                inNav = true;
                                break;
                            }
                            n = n.parentElement;
                        }
                        if (!inNav) return true;
                        n = e;
                        while (n && n !== document.body) {
                            const id = (n.id || '').toLowerCase();
                            const cls = (typeof n.className === 'string' ? n.className : '').toLowerCase();
                            if (/cookie|consent|onetrust|optanon|privacy/.test(id + ' ' + cls)) return true;
                            n = n.parentElement;
                        }
                        return false;
                    }"""):
                        continue
                    await el_o.click(force=True, timeout=800, no_wait_after=True)
                    await asyncio.sleep(0.45)
                    _absorb(await page.evaluate(DISCOVER_CLICKABLES_JS))
            except Exception:
                pass
        try:
            await page.evaluate(EXPOSE_HIDDEN_JS)
            await asyncio.sleep(0.4)
            _absorb(await page.evaluate(DISCOVER_CLICKABLES_JS))
        except Exception:
            pass

        try:
            await page.evaluate(CLOSE_CONSENT_UI_JS)
        except Exception:
            pass
        try:
            await page.evaluate(BLOCK_NAVIGATION_JS)
            await page.evaluate(INSTRUMENT_TRACKING_JS)
            await page.evaluate(HARVEST_TRACKING_JS)
        except Exception:
            pass

        # Whole-page text, used to tell "this content is gone from the site"
        # apart from "the control is here but did not match this row".
        try:
            page_text = _norm_str(await page.evaluate("() => document.body.innerText"))
        except Exception:
            page_text = ""

        # The menu-expansion sweep above found things that are not visible in
        # the page's resting state; keep that as a fallback pool.
        base_elements = list(discovered.values())
        elements = base_elements
        sys.stdout.write(f"[SDR]   {len(base_elements)} clickable elements available for matching\n")
        sys.stdout.flush()

        async def _reset_page():
            """Reload the page and rebuild the element list.

            Needed between candidate attempts for one SDR row. The first click
            can open a form, switch a tab or expand a panel, and the second
            candidate then gets clicked in a page state that never occurs for
            a real user — which is how the hero "Find a Provider" kept being
            reported for the row that means the promo one further down.
            """
            try:
                await _sdr_goto(page, page_url)
                await _sdr_try_login(page, auth_user, auth_pass)
                await asyncio.sleep(2.0)
                try:
                    await page.wait_for_load_state("networkidle", timeout=12000)
                except Exception:
                    pass
                await page.evaluate("""async () => {
                    await new Promise((resolve) => {
                        let total = 0;
                        const timer = setInterval(() => {
                            window.scrollBy(0, 600); total += 600;
                            if (total >= document.body.scrollHeight || total > 25000) {
                                clearInterval(timer); resolve();
                            }
                        }, 60);
                    });
                    window.scrollTo(0, 0);
                }""")
                await asyncio.sleep(1.2)
                await page.evaluate(BLOCK_NAVIGATION_JS)
                await page.evaluate(EXPOSE_HIDDEN_JS)
                await page.evaluate(CLOSE_CONSENT_UI_JS)
                await page.evaluate(INSTRUMENT_TRACKING_JS)
                await page.evaluate(HARVEST_TRACKING_JS)
                await asyncio.sleep(0.4)
                batch = await page.evaluate(DISCOVER_CLICKABLES_JS)
                return [b for b in (batch or []) if b.get("uid")]
            except Exception:
                return []

        async def _fresh_elements():
            """Re-scan (and re-stamp) the page right before a click.

            These pages re-render after every interaction — an accordion that
            read "Expand: X" becomes "Collapse: X", React swaps nodes, and the
            data-tvuid stamped minutes ago goes with them. Matching against a
            list captured once at page load quietly drifts onto stale or wrong
            elements, so the list is refreshed per row.
            """
            try:
                batch = await page.evaluate(DISCOVER_CLICKABLES_JS)
                return [b for b in (batch or []) if b.get("uid")]
            except Exception:
                return []

        # Everything the SDR rows on this page end up clicking, so the sweep
        # afterwards can tell specified from unspecified.
        tested_keys = set()

        for ci, case in enumerate(page_cases):
            label = (case.get("button_name") or case.get("link_text") or "")[:34]

            # Not every SDR row describes a click. Page-load, history-change
            # and page-view rows have no button to press, and consent-banner
            # rows are deliberately out of scope for the audit. Calling those
            # "Fail — element not found" would be wrong: nothing was tested.
            # A row can name an event and still name nothing to click. Report
            # it as untested rather than dropping it from the sheet entirely.
            if not case.get("identifiable", True):
                _record(case, "SKIPPED",
                        "Row names no control — it has no Link/Button Name, Link Text "
                        "or Link URL to find on the page")
                sys.stdout.write(f"[SDR]   [{ci+1}/{len(page_cases)}] \"{label}\" -> NO TARGET (skipped)\n")
                sys.stdout.flush()
                continue

            _ct = (case.get("click_type") or "").lower()
            _nm = (case.get("button_name") or "").lower()
            _blob = _ct + " " + _nm
            if any(k in _blob for k in ("page load", "page view", "pageview",
                                        "history change", "scroll", "page_load",
                                        "timer", "impression")):
                _record(case, "SKIPPED",
                        f"Not a click test — SDR marks this as '{case.get('click_type') or case.get('button_name')}'")
                sys.stdout.write(f"[SDR]   [{ci+1}/{len(page_cases)}] \"{label}\" -> NOT A CLICK (skipped)\n")
                sys.stdout.flush()
                continue
            # Consent links ("Cookie Settings", "Your Privacy Choices") used to
            # be refused here. They are ordinary tracked links and the sheet
            # expects them to fire, so they are tested like any other row; the
            # panel they open is dismissed afterwards so it cannot cover the
            # rows that follow.

            # --- locate the element this SDR row is about ---
            # Two passes. If ANY element's label agrees with the row's name,
            # only those are considered; a link URL on its own is too weak to
            # pick between rows that all point at the same destination. Only
            # when nothing matches by label (icon buttons, logos, image links)
            # does a URL-only match get to win.
            # Match against the page as it is NOW, not as it was before the
            # previous row's click re-rendered it.
            fresh = await _fresh_elements()
            elements = fresh or base_elements

            def _score_all(cands):
                out = []
                for e in cands:
                    sc, tsc, nsc = _sdr_score_element(case, e, page_url)
                    if sc > 0:
                        out.append((sc, tsc, nsc, e))
                return out

            want_region = _sdr_region(case)

            def _pick_best(cands):
                # When the sheet pins the row to a region, a same-named control
                # in a different region is simply not what the row is about.
                # The Skinvive logo sits in both the header and the footer with
                # identical text and link; without this the header one wins
                # every time and the footer row is judged against it.
                # ...but only when that region actually holds something that
                # answers to the row's name. A Location like
                # "header - more from allergan aesthetics" describes a menu
                # whose links the page renders in the body; binding on the word
                # "header" alone would throw the real links away and report the
                # row as missing.
                if want_region:
                    in_region = [x for x in cands
                                 if _sdr_el_in_region(x[3], want_region) and x[2] > 0]
                    if in_region:
                        cands = in_region
                # Only a STRONG label match counts as agreement — exact, a
                # substring, or near-total word overlap. Anything weaker is
                # coincidence: "Expand All" shares "all" with "View All FAQs",
                # and "Botox Logo" shares "botox" with "About BOTOX Cosmetic".
                # Promoting those into the tier would let a coincidental word
                # beat the element the row's link URL actually points at.
                by_name = [x for x in cands if x[2] >= 8]
                by_text = [x for x in cands if x[1] >= 8]
                pool = by_name or by_text or cands
                b, bs = None, 0
                for sc, tsc, nsc, e in pool:
                    if sc > bs:
                        b, bs = e, sc
                return b, bs

            def _ordered(cands):
                """Best-first, but only within the tier that actually agrees."""
                reg = _sdr_region(case)
                if reg:
                    in_reg = [x for x in cands
                              if _sdr_el_in_region(x[3], reg) and x[2] > 0]
                    if in_reg:
                        cands = in_reg
                by_name = [x for x in cands if x[2] >= 8]
                by_text = [x for x in cands if x[1] >= 8]
                pool = by_name or by_text or cands
                return [e for _s, _t, _n, e in sorted(pool, key=lambda x: -x[0])]

            scored_now = _score_all(elements)
            best, best_score = _pick_best(scored_now)
            candidates = _ordered(scored_now)
            # Hidden targets (collapsed sub-menu links) only exist in the pool
            # captured while the menus were forced open.
            if (best is None or best_score < 4) and elements is not base_elements:
                alt = _score_all(base_elements)
                b2, bs2 = _pick_best(alt)
                if bs2 > best_score:
                    best, best_score = b2, bs2
                    candidates = _ordered(alt)
            if best is None or best_score < 4:
                # Distinguish "this content is no longer on the page" from
                # "the control is there but could not be matched". The first
                # means the SDR has drifted from the live site and the row
                # needs rewriting; the second is a matching problem. Guessing
                # between them wastes the reader's time.
                detail = ""
                probe = (_sdr_location_detail(case) or case.get("button_name") or "").strip()
                if probe and page_text:
                    if _norm_str(probe) and _norm_str(probe) not in page_text:
                        detail = (f" — the text '{probe}' does not appear anywhere on this page,"
                                  f" so this content looks to have been removed or moved since"
                                  f" the SDR was written")
                    else:
                        detail = (f" — '{probe}' does appear on the page, so the control exists"
                                  f" but could not be matched to this row")
                _record(case, "FAIL",
                        f"Element not found on page (looked for name '{case.get('button_name')}'"
                        f"{', url ' + case['link_url'] if case.get('link_url') else ''}){detail}")
                sys.stdout.write(f"[SDR]   [{ci+1}/{len(page_cases)}] \"{label}\" -> NOT FOUND\n")
                sys.stdout.flush()
                continue

            # A row is often ambiguous: several controls share a name and the
            # SDR tells them apart only by Location ("footer_promo", "isi
            # banner"). Guessing once and reporting the result gives confident
            # but wrong data — a header link judged against a footer row.
            # Instead, try the best candidates in turn and keep whichever one
            # actually satisfies the row. The first clean pass wins; otherwise
            # the closest fit is reported, so the reasons describe the element
            # the row most plausibly meant.
            attempts = []
            for _cand_i, _cand_meta in enumerate(candidates[:SDR_MAX_CANDIDATES]):
                best = _cand_meta
                if _cand_i > 0:
                    # Start this attempt from a page in its resting state, then
                    # find the same control again by fingerprint — the uid it
                    # carried belongs to the previous, now-discarded scan.
                    refreshed = await _reset_page()
                    want_fp = _sdr_element_fp(_cand_meta)
                    again = next((e for e in refreshed
                                  if _sdr_element_fp(e) == want_fp), None)
                    if again is None:
                        continue
                    best = again
                loc_el = None
                try:
                    cand = page.locator(f'[data-tvuid="{best["uid"]}"]')
                    if await cand.count() > 0:
                        loc_el = cand.first
                except Exception:
                    pass
                if loc_el is None:
                    continue   # stale handle; try the next candidate

                # --- make it visible / reachable ---
                try:
                    if not await loc_el.is_visible(timeout=400):
                        await page.evaluate(EXPOSE_HIDDEN_JS)
                        await asyncio.sleep(0.35)
                    if not await loc_el.is_visible(timeout=300):
                        # Still hidden: it sits inside something collapsed —
                        # an ISI tray, an accordion, a closed sub-menu. Open
                        # the ancestors that own it. Without this the click
                        # lands on a node the page is not showing, no handler
                        # runs, and the row is judged on whatever the next
                        # candidate happens to fire.
                        await loc_el.evaluate("""e => {
                            let node = e;
                            const toOpen = [];
                            while (node && node !== document.body) {
                                node = node.parentElement;
                                if (!node) break;
                                const expanded = node.getAttribute('aria-expanded');
                                const hasPopup = node.getAttribute('aria-haspopup');
                                const cls = (typeof node.className === 'string' ? node.className : '').toLowerCase();
                                if (expanded === 'false' || hasPopup
                                        || /isi|accordion|collapse|drawer|tray/.test(cls)) {
                                    toOpen.unshift(node);
                                }
                            }
                            for (const p of toOpen) {
                                const trig = p.querySelector(
                                    ':scope > button, :scope > a, :scope > [role="button"], button');
                                const t = trig || p;
                                try {
                                    t.dispatchEvent(new MouseEvent('mouseenter', {bubbles: true}));
                                    t.click();
                                } catch (err) {}
                            }
                        }""")
                        await asyncio.sleep(0.6)
                except Exception:
                    pass
                # Clear any consent panel before EVERY click, not only when a
                # hit-test has already failed. OneTrust's preference centre
                # lays a full-page dark filter over everything, and Playwright
                # reports exactly that: "onetrust-pc-dark-filter ... intercepts
                # pointer events". While it is up, no real click can land, so
                # the audit falls back to dispatching on the node — which the
                # tag manager's link triggers ignore, and the row is recorded
                # as firing nothing. Testing the consent links themselves
                # makes this panel appear far more often, so it is closed on
                # the way in to every row.
                try:
                    await page.evaluate(CLOSE_CONSENT_UI_JS)
                except Exception:
                    pass
                # Put the element in the MIDDLE of the viewport, not merely
                # somewhere in it. Sticky headers and pinned ISI banners live
                # at the top and bottom edges, and "scroll into view" happily
                # leaves an element underneath one of them — Playwright's own
                # words on this page: "aaaem-isi-banner subtree intercepts
                # pointer events". The centre is the one band nothing pins
                # itself to.
                try:
                    await loc_el.scroll_into_view_if_needed(timeout=2500)
                    # scrollIntoView scrolls the nearest scrollable ancestor,
                    # which for content inside an ISI tray is that tray and not
                    # the window — the element ends up perfectly centred inside
                    # a box that is itself off-screen, and elementFromPoint at
                    # its centre returns null. Move the window itself until the
                    # element really is in the middle of the viewport, away
                    # from the sticky bars pinned to the top and bottom edges.
                    await loc_el.evaluate("""e => {
                        e.scrollIntoView({block: 'center', inline: 'center'});
                        for (let i = 0; i < 3; i++) {
                            const r = e.getBoundingClientRect();
                            const mid = r.top + r.height / 2;
                            if (mid > innerHeight * 0.3 && mid < innerHeight * 0.7) break;
                            window.scrollBy(0, mid - innerHeight / 2);
                        }
                    }""")
                except Exception:
                    pass
                await asyncio.sleep(0.6)

                # --- drain, then open this case's capture window ---
                try:
                    await page.evaluate(HARVEST_TRACKING_JS)
                    await page.evaluate(READ_CLEAR_WITNESS_JS)
                except Exception:
                    pass
                await asyncio.sleep(0.15)
                bookmark = len(all_requests)
                current_cid["v"] = ci

                hit = None
                try:
                    hit = await loc_el.evaluate("""(e) => {
                        const rectOf = (n) => {
                            let q = n.getBoundingClientRect();
                            if (q.width > 2 && q.height > 2) return q;
                            for (const c of n.getClientRects()) {
                                if (c.width > 2 && c.height > 2) return c;
                            }
                            for (const k of n.children) {
                                const kr = k.getBoundingClientRect();
                                if (kr.width > 2 && kr.height > 2) return kr;
                            }
                            return q;
                        };
                        const r = rectOf(e);
                        if (!r.width || !r.height) return {ok: false, reason: 'no-box'};
                        const c = [[r.x + r.width/2, r.y + r.height/2],
                                   [r.x + r.width*0.25, r.y + r.height/2],
                                   [r.x + r.width*0.75, r.y + r.height/2]];
                        for (const [x, y] of c) {
                            if (x < 0 || y < 0 || x > innerWidth || y > innerHeight) continue;
                            const t = document.elementFromPoint(x, y);
                            if (t && (t === e || e.contains(t) || t.contains(e))) return {ok: true, x, y};
                        }
                        const t0 = document.elementFromPoint(r.x + r.width/2, r.y + r.height/2);
                        return {ok: false,
                                blocker: t0 ? (t0.tagName + '.' + String(t0.className || '').slice(0,44)) : 'null'};
                    }""")
                except Exception:
                    pass
                el_res_blocked = ((hit or {}).get("blocker")
                                  or ("ok" if (hit or {}).get("ok") else "?"))
                # A control with no box at all is not broken — some only
                # materialise at a certain scroll position. The back-to-top
                # arrow is display:none until the page has moved, so at the
                # top of the page it measures zero and cannot be clicked.
                # Move down the page and look again before giving up on it.
                if hit and not hit.get("ok") and hit.get("reason") == "no-box":
                    # A control with no box is not necessarily absent. Two
                    # different situations produce it, and both are ordinary:
                    # the navigation is collapsed behind a menu button (every
                    # nav link then measures zero), or the control only
                    # materialises once the page has been scrolled. Try
                    # opening the menu first, then moving down the page.
                    try:
                        await page.evaluate(EXPOSE_HIDDEN_JS)
                        await page.evaluate("""() => {
                            const sels = [
                                'button[aria-label*="menu" i]', 'button[aria-label*="navigation" i]',
                                'button[class*="hamburger" i]', 'button[class*="menu-toggle" i]',
                                'button[class*="nav-toggle" i]', 'button[class*="navbar-toggle" i]',
                                '[aria-haspopup="true"]', 'button[aria-expanded="false"]',
                                '[class*="menu-icon" i]', '[class*="burger" i]'];
                            for (const sel of sels) {
                                for (const n of document.querySelectorAll(sel)) {
                                    const r = n.getBoundingClientRect();
                                    if (r.width < 3 || r.height < 3) continue;
                                    const blob = ((n.id || '') + ' ' +
                                        (typeof n.className === 'string' ? n.className : '')).toLowerCase();
                                    if (/cookie|consent|onetrust|optanon|privacy/.test(blob)) continue;
                                    try { n.click(); } catch (e) {}
                                }
                            }
                        }""")
                        await asyncio.sleep(0.9)
                        hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                    except Exception:
                        pass
                    for depth in (900, 2000, 4000):
                        if hit and hit.get("ok"):
                            break
                        try:
                            await page.evaluate("window.scrollTo(0, %d)" % depth)
                            await asyncio.sleep(0.7)
                            hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                        except Exception:
                            break
                    el_res_blocked = ((hit or {}).get("reason")
                                      or ("ok" if (hit or {}).get("ok") else "?"))

                # Nothing at all at the element's centre means it is not
                # really on screen, however "visible" it reports as — the ISI
                # links sit inside a tray that is scrolled or collapsed, so
                # their rect lands outside the viewport entirely. Opening the
                # tray puts them where a person would see them.
                if hit and not hit.get("ok") and str(hit.get("blocker")) in ("null", "None", ""):
                    try:
                        await loc_el.evaluate("""e => {
                            let n = e;
                            while (n && n !== document.body) {
                                const cls = (typeof n.className === 'string' ? n.className : '').toLowerCase();
                                const id = (n.id || '').toLowerCase();
                                if (/isi|accordion|collapse|drawer|tray|expand/.test(cls + ' ' + id)) {
                                    const t = n.querySelector('button, [role="button"]');
                                    if (t) { try { t.click(); } catch (err) {} }
                                }
                                n = n.parentElement;
                            }
                        }""")
                        await asyncio.sleep(0.8)
                        await loc_el.evaluate("""e => {
                            e.scrollIntoView({block: 'center', inline: 'center'});
                            const r = e.getBoundingClientRect();
                            const mid = r.top + r.height / 2;
                            if (mid < 0 || mid > innerHeight) window.scrollBy(0, mid - innerHeight / 2);
                        }""")
                        await asyncio.sleep(0.5)
                        hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                        el_res_blocked = ((hit or {}).get("reason")
                                          or ("ok" if (hit or {}).get("ok") else "?"))
                    except Exception:
                        pass

                pierced = False
                if not (hit and hit.get("ok")):
                    try:
                        await page.evaluate(CLOSE_CONSENT_UI_JS)
                        await page.keyboard.press("Escape")
                        await asyncio.sleep(0.2)
                        hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                        pierced = True
                    except Exception:
                        pass

                # The control is open and reachable at this instant; once it
                # is clicked the menu holding it usually closes and there is
                # nothing left to point at. So the picture is taken now.
                shot_click = await _sdr_capture_click_shot(
                    page, SDR_SHOT_DIR, case["excel_row"], len(attempts) + 1,
                    best.get("uid"))

                clicked, click_method = False, ""
                if hit and hit.get("ok"):
                    try:
                        for t in ("mouseMoved", "mousePressed", "mouseReleased"):
                            args = {"type": t, "x": hit["x"], "y": hit["y"]}
                            if t != "mouseMoved":
                                args.update({"button": "left", "clickCount": 1})
                            await cdp.send("Input.dispatchMouseEvent", args)
                            if t == "mousePressed":
                                await asyncio.sleep(0.05)
                        clicked, click_method = True, "cdp-click"
                    except Exception:
                        pass
                if not clicked:
                    try:
                        await loc_el.evaluate("e => e.click()")
                        clicked, click_method = True, "js-click"
                    except Exception:
                        pass
                if pierced:
                    try:
                        await page.evaluate(RESTORE_OVERLAY_JS)
                    except Exception:
                        pass

                verified = False
                if clicked:
                    await asyncio.sleep(0.25)
                    try:
                        w = await page.evaluate(READ_CLEAR_WITNESS_JS)
                        verified = best["uid"] in ((w or {}).get("uids") or [])
                    except Exception:
                        pass

                if not clicked:
                    current_cid["v"] = -1
                    continue   # this candidate is unclickable; try the next one

                # --- wait for the beacon, then close the window ---
                await asyncio.sleep(1.0)
                try:
                    early = await page.evaluate(PEEK_TRACKING_JS)
                except Exception:
                    early = []
                expect = bool(early)
                # A site that tags to two GA4 properties does not always send both
                # hits together — one can arrive seconds after the other. A pure
                # "quiet for 1.2s" rule closes the window in the gap between them
                # and then reports the event as missing from the property that had
                # not answered yet. So when an event is expected, hold the window
                # open for a minimum period regardless of quiet, and only let the
                # quiet detector extend past that.
                started = time.time()
                min_wait = SDR_MIN_BEACON_WAIT if expect else 0.8
                deadline = started + (SDR_BEACON_MAX_WAIT if expect else CLICK_IDLE_WAIT)
                last_n, last_change = len(all_requests), time.time()
                quiet = 2.0 if expect else 0.8
                while time.time() < deadline:
                    await asyncio.sleep(0.25)
                    if len(all_requests) != last_n:
                        last_n, last_change = len(all_requests), time.time()
                        continue
                    if time.time() - started < min_wait:
                        continue
                    if time.time() - last_change >= quiet:
                        break
                current_cid["v"] = -1

                # A click that produced nothing at all, made by dispatching on
                # the node rather than through the browser's input pipeline,
                # is usually not a real "nothing" — tag managers hang their
                # link triggers off trusted events, so an untrusted click runs
                # GA4's own auto-events and none of the site's. Scroll it to
                # the middle of the viewport, where nothing sticky overlaps,
                # and try once more for real before believing the result.
                def _fired_expected():
                    want = _norm_str(case["expected_event"])
                    for r in all_requests[bookmark:]:
                        if '/g/collect' not in r["url"]:
                            continue
                        for ev in parse_ga4_event(r["url"], r["post"]):
                            if _norm_str(ev.get("event")) == want:
                                return True
                    return False

                if (not _fired_expected()
                        and click_method and 'cdp-click' not in click_method):
                    try:
                        await loc_el.evaluate(
                            "e => e.scrollIntoView({block: 'center', inline: 'center'})")
                        await asyncio.sleep(0.5)
                        retry_hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                        if retry_hit and retry_hit.get("ok"):
                            for t in ("mouseMoved", "mousePressed", "mouseReleased"):
                                args = {"type": t, "x": retry_hit["x"], "y": retry_hit["y"]}
                                if t != "mouseMoved":
                                    args.update({"button": "left", "clickCount": 1})
                                await cdp.send("Input.dispatchMouseEvent", args)
                                if t == "mousePressed":
                                    await asyncio.sleep(0.05)
                            click_method += "+cdp-retry"
                            await asyncio.sleep(SDR_MIN_BEACON_WAIT)
                        try:
                            await page.evaluate(RESTORE_OVERLAY_JS)
                        except Exception:
                            pass
                    except Exception:
                        pass

                window_reqs = all_requests[bookmark:]
                try:
                    api_caps = await page.evaluate(HARVEST_TRACKING_JS)
                except Exception:
                    api_caps = []

                # --- collect what actually fired ---
                net_events = []
                for req in window_reqs:
                    ul = req["url"].lower()
                    if "/g/collect" in ul or (("google-analytics.com" in ul
                                               or "analytics.google.com" in ul) and "collect" in ul):
                        for ev in parse_ga4_event(req["url"], req["post"]):
                            if ev.get("event") and not _is_noise_event(ev["event"]):
                                net_events.append(ev)
                                if ev.get("measurement_id"):
                                    detected_ids.add(ev["measurement_id"])

                # Nothing from the page's own JS is used to judge a row —
                # only what went over the wire, which is what Omnibug shows.
                # Only hits on the selected GA4 property count as evidence.
                if ga4_id and ga4_mode != "any":
                    scoped = [e for e in net_events if e.get("measurement_id") == ga4_id]
                else:
                    scoped = list(net_events)

                expected_event = case["expected_event"]
                cands = [e for e in scoped if _norm_str(e.get("event")) == _norm_str(expected_event)]

                reasons, param_diff, notes = [], [], []
                matched = None
                if cands:
                    # Several hits can carry the same event name; judge against the
                    # one that satisfies the most of this row's expectations.
                    def fit(ev):
                        return sum(1 for p, v in case["expected_params"].items()
                                   if _sdr_param_matches(
                                       p, v, _sdr_actual_value(ev.get("params") or {}, p), page_url))
                    matched = max(cands, key=fit)

                if matched is None:
                    fired = sorted({e.get("event", "") for e in scoped})
                    other_prop = sorted({e.get("measurement_id", "") for e in net_events
                                         if _norm_str(e.get("event")) == _norm_str(expected_event)})
                    if other_prop:
                        reasons.append(
                            f"'{expected_event}' fired only on {', '.join(other_prop)}"
                            f" — not on the selected property {ga4_id}")
                    else:
                        reasons.append(
                            f"Expected event '{expected_event}' did not fire"
                            + (f". Events seen: {', '.join(fired)}" if fired else " (no GA4 event fired)"))
                else:
                    actual_params = matched.get("params") or {}
                    for pname, pexp in case["expected_params"].items():
                        pact = _sdr_actual_value(actual_params, pname)
                        ok = _sdr_param_matches(pname, pexp, pact, page_url)
                        cosmetic = False
                        if not ok and _sdr_cosmetic_match(pexp, pact):
                            # Same value, written differently. Counts as a pass.
                            ok = True
                            cosmetic = True
                            notes.append(f"{pname}: SDR says '{pexp}', site sends "
                                         f"'{_sdr_show(pact)}' (spacing/punctuation only)")

                        param_diff.append({"param": pname, "expected": pexp,
                                           "actual": "" if pact is None else str(pact),
                                           "match": ok, "cosmetic": cosmetic})
                        if not ok:
                            if pact is None or str(pact) == "":
                                # Exactly what a person reading Omnibug would
                                # conclude: the parameter is not on this hit. If it
                                # rides on some other hit, say which — that is
                                # still visible in Omnibug and is a real clue.
                                anywhere = sorted({
                                    e.get("event", "") for e in net_events
                                    if _sdr_actual_value(e.get("params") or {}, pname) not in (None, "")
                                })
                                if anywhere:
                                    reasons.append(
                                        f"{pname}: expected '{pexp}' but not sent on this event "
                                        f"(it is on the {', '.join(anywhere)} hit)")
                                else:
                                    reasons.append(
                                        f"{pname}: expected '{pexp}' but not present on the GA4 hit")
                            else:
                                note = ""
                                # When the value the site sent matches THIS row's
                                # own button name but the SDR's expected value does
                                # not, the sheet is almost certainly wrong — these
                                # templates are filled by copying a row and the
                                # text field gets left behind. Saying so turns an
                                # unexplained failure into a one-line sheet fix.
                                if pname in ("link_text", "link_url"):
                                    own = _norm_str(case.get("button_name"))
                                    def _tk(v):
                                        return {t for t in re.split(r'[^a-z0-9]+', _norm_str(v)) if len(t) > 2}
                                    a_t, o_t, e_t = _tk(pact), _tk(own), _tk(pexp)
                                    if o_t:
                                        a_fit = len(a_t & o_t) / len(o_t)
                                        e_fit = len(e_t & o_t) / len(o_t)
                                        if a_fit >= 0.8 and e_fit < 0.5:
                                            note = (" — note: what fired matches this row's own"
                                                    " Link/Button Name, so the SDR's expected value"
                                                    " looks copied from another row")
                                # A URL that does not match is a failure. Saying
                                # that only the host differs points straight at
                                # the fix: the sheet was written against another
                                # environment and never updated.
                                if (pname in ("link_url", "page_location", "file_location")
                                        and _sdr_url_host_only_diff(pexp, pact)):
                                    note += (" — only the host differs, so the SDR still points"
                                             " at a different environment than the page tested")
                                reasons.append(
                                    f"{pname}: expected '{pexp}' but got "
                                    f"'{_sdr_show(pact)}'{note}")

                status = "PASS" if not reasons else "FAIL"
                if status == "PASS" and not verified:
                    # Everything matched, but we could not prove the click landed on
                    # this exact element — say so rather than quietly passing it.
                    notes.append("click could not be verified on this exact element")
                # Cosmetic notes ride along with the verdict so a pass still shows
                # what differed, without turning it into a failure.
                if notes:
                    prefix = "Passed with notes — " if status == "PASS" else "Also (not counted as failure) — "
                    reasons.append(prefix + "; ".join(notes))

                if os.environ.get("TV_DEBUG_CAND"):
                    sys.stdout.write(
                        "      CAND[%d] zone=%s hints=%s method=%s blocked=%r -> event=%r params_ok=%d/%d" % (
                            _cand_i, best.get('zone'),
                            [h for h in (best.get('hints') or [])
                             if h in ('isi', 'header', 'footer', 'nav', 'sticky')],
                            click_method, el_res_blocked,
                            (matched or {}).get('event'),
                            sum(1 for d in param_diff if d["match"]), len(param_diff))
                        + os.linesep)
                    sys.stdout.flush()
                attempts.append({
                    # Rank by: did the expected event fire, then how many of
                    # the row's parameters agreed. The candidate that best
                    # satisfies the row is the control the row meant.
                    "fit": (1 if matched is not None else 0,
                            sum(1 for d in param_diff if d["match"]),
                            -len(reasons)),
                    "status": status, "reasons": list(reasons),
                    "matched": matched, "param_diff": list(param_diff),
                    "scoped": list(scoped), "best": best,
                    "click_method": click_method, "verified": verified,
                    "shot_click": shot_click,
                })
                if status == "PASS":
                    break
                # If another candidate is still to be tried, the loop reloads
                # the page for it rather than continuing from this click.

            if not attempts:
                _record(case, "FAIL", "Element found but could not be clicked")
                sys.stdout.write(f"[SDR]   [{ci+1}/{len(page_cases)}] \"{label}\" -> CLICK FAILED\n")
                sys.stdout.flush()
                continue

            pick = max(attempts, key=lambda a: a["fit"])
            status = pick["status"]
            reasons = pick["reasons"]
            matched = pick["matched"]
            param_diff = pick["param_diff"]
            scoped = pick["scoped"]
            best = pick["best"]
            click_method = pick["click_method"]
            verified = pick["verified"]
            shot_click = pick.get("shot_click", "")
            if len(attempts) > 1:
                reasons.append(f"(tried {len(attempts)} candidate elements; reporting the closest fit)")

            tested_keys.add(_sweep_key(best))
            _record(case, status, "; ".join(reasons), {
                "actual_event": (matched or {}).get("event", ""),
                "actual_params": (matched or {}).get("params", {}),
                "fired_events": sorted({e.get("event", "") for e in scoped}),
                "param_diff": param_diff,
                "matched_element": best.get("text", "")[:60],
                "measurement_id": (matched or {}).get("measurement_id", ""),
                "click_method": click_method,
                "click_verified": verified,
                "shot_click": shot_click,
            })
            # A page can hold dozens of rows and take ten minutes; waiting for
            # the page to end before saving would still lose most of a run
            # that dies part-way through one.
            if (ci + 1) % 5 == 0:
                _flush(partial=True)

            mark = "PASS" if status == "PASS" else "FAIL"
            sys.stdout.write(f"[SDR]   [{ci+1}/{len(page_cases)}] \"{label}\" -> {mark}"
                             f"{'' if status == 'PASS' else ' | ' + '; '.join(reasons)[:110]}\n")
            sys.stdout.flush()

            try:
                await page.keyboard.press("Escape")
            except Exception:
                pass
            try:
                for extra_pg in context.pages:
                    if extra_pg != page:
                        await extra_pg.close()
            except Exception:
                pass
            # A click may have escaped the blocker via a JS redirect.
            try:
                if _norm_url(page.url) != _norm_url(page_url):
                    await page.goto(page_url, wait_until="domcontentloaded", timeout=30000)
                    await asyncio.sleep(1.5)
                    await page.evaluate(BLOCK_NAVIGATION_JS)
                    await page.evaluate(INSTRUMENT_TRACKING_JS)
                    await page.evaluate(EXPOSE_HIDDEN_JS)
                    await asyncio.sleep(0.5)
                    discovered = {}
                    _absorb(await page.evaluate(DISCOVER_CLICKABLES_JS))
                    elements = list(discovered.values())
            except Exception:
                pass

        # ---- what the SDR never mentioned ----
        # The SDR says what should be tagged. It cannot say what was left out
        # of it, and a button that is on the page but not in the sheet is the
        # gap a manual pass is looking for. So every remaining control is
        # clicked too, and reported as tagged (with what it sent) or not.
        try:
            leftovers = [e for e in await _fresh_elements()
                         if _sweep_key(e) not in tested_keys]
        except Exception:
            leftovers = []
        if leftovers:
            sys.stdout.write("[SDR] %d control(s) on this page are not in the SDR "
                             "— checking each one" % len(leftovers) + os.linesep)
            sys.stdout.flush()

        seen_extra = set()
        page_extras = []
        for xi, el in enumerate(leftovers):
            key = _sweep_key(el)
            if key in seen_extra:
                continue
            seen_extra.add(key)
            name = (el.get("text") or "")[:60]
            row_id = "x%d" % (len(extras) + 1)
            try:
                # Fresh page per control. Clicking many things into one page
                # state leaves the site re-rendered and its triggers detached,
                # and the button then looks untagged when it is not.
                # A page load per control is a lot to ask of a CDN in a tight
                # loop; a breath between them keeps the audit from looking
                # like something worth rate-limiting.
                await asyncio.sleep(1.2)
                try:
                    if await _sdr_goto(page, page_url, tries=2):
                        continue        # the site is not answering; skip this one
                    await _sdr_try_login(page, auth_user, auth_pass)
                    await asyncio.sleep(1.5)
                    await page.evaluate(BLOCK_NAVIGATION_JS)
                    await page.evaluate(INSTRUMENT_TRACKING_JS)
                    await page.evaluate(EXPOSE_HIDDEN_JS)
                    await asyncio.sleep(1.0)
                except Exception:
                    pass

                # The page re-renders after every click, so the element is
                # located again rather than trusted from the earlier scan.
                current = next((e for e in await _fresh_elements()
                                if _sweep_key(e) == key), None)
                if not current:
                    continue
                uid = current.get("uid")
                loc_el = page.locator('[data-tvuid="%s"]' % uid).first
                shot = await _sdr_capture_click_shot(page, SDR_SHOT_DIR, row_id, 1, uid)

                bookmark = len(all_requests)

                async def _press():
                    """One attempt at a real click on this control.

                    A tag manager hangs its link triggers off trusted events,
                    so a dispatched click runs GA4's own auto-events and none
                    of the site's — which reads back as "not tagged" for a
                    button that is tagged. The pointer path is tried first and
                    the dispatch is only a fallback.
                    """
                    try:
                        await page.evaluate(CLOSE_CONSENT_UI_JS)
                    except Exception:
                        pass
                    try:
                        await loc_el.evaluate(
                            "e => e.scrollIntoView({block: 'center', inline: 'center'})")
                        await asyncio.sleep(0.3)
                    except Exception:
                        pass
                    hit = None
                    try:
                        hit = await loc_el.evaluate(PIERCE_OVERLAY_JS)
                    except Exception:
                        pass
                    method = ""
                    if hit and hit.get("ok"):
                        try:
                            for t in ("mouseMoved", "mousePressed", "mouseReleased"):
                                a = {"type": t, "x": hit["x"], "y": hit["y"]}
                                if t != "mouseMoved":
                                    a.update({"button": "left", "clickCount": 1})
                                await cdp.send("Input.dispatchMouseEvent", a)
                                if t == "mousePressed":
                                    await asyncio.sleep(0.05)
                            method = "cdp-click"
                        except Exception:
                            pass
                    if not method:
                        try:
                            await loc_el.evaluate("e => e.click()")
                            method = "js-click"
                        except Exception:
                            pass
                    try:
                        await page.evaluate(RESTORE_OVERLAY_JS)
                    except Exception:
                        pass
                    return method

                click_method = await _press()
                if not click_method:
                    continue

                # Same wait the rows get: hold the window open when something
                # is on its way, close it quickly when nothing is.
                await asyncio.sleep(1.0)
                try:
                    expect = bool(await page.evaluate(PEEK_TRACKING_JS))
                except Exception:
                    expect = False
                started = time.time()
                min_wait = SDR_MIN_BEACON_WAIT if expect else 0.8
                deadline = started + (SDR_BEACON_MAX_WAIT if expect else CLICK_IDLE_WAIT)
                last_n, last_change = len(all_requests), time.time()
                quiet = 2.0 if expect else 0.8
                while time.time() < deadline:
                    await asyncio.sleep(0.25)
                    if len(all_requests) != last_n:
                        last_n, last_change = len(all_requests), time.time()
                        continue
                    if time.time() - started < min_wait:
                        continue
                    if time.time() - last_change >= quiet:
                        break

                def _collect():
                    got = []
                    for req in all_requests[bookmark:]:
                        ul = req["url"].lower()
                        if "/g/collect" in ul or (("google-analytics.com" in ul
                                                   or "analytics.google.com" in ul)
                                                  and "collect" in ul):
                            for ev in parse_ga4_event(req["url"], req["post"]):
                                if ev.get("event") and not _is_noise_event(ev["event"]):
                                    got.append(ev)
                                    if ev.get("measurement_id"):
                                        detected_ids.add(ev["measurement_id"])
                    if ga4_id and ga4_mode != "any":
                        return [e for e in got if e.get("measurement_id") == ga4_id]
                    return got

                scoped_x = _collect()
                # Nothing here names this control, and the click went through
                # the dispatch fallback: that is the signature of a trigger
                # that only listens to real pointer events. Worth one honest
                # second try before calling the button untagged.
                _, sc = _sweep_pick_event(scoped_x, name, current.get("href", ""))
                if sc < 3 and click_method != "cdp-click":
                    if await _press():
                        await asyncio.sleep(1.0)
                        started = time.time()
                        while time.time() - started < SDR_MIN_BEACON_WAIT:
                            await asyncio.sleep(0.3)
                            _, sc2 = _sweep_pick_event(
                                _collect(), name, current.get("href", ""))
                            if sc2 >= 3:
                                break
                        scoped_x = _collect()

                # The verdict waits until the whole page is swept, because
                # telling a click's own hit from a timer's needs to know what
                # fires in every window.
                rec = {
                    "excel_row": row_id,
                    "page_url": page_url,
                    "location": current.get("zone", "") or "",
                    "button_name": name,
                    "link_url": current.get("href", "") or "",
                    "status": "NOT TAGGED",
                    "actual_event": "", "actual_params": {}, "measurement_id": "",
                    "fired_events": [], "param_diff": [], "reason": "",
                    "shot_click": shot,
                    "_raw_events": scoped_x,
                }
                extras.append(rec)
                page_extras.append(rec)
                sys.stdout.write("[SDR]   not-in-SDR [%d/%d] \"%s\" -> %d hit(s)%s"
                                 % (xi + 1, len(leftovers), name[:34],
                                    len(scoped_x), os.linesep))
                sys.stdout.flush()
            except Exception as e:
                sys.stdout.write("[SDR]   not-in-SDR \"%s\" -> could not test (%s)%s"
                                 % (name[:34], str(e)[:60], os.linesep))
                sys.stdout.flush()

        _sweep_resolve(page_extras)
        if page_extras:
            n_tag = sum(1 for e in page_extras if e["status"] == "TAGGED")
            sys.stdout.write("[SDR] Not in SDR on this page: %d tagged, %d not tagged"
                             % (n_tag, len(page_extras) - n_tag) + os.linesep)
            sys.stdout.flush()

        # Page finished — checkpoint before moving to the next.
        _flush(partial=True)

    # Draw the read-out panels now that every row has its captured hit. Doing
    # it here, rather than during the audit, keeps the clicking at full speed.
    try:
        n_panels = await _sdr_render_param_panels(
            browser, results + [e for e in extras if e.get("actual_event")],
            SDR_SHOT_DIR)
        if n_panels:
            sys.stdout.write("[SDR] Rendered %d parameter panels" % n_panels + os.linesep)
            sys.stdout.flush()
    except Exception:
        pass
    _flush(partial=False)
    try:
        await context.close()
    except Exception:
        pass

    p = sum(1 for r in results if r["status"] == "PASS")
    f = sum(1 for r in results if r["status"] == "FAIL")
    s = sum(1 for r in results if r["status"] == "SKIPPED")
    sys.stdout.write(f"[SDR] DONE — {p} PASS, {f} FAIL, {s} SKIPPED (of {len(results)})\n")
    if detected_ids:
        sys.stdout.write(f"[SDR] GA4 properties seen on site: {', '.join(sorted(detected_ids))}\n")
    sys.stdout.flush()
    return {"meta": parsed, "results": results, "not_in_sdr": extras,
            "detected_ga4_ids": sorted(detected_ids), "ga4_id": ga4_id}


# How wide a screenshot is drawn inside a cell. Excel anchors a picture to a
# cell rather than putting it in one, so the row has to be made tall enough
# to hold it or the images overlap each other.
SHOT_W = 300
SHOT_COLS = ("Tool QA — Params Fired", "Tool QA — Click Screenshot")


# A picture is anchored to a cell rather than held in one, so sorting or
# filtering the sheet leaves it where it was. The cell itself therefore also
# carries a link to the same file, which does travel with the row — and the
# picture is nudged down far enough to leave that line readable.
LINK_BAND_PX = 17
_EMU_PER_PX = 9525


def _embed_shot(ws, path, row, col, label, width=SHOT_W):
    """Put one PNG at a cell: anchored picture, plus a link on the cell.

    Returns the drawn height in pixels so the caller can size the row to the
    tallest picture on it.
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import (
        AnchorMarker, OneCellAnchor)
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.styles import Font
    if not path or not os.path.exists(path):
        return 0
    try:
        full = os.path.abspath(path)
        cell = ws.cell(row, col)
        cell.value = label
        try:
            cell.hyperlink = "file:///" + full.replace("\\", "/")
            cell.font = Font(color="0563C1", underline="single", size=9)
        except Exception:
            pass

        img = XLImage(path)
        w, h = img.width or width, img.height or width
        if w > width:
            h = int(h * (width / float(w)))
            w = width
        img.width, img.height = w, h
        # Sit the picture just below the link line, in the same cell.
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=col - 1, colOff=0,
                               row=row - 1, rowOff=LINK_BAND_PX * _EMU_PER_PX),
            ext=XDRPositiveSize2D(int(w * _EMU_PER_PX), int(h * _EMU_PER_PX)))
        ws.add_image(img)
        return int(h) + LINK_BAND_PX
    except Exception:
        return 0


def _pick_shot_columns(ws, header_row, count=2):
    """Where the evidence columns can go: tight against the table, and empty.

    Two things go wrong on real sheets, and they pull in opposite directions.
    `ws.max_column` counts a cell carrying only formatting — one SDR is styled
    out to column X with nothing in it after O — so appending there strands the
    evidence nine columns from the table. But a sheet can also hold something
    off to the right with no header over it, and starting right after the last
    header would then write over it.

    So: start immediately after the last header, and take the first run of
    `count` columns that is empty from the header row down. Tight when there
    is nothing in the way, and never destructive when there is.
    """
    last_hdr = 0
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value or "").strip():
            last_hdr = c
    start = (last_hdr or ws.max_column) + 1
    bottom = min(ws.max_row, 5000)

    def empty(col):
        for r in range(header_row, bottom + 1):
            if str(ws.cell(r, col).value or "").strip():
                return False
        return True

    limit = max(ws.max_column, start) + 40
    while start < limit:
        if all(empty(start + i) for i in range(count)):
            return start
        start += 1
    return ws.max_column + 1


def _add_shot_columns(ws, header_row, results, row_of=lambda r: r.get("excel_row")):
    """Add the two picture columns at the end and fill them.

    They go at the end rather than being inserted, so every column the sheet
    already had keeps its position — a row's data is where it always was, with
    the evidence appended beside it.
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    if not any(r.get("shot_params") or r.get("shot_click") for r in results):
        return
    c_par = _pick_shot_columns(ws, header_row, 2)
    c_click = c_par + 1
    for col, name in ((c_par, SHOT_COLS[0]), (c_click, SHOT_COLS[1])):
        h = ws.cell(header_row, col)
        h.value = name
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="4472C4")
        h.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = SHOT_W / 7.0

    for r in results:
        row = row_of(r)
        if not row:
            continue
        tall = max(
            _embed_shot(ws, r.get("shot_params"), row, c_par, "open params PNG"),
            _embed_shot(ws, r.get("shot_click"), row, c_click, "open screenshot PNG"))
        if tall:
            # px -> points, and never shrink a row the sheet already sized.
            want = tall * 0.75 + 4
            cur = ws.row_dimensions[row].height or 0
            ws.row_dimensions[row].height = max(cur, want)


QA_DETAIL_SHEET = "Tool QA Details"


def _write_not_in_sdr_sheet(wb, extras):
    """One sheet listing the controls the SDR never mentioned.

    A verdict sheet can only report on rows somebody wrote down. This is the
    other half of coverage: what is on the page, whether it is tagged, and
    what it sent — with a picture of the control either way.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if not extras:
        return
    if NOT_IN_SDR_SHEET in wb.sheetnames:
        del wb[NOT_IN_SDR_SHEET]
    ws = wb.create_sheet(NOT_IN_SDR_SHEET)

    headers = ["Page URL", "Where", "Button / Link", "Link URL", "Tagged?",
               "GA4 Event", "GA4 Property", "Parameters Sent", "Note"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        h = ws.cell(1, c)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="4472C4")
        h.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    fill_yes = PatternFill("solid", fgColor="C6EFCE")
    fill_no = PatternFill("solid", fgColor="FFC7CE")
    font_yes = Font(color="006100", bold=True)
    font_no = Font(color="9C0006", bold=True)

    ordered = sorted(extras, key=lambda e: (e.get("page_url", ""),
                                            e.get("location", ""),
                                            e.get("button_name", "")))
    for e in ordered:
        params = "; ".join("%s: %s" % (k, v)
                           for k, v in (e.get("actual_params") or {}).items())
        ws.append([
            e.get("page_url", ""),
            e.get("location", "") or "-",
            e.get("button_name", "") or "(no label)",
            e.get("link_url", "") or "-",
            "Tagged" if e.get("status") == "TAGGED" else "Not tagged",
            e.get("actual_event", "") or "-",
            e.get("measurement_id", "") or "-",
            params or "-",
            e.get("reason", "") or "",
        ])
        row = ws.max_row
        cell = ws.cell(row, 5)
        if e.get("status") == "TAGGED":
            cell.fill, cell.font = fill_yes, font_yes
        else:
            cell.fill, cell.font = fill_no, font_no

    for i, w in enumerate([40, 12, 34, 40, 12, 22, 16, 60, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        r[7].alignment = Alignment(wrap_text=True, vertical="top")
        r[8].alignment = Alignment(wrap_text=True, vertical="top")

    # Same two evidence columns as the SDR sheet: the control, and what it
    # sent if it sent anything.
    line_of = {id(e): i + 2 for i, e in enumerate(ordered)}
    _add_shot_columns(ws, 1, ordered, row_of=lambda e: line_of.get(id(e)))


def write_sdr_verdicts(sdr_path, meta, results, out_path, qa_column="", extras=None):
    """Write the verdicts back into a copy of the operator's own SDR.

    Two things have to be true at once, so the workbook carries both:

    * The SDR sheet itself stays EXACTLY as it was handed over — same tabs,
      same columns, same headers, same order. Nothing is inserted. Only the QA
      column the team already uses gets filled in with Pass / Fail, so the
      sheet drops straight back into their process. The reason a row failed is
      attached as a cell note on that verdict, which carries the detail
      without altering the layout.

    * The full evidence still has to be somewhere, so it goes into a separate
      "Tool QA Details" sheet: what fired, on which property, which parameters
      differed and how the element was matched.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(sdr_path)
    sheet = meta.get("sheet") or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet '{sheet}' not in workbook")
    ws = wb[sheet]
    header_row = meta.get("header_row", 2)
    qa_cols = meta.get("qa_columns", {}) or {}

    # Prefer the requested QA column, else a Live/Prod one, else any QA column.
    target_col = None
    if qa_column and qa_column in qa_cols:
        target_col = qa_cols[qa_column]
    if target_col is None:
        for name, col in qa_cols.items():
            n = name.lower()
            if 'live' in n or 'prod' in n:
                target_col = col
                break
    if target_col is None and qa_cols:
        target_col = list(qa_cols.values())[0]
    if target_col is None:
        # No QA column at all — only then is one added, at the end, because
        # there is nowhere else for a verdict to go.
        target_col = ws.max_column + 1
        ws.cell(header_row, target_col).value = "Live/Prod QA"
        ws.cell(header_row, target_col).font = Font(bold=True)

    fill_pass = PatternFill("solid", fgColor="C6EFCE")
    fill_note = PatternFill("solid", fgColor="FFF2CC")
    fill_fail = PatternFill("solid", fgColor="FFC7CE")
    fill_skip = PatternFill("solid", fgColor="E7E6E6")
    font_pass = Font(color="006100", bold=True)
    font_fail = Font(color="9C0006", bold=True)
    font_skip = Font(color="595959", bold=True)

    for r in results:
        row = r.get("excel_row")
        if not row:
            continue
        status = r.get("status", "")
        verdict = {"PASS": "Pass", "FAIL": "Fail", "SKIPPED": "Not Tested"}.get(status, status)
        reason = (r.get("reason") or "").strip()
        cell = ws.cell(row, target_col)
        cell.value = verdict
        if status == "PASS":
            cell.fill = fill_note if reason else fill_pass
            cell.font = font_pass
        elif status == "FAIL":
            cell.fill, cell.font = fill_fail, font_fail
        else:
            cell.fill, cell.font = fill_skip, font_skip
        # The explanation rides along as a note so the columns stay untouched.
        if reason:
            try:
                c = Comment(reason[:1500], "Tag Validator")
                c.width, c.height = 420, 160
                cell.comment = c
            except Exception:
                pass

    # ---- the pictures, appended beside the row they belong to ----
    _add_shot_columns(ws, header_row, results)

    # ---- the evidence sheet ----
    if QA_DETAIL_SHEET in wb.sheetnames:
        del wb[QA_DETAIL_SHEET]
    det = wb.create_sheet(QA_DETAIL_SHEET)
    headers = ["SDR Row", "Page URL", "Location", "Link/Button Name",
               "Expected Event", "Actual Event", "GA4 Property", "Result",
               "Why It Failed / Notes", "Parameter Differences",
               "All Events Fired", "Element Clicked", "Click Method",
               "Click Verified"]
    det.append(headers)
    for c in range(1, len(headers) + 1):
        h = det.cell(1, c)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="4472C4")
        h.alignment = Alignment(vertical="center")
    det.freeze_panes = "A2"

    for r in sorted(results, key=lambda x: x.get("excel_row") or 0):
        diffs = "; ".join(
            f"{d['param']}: expected '{d['expected']}', got '{d['actual'] or '(not sent)'}'"
            + (" [spacing only — passed]" if d.get("cosmetic") else "")
            for d in r.get("param_diff", []) if not d["match"] or d.get("cosmetic")
        )
        det.append([
            r.get("excel_row"),
            r.get("page_url", ""),
            r.get("location", ""),
            r.get("button_name", ""),
            r.get("expected_event", ""),
            r.get("actual_event", "") or "-",
            r.get("measurement_id", "") or "-",
            {"PASS": "Pass", "FAIL": "Fail", "SKIPPED": "Not Tested"}.get(r.get("status"), r.get("status")),
            r.get("reason", "") or "",
            diffs or "-",
            ", ".join(r.get("fired_events", [])) or "-",
            r.get("matched_element", "") or "-",
            r.get("click_method", "") or "-",
            "Yes" if r.get("click_verified") else "No",
        ])

    widths = [9, 42, 22, 32, 20, 20, 18, 11, 72, 60, 30, 34, 14, 13]
    for i, w in enumerate(widths, start=1):
        det.column_dimensions[get_column_letter(i)].width = w
    for row in det.iter_rows(min_row=2, max_row=det.max_row):
        row[8].alignment = Alignment(wrap_text=True, vertical="top")
        row[9].alignment = Alignment(wrap_text=True, vertical="top")
        v = row[7].value
        if v == "Fail":
            row[7].fill, row[7].font = fill_fail, font_fail
        elif v == "Pass":
            row[7].fill, row[7].font = fill_pass, font_pass
        else:
            row[7].fill, row[7].font = fill_skip, font_skip

    _write_not_in_sdr_sheet(wb, extras or [])

    wb.save(out_path)
    return out_path


async def run_batch(browser, urls_batch, start_index, total, mode=None):
    if mode == 'clicks':
        # Click audit is sequential per URL (clicking is inherently serial)
        results = []
        for i, url in enumerate(urls_batch):
            results.append(await validate_clicks(browser, url, start_index + i, total))
        return results
    fn = validate_tags
    tasks = [fn(browser, url, start_index + i, total) for i, url in enumerate(urls_batch)]
    return await asyncio.gather(*tasks)


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", help="Mode: 'tealium', 'ga4', 'clicks' or 'sdr'")
    parser.add_argument("--sdr", help="Path to SDR Excel file (for --mode sdr)")
    parser.add_argument("--start-url", dest="start_url",
                        help="Start URL for SDR audit (defaults to first http URL found in SDR)")
    parser.add_argument("--sdr-sheet", dest="sdr_sheet",
                        help="Which sheet of the SDR workbook to validate")
    parser.add_argument("--ga4-id", dest="ga4_id",
                        help="GA4 measurement ID to validate against, e.g. G-XXXXXXX")
    parser.add_argument("--ga4-mode", dest="ga4_mode", default="specific",
                        help="'specific' = only the chosen GA4 property counts; 'any' = any property")
    parser.add_argument("--qa-column", dest="qa_column",
                        help="Which QA column to fill in the SDR (e.g. 'Live/Prod QA')")
    parser.add_argument("--resume", dest="resume", action="store_true",
                        help="Continue a previous SDR run instead of re-testing rows already done")
    parser.add_argument("--list-sdr-sheets", dest="list_sdr_sheets_path",
                        help="Print the sheets in an SDR workbook as JSON, then exit")
    args = parser.parse_args()

    # Sheet discovery for the UI picker — cheap, no browser needed.
    if args.list_sdr_sheets_path:
        try:
            print(json.dumps(list_sdr_sheets(args.list_sdr_sheets_path), indent=2))
        except Exception as e:
            print(json.dumps({"error": str(e)[:200]}))
        return

    # ---- SDR mode runs an entirely different audit pipeline ----
    if args.mode == 'sdr':
        sdr_path = args.sdr or 'sdr_input.xlsx'
        if not os.path.exists(sdr_path):
            sys.stdout.write(f"[SDR] SDR file not found: {sdr_path}\n"); sys.stdout.flush()
            return

        sheet_name = (getattr(args, 'sdr_sheet', '') or '').strip() or None
        ga4_id = (getattr(args, 'ga4_id', '') or '').strip()
        ga4_mode = (getattr(args, 'ga4_mode', '') or 'specific').strip()

        start_url = (args.start_url or '').strip()
        if not start_url:
            probe = parse_sdr_file(sdr_path, sheet_name)
            for c in probe.get("cases", []):
                pu = c.get("page_url", "")
                if pu and pu.lower().startswith('http'):
                    start_url = pu
                    break
            if not start_url:
                for c in probe.get("cases", []):
                    lu = c.get("link_url", "")
                    if lu.startswith('http'):
                        from urllib.parse import urlparse as _up
                        _p = _up(lu)
                        start_url = f"{_p.scheme}://{_p.netloc}/"
                        break
        if not start_url:
            sys.stdout.write("[SDR] Could not determine start URL — pass --start-url\n"); sys.stdout.flush()
            return

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage'])
            sdr_out = await validate_sdr(browser, sdr_path, start_url,
                                         sheet_name=sheet_name,
                                         ga4_id=ga4_id, ga4_mode=ga4_mode,
                                         resume=bool(getattr(args, 'resume', False)),
                                         qa_column=(getattr(args, 'qa_column', '') or ''))
            await browser.close()

        meta = sdr_out.get("meta", {})
        results = sdr_out.get("results", [])

        # Group failures by parameter. A convention mismatch (the SDR writes
        # "navigation_link", the site sends "navigation link") shows up as a
        # failure on every single row, which reads like a hundred separate
        # bugs. Naming the pattern once tells the reader it is one fix.
        from collections import Counter as _Ctr
        param_fails = _Ctr()
        sep_only = _Ctr()
        for r in results:
            for d in r.get("param_diff", []):
                if d["match"]:
                    continue
                param_fails[d["param"]] += 1
                e_norm = re.sub(r'[\s_\-]+', ' ', str(d["expected"]).strip().lower())
                a_norm = re.sub(r'[\s_\-]+', ' ', str(d["actual"] or '').strip().lower())
                if a_norm and e_norm == a_norm:
                    sep_only[d["param"]] += 1
        patterns = []
        for pname, n in param_fails.most_common():
            entry = {"param": pname, "failed_rows": n,
                     "separator_only": sep_only.get(pname, 0)}
            if sep_only.get(pname, 0) == n:
                entry["note"] = ("values are identical apart from separators "
                                 "(SDR uses underscores, site sends spaces) — one convention fix")
            elif not sep_only.get(pname, 0):
                entry["note"] = "genuinely different or not sent"
            patterns.append(entry)
        if patterns:
            sys.stdout.write("[SDR] --- failure patterns (one root cause each) ---\n")
            for p in patterns[:8]:
                sys.stdout.write(f"[SDR]   {p['param']}: {p['failed_rows']} row(s)"
                                 f"{' — ' + p['note'] if p.get('note') else ''}\n")
            sys.stdout.flush()

        with open('sdr_results.json', 'w', encoding='utf-8') as f:
            json.dump({"generated": datetime.datetime.now().isoformat(),
                       "start_url": start_url,
                       "sheet": meta.get("sheet", ""),
                       "ga4_id": sdr_out.get("ga4_id", ""),
                       "detected_ga4_ids": sdr_out.get("detected_ga4_ids", []),
                       "not_in_sdr": sdr_out.get("not_in_sdr", []),
                       "failure_patterns": patterns,
                       "partial": False,
                       "completed": len(results),
                       "total_cases": len(results),
                       "results": results}, f, indent=2, default=str)

        # 1) Flat report — one row per SDR case, with the reason spelled out.
        flat_rows = []
        for r in results:
            failed = [d for d in r.get("param_diff", []) if not d["match"]]
            flat_rows.append({
                "SDR Row": r["excel_row"],
                "Page URL": r["page_url"],
                "Location": r["location"],
                "Link/Button Name": r["button_name"],
                "Expected Event": r["expected_event"],
                "Actual Event": r.get("actual_event", "") or "-",
                "GA4 Property": r.get("measurement_id", "") or "-",
                "Result": r["status"],
                "Why Failed": r["reason"] or "-",
                "Failed Params": "; ".join(
                    f"{d['param']}: expected '{d['expected']}', got '{d['actual'] or '(not sent)'}'"
                    for d in failed) or "-",
                "All Events Fired": ", ".join(r.get("fired_events", [])) or "-",
                "Matched Element": r.get("matched_element", ""),
                "Click Method": r.get("click_method", ""),
                "Click Verified": "Yes" if r.get("click_verified") else "No",
            })
        pd.DataFrame(flat_rows).to_excel('sdr_results.xlsx', index=False)
        # pandas cannot place pictures, so the two evidence columns are added
        # in a second pass. flat_rows follows `results` in order, so row N of
        # the sheet is result N.
        try:
            import openpyxl
            _wb = openpyxl.load_workbook('sdr_results.xlsx')
            _ws = _wb.active
            _line = {id(r): i + 2 for i, r in enumerate(results)}
            _add_shot_columns(_ws, 1, results, row_of=lambda r: _line.get(id(r)))
            _wb.save('sdr_results.xlsx')
        except Exception as _e:
            sys.stdout.write("[SDR] Could not add evidence columns to the flat "
                             "report: %s" % _e + os.linesep)

        # 2) The operator's own SDR, filled in. This is the deliverable they
        #    actually hand over: the same sheet, same rows, with the QA column
        #    marked Pass/Fail and the reason written in the column beside it.
        try:
            write_sdr_verdicts(sdr_path, meta, results, 'sdr_filled.xlsx',
                               qa_column=(getattr(args, 'qa_column', '') or ''),
                               extras=sdr_out.get("not_in_sdr") or [])
            sys.stdout.write("[SDR] Wrote sdr_filled.xlsx (your SDR with QA + reason columns)\n")
        except Exception as e:
            sys.stdout.write(f"[SDR] [WARN] could not write filled SDR: {str(e)[:140]}\n")

        passed = sum(1 for r in results if r["status"] == "PASS")
        sys.stdout.write(f"[SDR] Done: {passed}/{len(results)} test cases passed\n")
        sys.stdout.flush()
        return

    input_file, output_file = "input_sites.xlsx", "validation_results.xlsx"
    if not os.path.exists(input_file): return
    df = pd.read_excel(input_file)
    url_col = next((col for col in df.columns if any(n in str(col).lower() for n in ['url', 'link', 'website', 'site', 'address'])), None)
    if not url_col: return
    urls = [("https://" + str(u).strip() if not str(u).strip().startswith("http") else str(u).strip()) for u in df[url_col] if pd.notna(u)]
    total = len(urls)

    all_results = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage'])
        for i in range(0, total, CONCURRENCY):
            all_results.extend(await run_batch(browser, urls[i:i + CONCURRENCY], i + 1, total, args.mode))
        await browser.close()

    # Persist rich data for the UI
    if args.mode == 'clicks':
        click_rich = [r.pop("_click_rich") for r in all_results if "_click_rich" in r]
        with open("click_results.json", "w", encoding="utf-8") as f:
            json.dump({"generated": datetime.datetime.now().isoformat(),
                       "results": click_rich}, f, indent=2)
    else:
        for r in all_results:
            r.pop("_rich", None)
            r.pop("_click_rich", None)

    res_df = pd.DataFrame(all_results)
    if args.mode == 'tealium':
        cols = ['URL', 'Tealium_Loaded', 'Tealium_Account', 'Tealium_Profile', 'Tealium_Env', 'Adobe_Loaded', 'Adobe_ReportSuite', 'Adobe_PageView', 'Error']
        res_df[[c for c in cols if c in res_df.columns]].to_excel(output_file, index=False)
    elif args.mode == 'ga4':
        cols = ['URL', 'GTM_Loaded', 'GTM_ID', 'GA4_Fired', 'GA4_Measurement_ID', 'GA4_PageView', 'Error']
        res_df[[c for c in cols if c in res_df.columns]].to_excel(output_file, index=False)
    elif args.mode == 'clicks':
        cols = ['URL', 'Total_Elements', 'With_Tracking', 'Without_Tracking', 'Skipped', 'Error']
        summary_df = res_df[[c for c in cols if c in res_df.columns]]
        
        detail_rows = []
        for r in click_rich:
            page_url = r.get("URL", "")
            for el_res in r.get("elements", []):
                el = el_res.get("element", {})
                
                ga4_events = el_res.get("ga4_events", [])
                ga4_1_name = ga4_events[0]["event"] if len(ga4_events) > 0 else "--"
                ga4_1_id = ga4_events[0]["measurement_id"] if len(ga4_events) > 0 else "--"
                ga4_1_params = "; ".join([f"{k}={v}" for k, v in ga4_events[0].get("params", {}).items()]) if len(ga4_events) > 0 else "--"
                
                ga4_2_name = ga4_events[1]["event"] if len(ga4_events) > 1 else "--"
                ga4_2_id = ga4_events[1]["measurement_id"] if len(ga4_events) > 1 else "--"
                ga4_2_params = "; ".join([f"{k}={v}" for k, v in ga4_events[1].get("params", {}).items()]) if len(ga4_events) > 1 else "--"
                
                ga4_3_name = ga4_events[2]["event"] if len(ga4_events) > 2 else "--"
                ga4_3_id = ga4_events[2]["measurement_id"] if len(ga4_events) > 2 else "--"
                ga4_3_params = "; ".join([f"{k}={v}" for k, v in ga4_events[2].get("params", {}).items()]) if len(ga4_events) > 2 else "--"
                
                ga4_extra = ", ".join([e["event"] for e in ga4_events[3:]]) if len(ga4_events) > 3 else "--"
                
                adobe_calls = el_res.get("adobe_calls", [])
                adobe_evs = ", ".join([(c.get("link_name") or c.get("link_type")) for c in adobe_calls]) or "--"
                
                sdk_calls = el_res.get("adobe_websdk", [])
                if sdk_calls:
                    sdk_evs = ", ".join([w.get("event_type") for w in sdk_calls])
                    if adobe_evs == "--":
                        adobe_evs = sdk_evs
                    else:
                        adobe_evs += " | " + sdk_evs
                        
                other_evs = ", ".join([o.get("vendor", "") for o in el_res.get("other_analytics", [])]) or "--"
                
                # Raw request URLs for full QA visibility — capped so the
                # cell stays inside Excel's 32k character limit.
                net_urls = "\n".join(nr.get("url", "") for nr in el_res.get("network_requests", [])[:60])

                detail_rows.append({
                    "Page URL": page_url,
                    "Element Tag": el.get("tag", ""),
                    "Element Text": el.get("text", ""),
                    "Element ID": el.get("id", ""),
                    "Element Href": el.get("href", ""),
                    "Element Selector": el.get("selector", ""),
                    "Zone": el.get("zone", ""),
                    "Frame": el.get("frame_url", ""),
                    "Is Download": "Yes" if el.get("is_download") else "No",
                    "Click Status": "Skipped" if el_res.get("skipped") else (el_res.get("click_method") or "clicked"),
                    "Click Verified": ("--" if el_res.get("skipped")
                                       else ("Yes" if el_res.get("click_verified") else "No")),
                    "Skip Reason": el_res.get("skip_reason", ""),
                    "Blocked By": el_res.get("blocked_by", ""),
                    "Tracking Detected": "Yes" if el_res.get("has_tracking") else "No",
                    "GA4 Event 1 Name": ga4_1_name,
                    "GA4 Event 1 ID": ga4_1_id,
                    "GA4 Event 1 Params": ga4_1_params,
                    "GA4 Event 2 Name": ga4_2_name,
                    "GA4 Event 2 ID": ga4_2_id,
                    "GA4 Event 2 Params": ga4_2_params,
                    "GA4 Event 3 Name": ga4_3_name,
                    "GA4 Event 3 ID": ga4_3_id,
                    "GA4 Event 3 Params": ga4_3_params,
                    "GA4 Extra Events": ga4_extra,
                    "Adobe Calls Fired": adobe_evs,
                    "Other Analytics Fired": other_evs,
                    "Analytics Requests": el_res.get("total_requests", 0),
                    "All Network Requests": el_res.get("network_count", 0),
                    "Network Request URLs": net_urls,
                })
        detail_df = pd.DataFrame(detail_rows)
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            detail_df.to_excel(writer, sheet_name='Clicked_Elements_Detail', index=False)
    else:
        res_df.to_excel(output_file, index=False)


if __name__ == "__main__":
    asyncio.run(main())
